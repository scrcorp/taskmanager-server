"""Integration tests — dashboard export Overtime 시트 주 단위 판정 (P0-4).

GET /api/v1/console/dashboard/export 의 Overtime 시트가:
    - 범위 전체 합계가 아니라 user × 주(일요일 시작) 단위 row 로 나오는지
    - net(C1) 시간으로 주간 한도와 비교하는지
    - 한도가 매장별 LaborLawSetting 로 resolve 되는지
+ GET /api/v1/console/dashboard/overtime-summary 의 net + 매장별 한도 판정.

주의: 테스트 데이터는 고정 과거 2주(2026-05-03 ~ 05-16)에 생성하고,
테스트 유저의 해당 범위 attendance 를 전후로 purge 해 dev DB 잔재와 격리한다.
"""

from __future__ import annotations

import re
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from urllib.parse import unquote
from uuid import UUID

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import delete

from app.database import async_session
from app.models.attendance import Attendance
from app.models.attendance_break import BREAK_TYPE_UNPAID_MEAL, AttendanceBreak
from app.models.organization import LaborLawSetting


pytestmark = pytest.mark.asyncio


# 고정 과거 2주 — week1: 05-03(일)~05-09(토), week2: 05-10(일)~05-16(토)
WEEK1_START = date(2026, 5, 3)
WEEK1_END = date(2026, 5, 9)
WEEK2_START = date(2026, 5, 10)
WEEK2_END = date(2026, 5, 16)


async def _purge_range_attendance(user_ids: list[UUID]) -> None:
    """테스트 유저의 대상 범위 attendance 제거 (매장 무관 — dev DB 잔재 격리)."""
    async with async_session() as db:
        await db.execute(
            delete(Attendance).where(
                Attendance.user_id.in_(user_ids),
                Attendance.work_date >= WEEK1_START,
                Attendance.work_date <= WEEK2_END,
            )
        )
        await db.commit()


async def _delete_labor_settings(store_ids: list[UUID]) -> None:
    async with async_session() as db:
        await db.execute(
            delete(LaborLawSetting).where(LaborLawSetting.store_id.in_(store_ids))
        )
        await db.commit()


@pytest_asyncio.fixture
async def clean_export_range(test_users: dict, test_store_id: UUID, second_store_id: UUID):
    """대상 2주 attendance + 테스트 매장 LaborLawSetting 전후 정리."""
    user_ids = [info["id"] for info in test_users.values()]
    store_ids = [test_store_id, second_store_id]
    await _purge_range_attendance(user_ids)
    await _delete_labor_settings(store_ids)
    yield
    await _purge_range_attendance(user_ids)
    await _delete_labor_settings(store_ids)


async def _create_attendance(
    *,
    org_id: UUID,
    store_id: UUID,
    user_id: UUID,
    work_date: date,
    total_work_minutes: int,
    total_break_minutes: int = 0,
    breaks: list[tuple[str, int]] | None = None,
) -> UUID:
    """clocked_out attendance + break 세션 직접 생성 (과거 데이터 시뮬레이션)."""
    async with async_session() as db:
        att = Attendance(
            organization_id=org_id,
            store_id=store_id,
            user_id=user_id,
            work_date=work_date,
            status="clocked_out",
            total_work_minutes=total_work_minutes,
            total_break_minutes=total_break_minutes,
        )
        db.add(att)
        await db.flush()
        base = datetime.combine(work_date, time(12, 0), tzinfo=timezone.utc)
        offset = 0
        for break_type, duration in breaks or []:
            db.add(
                AttendanceBreak(
                    attendance_id=att.id,
                    started_at=base + timedelta(minutes=offset),
                    ended_at=base + timedelta(minutes=offset + duration),
                    break_type=break_type,
                    duration_minutes=duration,
                )
            )
            offset += duration + 5
        await db.commit()
        return att.id


def _overtime_rows_for(content: bytes, user_name: str) -> list[dict]:
    """응답 xlsx 바이트에서 Overtime 시트의 해당 유저 row 를 dict 로 추출."""
    wb = load_workbook(BytesIO(content), read_only=True)
    assert "Overtime" in wb.sheetnames
    ws = wb["Overtime"]
    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    assert headers == [
        "User", "Week Start", "Week End", "Total Hours", "Max Weekly", "Overtime Hours",
    ]
    rows = [dict(zip(headers, r)) for r in rows_iter if r[0] == user_name]
    wb.close()
    return rows


# ── P0-4: export Overtime 시트 — user × 주 단위 row ────────────────


async def test_export_overtime_sheet_splits_range_into_weeks(
    async_client: AsyncClient,
    admin_headers: dict,
    clean_export_range,
    seed_organization: dict,
    test_users: dict,
    test_store_id: UUID,
) -> None:
    """2주 범위: 45h net 주만 초과 5.0, 20h 주는 0.0 — 주별 row 2개.

    기존 버그였다면 단일 row 로 65h vs 40h = 25h 초과가 나왔을 케이스.
    범위 경계도 검증: date_from(수)~date_to(화)지만 경계 주 전체가 집계에 포함.
    """
    org_id = seed_organization["id"]
    staff = test_users["teststaff"]

    # week1: 일요일 포함 5일 × (gross 550, unpaid 10 → net 540) = 2700분 = 45h
    for i in range(5):
        await _create_attendance(
            org_id=org_id, store_id=test_store_id, user_id=staff["id"],
            work_date=WEEK1_START + timedelta(days=i), total_work_minutes=550,
            total_break_minutes=10, breaks=[(BREAK_TYPE_UNPAID_MEAL, 10)],
        )
    # week2: 3일 × 400분 = 1200분 = 20h
    for i in range(3):
        await _create_attendance(
            org_id=org_id, store_id=test_store_id, user_id=staff["id"],
            work_date=WEEK2_START + timedelta(days=i), total_work_minutes=400,
        )

    resp = await async_client.get(
        "/api/v1/console/dashboard/export",
        headers=admin_headers,
        params={
            # 수요일 ~ 다음 주 화요일 — 경계 주가 통째로 포함되어야 주간 합이 안 잘림
            "date_from": str(WEEK1_START + timedelta(days=3)),
            "date_to": str(WEEK2_START + timedelta(days=2)),
            "store_id": str(test_store_id),
        },
    )
    assert resp.status_code == 200, resp.text

    rows = _overtime_rows_for(resp.content, staff["full_name"])
    assert len(rows) == 2, rows

    week1, week2 = sorted(rows, key=lambda r: r["Week Start"])
    assert week1["Week Start"] == str(WEEK1_START)
    assert week1["Week End"] == str(WEEK1_END)
    assert week1["Total Hours"] == 45.0  # net (gross 였다면 45.8h)
    assert week1["Max Weekly"] == 40
    assert week1["Overtime Hours"] == 5.0

    assert week2["Week Start"] == str(WEEK2_START)
    assert week2["Week End"] == str(WEEK2_END)
    assert week2["Total Hours"] == 20.0
    assert week2["Overtime Hours"] == 0.0


async def test_export_overtime_sheet_uses_store_threshold(
    async_client: AsyncClient,
    admin_headers: dict,
    clean_export_range,
    seed_organization: dict,
    test_users: dict,
    test_store_id: UUID,
) -> None:
    """매장 한도 30h 설정 시 35h net → overtime 5.0 (org 임의 row 아님)."""
    org_id = seed_organization["id"]
    sv = test_users["testsv"]

    async with async_session() as db:
        db.add(LaborLawSetting(
            organization_id=org_id, store_id=test_store_id,
            federal_max_weekly=40, state_max_weekly=None, store_max_weekly=30,
        ))
        await db.commit()

    for i in range(5):
        # 일 420분 → 주간 2100분 = 35h
        await _create_attendance(
            org_id=org_id, store_id=test_store_id, user_id=sv["id"],
            work_date=WEEK1_START + timedelta(days=i), total_work_minutes=420,
        )

    resp = await async_client.get(
        "/api/v1/console/dashboard/export",
        headers=admin_headers,
        params={
            "date_from": str(WEEK1_START),
            "date_to": str(WEEK1_END),
            "store_id": str(test_store_id),
        },
    )
    assert resp.status_code == 200, resp.text

    rows = _overtime_rows_for(resp.content, sv["full_name"])
    assert len(rows) == 1, rows
    assert rows[0]["Max Weekly"] == 30
    assert rows[0]["Total Hours"] == 35.0
    assert rows[0]["Overtime Hours"] == 5.0


# ── overtime-summary — net + 매장별 한도 판정 ──────────────────────


async def test_overtime_summary_uses_net_and_store_threshold(
    async_client: AsyncClient,
    admin_headers: dict,
    clean_export_range,
    seed_organization: dict,
    test_users: dict,
    test_store_id: UUID,
) -> None:
    """gross 41h / net 39h 유저는 미집계, net 41h 유저만 초과 1.0h."""
    org_id = seed_organization["id"]
    sv = test_users["testsv"]    # gross 41h, net 39h — 집계되면 안 됨
    gm = test_users["testgm"]    # net 41h — overtime_users 에 포함

    for i in range(6):
        await _create_attendance(
            org_id=org_id, store_id=test_store_id, user_id=sv["id"],
            work_date=WEEK1_START + timedelta(days=i), total_work_minutes=410,
            total_break_minutes=20, breaks=[(BREAK_TYPE_UNPAID_MEAL, 20)],
        )
        await _create_attendance(
            org_id=org_id, store_id=test_store_id, user_id=gm["id"],
            work_date=WEEK1_START + timedelta(days=i), total_work_minutes=410,
        )

    resp = await async_client.get(
        "/api/v1/console/dashboard/overtime-summary",
        headers=admin_headers,
        params={
            "week_date": str(WEEK1_START + timedelta(days=3)),
            "store_id": str(test_store_id),
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()

    assert body["week_start"] == str(WEEK1_START)
    assert body["week_end"] == str(WEEK1_END)
    assert body["max_weekly_hours"] == 40
    assert body["total_users_with_attendance"] == 2
    # gross 판정이었다면 2명(41h, 41h) — net 판정이라 gm 1명만
    assert body["overtime_users"] == 1
    assert body["total_overtime_hours"] == 1.0


# ---------------------------------------------------------------------------
# 파일명 — 무엇을/어디/언제 받은 파일인지 이름만으로 구분되어야 한다
# ---------------------------------------------------------------------------


async def test_export_filename_carries_store_and_range(
    async_client: AsyncClient,
    admin_headers: dict,
    test_store_id: UUID,
) -> None:
    """매장+기간 필터 → 파일명에 매장명·기간·생성시각. (예전엔 항상 dashboard_export.xlsx)"""
    resp = await async_client.get(
        "/api/v1/console/dashboard/export",
        headers=admin_headers,
        params={
            "date_from": str(WEEK1_START),
            "date_to": str(WEEK1_END),
            "store_id": str(test_store_id),
        },
    )
    assert resp.status_code == 200, resp.text
    dispo = unquote(resp.headers["content-disposition"])
    assert re.search(r"Dashboard_.+_\d{8}-\d{4}_\d{8}-\d{4}Z\.xlsx", dispo), dispo
    assert "dashboard_export.xlsx" not in dispo


async def test_export_filename_without_filters_says_all_stores(
    async_client: AsyncClient, admin_headers: dict
) -> None:
    """필터 없이 받아도 생성시각은 붙는다 — 두 번 받은 파일이 구분된다."""
    resp = await async_client.get(
        "/api/v1/console/dashboard/export", headers=admin_headers
    )
    assert resp.status_code == 200, resp.text
    dispo = resp.headers["content-disposition"]
    assert re.search(r"Dashboard_AllStores_\d{8}-\d{4}Z\.xlsx", dispo), dispo
