"""Integration — 콘솔 Payroll API (Phase 3: periods/preview/confirm/entries/events).

대상: app/api/console/payroll.py + app/services/payroll_confirm_service.py
    - periods 목록 auto-ensure / 범위 검증
    - preview 행+요약, confirmed 기간 preview 409
    - 마감 게이트 ①~④ + 멀티스토어 정합 — 게이트별 개별 실패 payload
    - all-green confirm: entries 동결(스냅샷·breakdown==스칼라), events 스탬프
    - double-confirm 409 / 권한 거부(GM: payroll:*, staff: schedules:update)
    - events 목록 필터 + 귀책 태그 플로우 (invalid/frozen 400)

스펙: docs/99_inbox/2026-08-03 payroll-v1-스키마-스펙.md §4/§5/§6 + 설계방향 L1~L6.
"""

from __future__ import annotations

import uuid as uuid_mod
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from typing import AsyncIterator
from uuid import UUID

import pytest_asyncio
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.core.payroll_rules import (
    EVENT_KIND_DAILY_OT,
    EVENT_KIND_MEAL_PENALTY,
    EVENT_KIND_REST_PENALTY,
)
from app.database import async_session
from app.models.attendance import Attendance
from app.models.org_member import OrgMember, OrgMemberStore
from app.models.organization import Store
from app.models.payroll import PayPeriod, PayrollEntry, PayrollEvent
from app.models.rate import HourlyRateHistory
from app.models.schedule import Schedule
from app.models.tip import TipEntry, TipPeriod
from app.models.user import User
from app.schemas.payroll import (
    CALC_VERSION,
    CODE_ALREADY_CONFIRMED,
    CODE_CLOSE_GATES_FAILED,
    DayDetail,
    EntryBreakdown,
    PayrollPreviewRow,
    RateSegment,
    GATE_MULTI_STORE_WEEK,
    VALIDATION_BELOW_MINIMUM_WAGE,
    VALIDATION_OPEN_SHIFT,
    VALIDATION_OVERLAPPING_ATTENDANCE,
    VALIDATION_RATE_MISSING,
    VALIDATION_TIP_PERIOD_NOT_CONFIRMED,
    VALIDATION_UNCONFIRMED_AUTO_CLOCKOUT,
)
from app.services.payroll_export_service import (
    DRAFT_BANNER,
    EXPORT_COLUMNS,
    EXPORT_SHEET_TITLE,
    RATE_CHANGES_COLUMNS,
    RATE_CHANGES_SHEET_TITLE,
    WARNINGS_SHEET_TITLE,
    preview_export_row,
)
from app.services.payroll_period_service import payroll_period_service

_BASE = "/api/v1/console/payroll"

# 기준 기간 — 2026-07-01 ~ 07-15 (과거 기간이라 auto-ensure 대상)
_JUL_START = date(2026, 7, 1)
_JUL_END = date(2026, 7, 15)
_JUL_MID = date(2026, 7, 10)
_SUN = date(2026, 7, 5)
_MON = date(2026, 7, 6)
_TUE = date(2026, 7, 7)
_WED = date(2026, 7, 8)
_THU = date(2026, 7, 9)
_FRI = date(2026, 7, 10)

_RATE = Decimal("20.00")  # store default 시급


# ---------------------------------------------------------------------------
# 픽스처 — throwaway store 2개 + 기본 user/member (종료 시 전부 정리)
# ---------------------------------------------------------------------------


@pytest_asyncio.fixture
async def api_ctx(
    seed_organization: dict, seed_roles: dict[str, UUID]
) -> AsyncIterator[dict]:
    """매장 A/B + 기본 user (empid A=7, B=99 — 스냅샷 store-scope 검증용)."""
    org_id: UUID = seed_organization["id"]
    suffix = uuid_mod.uuid4().hex[:8]
    async with async_session() as db:
        store_a = Store(
            organization_id=org_id,
            name=f"__payroll_api_store_A_{suffix}",
            timezone="UTC",
            day_start_time={"all": "00:00"},
            default_hourly_rate=_RATE,
        )
        store_b = Store(
            organization_id=org_id,
            name=f"__payroll_api_store_B_{suffix}",
            timezone="UTC",
            day_start_time={"all": "00:00"},
            default_hourly_rate=_RATE,
        )
        db.add_all([store_a, store_b])
        await db.commit()
        await db.refresh(store_a)
        await db.refresh(store_b)

    ctx = {
        "org_id": org_id,
        "store_id": store_a.id,
        "store_b_id": store_b.id,
        "role_id": seed_roles["staff"],
        "user_ids": [],
        "member_ids": [],
    }
    crewid = 900_000 + int(uuid_mod.uuid4().hex[:4], 16)
    main = await _mk_user(ctx, "Payroll Main", crewid=crewid, empid=7, empid_b=99)
    ctx["user_id"] = main["user_id"]
    ctx["member_id"] = main["member_id"]
    ctx["crewid"] = crewid

    yield ctx

    async with async_session() as db:
        store_ids = [ctx["store_id"], ctx["store_b_id"]]
        await db.execute(
            delete(PayrollEvent).where(PayrollEvent.store_id.in_(store_ids))
        )
        await db.execute(
            delete(PayrollEntry).where(PayrollEntry.store_id.in_(store_ids))
        )
        await db.execute(delete(PayPeriod).where(PayPeriod.store_id.in_(store_ids)))
        await db.execute(delete(TipEntry).where(TipEntry.store_id.in_(store_ids)))
        await db.execute(delete(TipPeriod).where(TipPeriod.store_id.in_(store_ids)))
        await db.execute(delete(Attendance).where(Attendance.store_id.in_(store_ids)))
        await db.execute(delete(Schedule).where(Schedule.store_id.in_(store_ids)))
        await db.execute(
            delete(HourlyRateHistory).where(
                HourlyRateHistory.org_member_id.in_(ctx["member_ids"])
            )
        )
        await db.execute(delete(OrgMember).where(OrgMember.id.in_(ctx["member_ids"])))
        await db.execute(delete(User).where(User.id.in_(ctx["user_ids"])))
        await db.execute(delete(Store).where(Store.id.in_(store_ids)))
        await db.commit()


# ---------------------------------------------------------------------------
# 헬퍼
# ---------------------------------------------------------------------------


async def _mk_user(
    ctx: dict,
    full_name: str,
    *,
    member_rate: Decimal | None = None,
    crewid: int | None = None,
    empid: int | None = None,
    empid_b: int | None = None,
) -> dict:
    """user + org_member (+ A/B 매장 empid) 생성. ctx 정리 목록에 등록."""
    suffix = uuid_mod.uuid4().hex[:8]
    async with async_session() as db:
        user = User(
            organization_id=ctx["org_id"],
            role_id=ctx["role_id"],
            username=f"__payroll_api_{suffix}",
            full_name=full_name,
            password_hash="x",
            is_active=True,
        )
        db.add(user)
        await db.commit()
        await db.refresh(user)
        member = OrgMember(
            user_id=user.id,
            organization_id=ctx["org_id"],
            role_id=ctx["role_id"],
            hourly_rate=member_rate,
            crewid=crewid,
        )
        db.add(member)
        await db.commit()
        await db.refresh(member)
        if empid is not None:
            db.add(
                OrgMemberStore(
                    org_member_id=member.id, store_id=ctx["store_id"], empid=empid
                )
            )
        if empid_b is not None:
            db.add(
                OrgMemberStore(
                    org_member_id=member.id, store_id=ctx["store_b_id"], empid=empid_b
                )
            )
        await db.commit()
    ctx["user_ids"].append(user.id)
    ctx["member_ids"].append(member.id)
    return {"user_id": user.id, "member_id": member.id}


async def _mk_attendance(
    ctx: dict,
    *,
    work_date: date,
    total_work_minutes: int | None,
    user_id: UUID | None = None,
    store_id: UUID | None = None,
    status: str = "clocked_out",
    clock_in: datetime | None = None,
    clock_out: datetime | None = None,
    anomalies: list[str] | None = None,
    schedule_id: UUID | None = None,
) -> UUID:
    async with async_session() as db:
        att = Attendance(
            organization_id=ctx["org_id"],
            store_id=store_id or ctx["store_id"],
            user_id=user_id or ctx["user_id"],
            schedule_id=schedule_id,
            work_date=work_date,
            status=status,
            total_work_minutes=total_work_minutes,
            clock_in=clock_in,
            clock_out=clock_out,
            anomalies=anomalies,
        )
        db.add(att)
        await db.commit()
        await db.refresh(att)
        return att.id


async def _mk_schedule(ctx: dict, *, work_date: date, start: time, end: time) -> UUID:
    """confirmed 스케줄 1건 → id.

    같은 날 attendance 를 둘 이상 만들려면 필요하다 — `uq_attendance_walkin` 이
    `schedule_id IS NULL` 인 행에 대해 (user, work_date) 를 하루 1건으로 막는다.
    """
    async with async_session() as db:
        sched = Schedule(
            organization_id=ctx["org_id"],
            user_id=ctx["user_id"],
            store_id=ctx["store_id"],
            operating_day=work_date,
            start_at=datetime.combine(work_date, start),
            end_at=datetime.combine(work_date, end),
            status="confirmed",
        )
        db.add(sched)
        await db.commit()
        await db.refresh(sched)
        return sched.id


async def _mk_tip_period(
    ctx: dict, *, status: str = "confirmed",
    start: date = _JUL_START, end: date = _JUL_END,
) -> UUID:
    async with async_session() as db:
        tp = TipPeriod(
            store_id=ctx["store_id"], start_date=start, end_date=end, status=status
        )
        db.add(tp)
        await db.commit()
        await db.refresh(tp)
        return tp.id


async def _ensure_period(ctx: dict, date_in_period: date = _JUL_MID) -> str:
    """테스트 대상 기간 행 보장 (시스템 생성 경로) → period id 문자열."""
    async with async_session() as db:
        period = await payroll_period_service.ensure_period(
            db, store_id=ctx["store_id"], date_in_period=date_in_period
        )
        await db.commit()
        return str(period.id)


async def _login(username: str) -> dict[str, str]:
    """username → Authorization 헤더 (토큰 직접 mint — 로그인 플로우 의존 없음)."""
    from app.utils.jwt import create_access_token

    async with async_session() as db:
        user = (
            await db.execute(select(User).where(User.username == username))
        ).scalar_one()
        token = create_access_token(
            {"sub": str(user.id), "org": str(user.organization_id)}
        )
    return {"Authorization": f"Bearer {token}"}


def _gate_map(body: dict) -> dict[str, dict]:
    """409 detail.gates[] → {gate_code: gate_dict}."""
    assert body["detail"]["code"] == CODE_CLOSE_GATES_FAILED, body
    return {g["gate"]: g for g in body["detail"]["gates"]}


def _money(value) -> Decimal:
    return Decimal(str(value))


async def _confirm(
    client: AsyncClient, headers: dict, period_id: str
):
    return await client.post(f"{_BASE}/periods/{period_id}/confirm", headers=headers)


async def _events_in_db(ctx: dict) -> list[PayrollEvent]:
    async with async_session() as db:
        return list(
            (
                await db.execute(
                    select(PayrollEvent)
                    .where(PayrollEvent.store_id == ctx["store_id"])
                    .order_by(PayrollEvent.work_date, PayrollEvent.kind)
                )
            )
            .scalars()
            .all()
        )


# ---------------------------------------------------------------------------
# Periods 목록 — auto-ensure + 범위 검증
# ---------------------------------------------------------------------------


async def test_list_periods_auto_ensures_range(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """과거 범위 조회 시 반월 기간이 자동 생성되고, 재조회에 중복이 없다."""
    params = {
        "store_id": str(api_ctx["store_id"]),
        "start": "2026-07-01",
        "end": "2026-07-31",
    }
    resp = await async_client.get(
        f"{_BASE}/periods/", params=params, headers=admin_headers
    )
    assert resp.status_code == 200, resp.text
    periods = resp.json()
    assert [(p["start_date"], p["end_date"]) for p in periods] == [
        ("2026-07-01", "2026-07-15"),
        ("2026-07-16", "2026-07-31"),
    ]
    assert all(p["status"] == "open" for p in periods)
    assert all(p["tip_period_status"] is None for p in periods)  # 팁 기간 미생성

    # 멱등 — 재호출해도 같은 2건
    resp2 = await async_client.get(
        f"{_BASE}/periods/", params=params, headers=admin_headers
    )
    assert resp2.status_code == 200
    assert len(resp2.json()) == 2

    # 기본 범위(파라미터 없음) — 오늘이 속한 현재 기간 포함
    resp3 = await async_client.get(
        f"{_BASE}/periods/",
        params={"store_id": str(api_ctx["store_id"])},
        headers=admin_headers,
    )
    assert resp3.status_code == 200
    today = datetime.now(timezone.utc).date()
    assert any(
        p["start_date"] <= today.isoformat() <= p["end_date"] for p in resp3.json()
    )


async def test_list_periods_invalid_range_400(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    resp = await async_client.get(
        f"{_BASE}/periods/",
        params={
            "store_id": str(api_ctx["store_id"]),
            "start": "2026-07-31",
            "end": "2026-07-01",
        },
        headers=admin_headers,
    )
    assert resp.status_code == 400
    assert "Start date" in resp.json()["detail"]

    resp2 = await async_client.get(
        f"{_BASE}/periods/",
        params={
            "store_id": str(api_ctx["store_id"]),
            "start": "2024-01-01",
            "end": "2026-07-01",
        },
        headers=admin_headers,
    )
    assert resp2.status_code == 400
    assert "too large" in resp2.json()["detail"]


async def test_list_periods_shows_tip_status(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """대응 tip_period 가 있으면 status 가 함께 노출 (게이트 ④ 사전 표시)."""
    await _mk_tip_period(api_ctx, status="open")
    resp = await async_client.get(
        f"{_BASE}/periods/",
        params={
            "store_id": str(api_ctx["store_id"]),
            "start": "2026-07-01",
            "end": "2026-07-15",
        },
        headers=admin_headers,
    )
    assert resp.status_code == 200
    period = next(p for p in resp.json() if p["start_date"] == "2026-07-01")
    assert period["tip_period_status"] == "open"


# ---------------------------------------------------------------------------
# Preview
# ---------------------------------------------------------------------------


async def test_preview_rows_and_validation_summary(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """행(계산값) + validation 요약 — Mon 10h 근무자 + 열린 근무자."""
    await _mk_attendance(api_ctx, work_date=_MON, total_work_minutes=600)
    open_user = await _mk_user(api_ctx, "Open Shift User")
    await _mk_attendance(
        api_ctx, user_id=open_user["user_id"], work_date=_TUE,
        total_work_minutes=None, status="working",
        clock_in=datetime(2026, 7, 7, 9, 0, tzinfo=timezone.utc),
    )
    period_id = await _ensure_period(api_ctx)

    resp = await async_client.get(
        f"{_BASE}/periods/{period_id}/preview", headers=admin_headers
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["period"]["id"] == period_id
    assert body["period"]["status"] == "open"

    rows = {r["member_name"]: r for r in body["rows"]}
    main = rows["Payroll Main"]
    assert (main["regular_minutes"], main["ot_minutes"]) == (480, 120)
    assert _money(main["regular_pay"]) == Decimal("160.00")
    assert _money(main["ot_pay"]) == Decimal("60.00")
    assert _money(main["penalty_pay"]) == Decimal("40.00")  # meal + rest
    assert main["empid"] == 7
    assert main["crewid"] == api_ctx["crewid"]

    summary = {s["code"]: s["count"] for s in body["validations_summary"]}
    assert summary.get(VALIDATION_OPEN_SHIFT) == 1

    # 멱등 — 같은 입력 재호출 시 같은 행
    resp2 = await async_client.get(
        f"{_BASE}/periods/{period_id}/preview", headers=admin_headers
    )
    assert resp2.status_code == 200
    assert resp2.json()["rows"] == body["rows"]


async def test_preview_unknown_period_404(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    resp = await async_client.get(
        f"{_BASE}/periods/{uuid_mod.uuid4()}/preview", headers=admin_headers
    )
    assert resp.status_code == 404


# ---------------------------------------------------------------------------
# 마감 게이트 — 각 게이트 개별 실패
# ---------------------------------------------------------------------------


async def test_gate_unconfirmed_auto_clockout_then_resolve(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict,
    test_users: dict,
) -> None:
    """게이트 ① — 미확인 자동퇴근이 confirm 을 막고, 확인 후엔 통과한다."""
    att_id = await _mk_attendance(
        api_ctx, work_date=_MON, total_work_minutes=240,
        anomalies=["auto_clocked_out"],
    )
    await _mk_tip_period(api_ctx)  # 게이트 ④ 통과 상태
    period_id = await _ensure_period(api_ctx)

    resp = await _confirm(async_client, admin_headers, period_id)
    assert resp.status_code == 409, resp.text
    gates = _gate_map(resp.json())
    assert set(gates) == {VALIDATION_UNCONFIRMED_AUTO_CLOCKOUT}
    item = gates[VALIDATION_UNCONFIRMED_AUTO_CLOCKOUT]["items"][0]
    assert item["user_id"] == str(api_ctx["user_id"])
    assert item["member_name"] == "Payroll Main"
    assert item["dates"] == ["2026-07-06"]
    assert "2026-07-06" in item["message"]

    # 확인 처리 (L6 플로우의 영속 상태) → 게이트 해소
    async with async_session() as db:
        att = await db.get(Attendance, att_id)
        att.auto_clock_out_confirmed_at = datetime.now(timezone.utc)
        att.auto_clock_out_confirmed_by = test_users["testadmin"]["id"]
        await db.commit()

    resp2 = await _confirm(async_client, admin_headers, period_id)
    assert resp2.status_code == 200, resp2.text
    assert resp2.json()["period"]["status"] == "confirmed"


async def test_gate_open_shift(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """게이트 ② — clock_out 없는 열린 근무."""
    await _mk_attendance(
        api_ctx, work_date=_TUE, total_work_minutes=None, status="working",
        clock_in=datetime(2026, 7, 7, 9, 0, tzinfo=timezone.utc),
    )
    await _mk_tip_period(api_ctx)
    period_id = await _ensure_period(api_ctx)

    resp = await _confirm(async_client, admin_headers, period_id)
    assert resp.status_code == 409, resp.text
    gates = _gate_map(resp.json())
    assert set(gates) == {VALIDATION_OPEN_SHIFT}
    item = gates[VALIDATION_OPEN_SHIFT]["items"][0]
    assert item["dates"] == ["2026-07-07"]
    assert "clock-out" in gates[VALIDATION_OPEN_SHIFT]["message"]


async def test_gate_overlapping_attendance_then_resolve(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """게이트 ⑦ — 시간이 겹친 두 근무는 같은 시간을 두 번 지급한다 (D15).

    겹침 clock-in 을 허용한 대가로 반드시 있어야 하는 **최후 방어선**이다. 확인
    도장(escape hatch)은 두지 않는다 — "승인된 겹침" 은 정의상 이중 지급이라, 해소
    경로는 한쪽을 정정/취소하는 것뿐이다.

    판정은 anomaly 플래그가 아니라 **시간 구간**이다. 콘솔 clock-in 경로에는 겹침
    가드가 아예 없어 라벨 없이 겹침이 만들어질 수 있고, 플래그 기반이면 그 경로에서
    생긴 겹침을 영영 못 잡는다 — 그래서 아래 두 row 에도 anomaly 를 붙이지 않는다.
    """
    # 09:00–14:00 / 13:00–17:00 — **둘 다 닫혀 있다.** 이중 지급은 퇴근 후에 일어난다.
    await _mk_attendance(
        api_ctx, work_date=_MON, total_work_minutes=300,
        schedule_id=await _mk_schedule(
            api_ctx, work_date=_MON, start=time(9, 0), end=time(14, 0)
        ),
        clock_in=datetime(2026, 7, 6, 9, 0, tzinfo=timezone.utc),
        clock_out=datetime(2026, 7, 6, 14, 0, tzinfo=timezone.utc),
    )
    second = await _mk_attendance(
        api_ctx, work_date=_MON, total_work_minutes=240,
        schedule_id=await _mk_schedule(
            api_ctx, work_date=_MON, start=time(13, 0), end=time(17, 0)
        ),
        clock_in=datetime(2026, 7, 6, 13, 0, tzinfo=timezone.utc),
        clock_out=datetime(2026, 7, 6, 17, 0, tzinfo=timezone.utc),
    )
    await _mk_tip_period(api_ctx)
    period_id = await _ensure_period(api_ctx)

    resp = await _confirm(async_client, admin_headers, period_id)
    assert resp.status_code == 409, resp.text
    gates = _gate_map(resp.json())
    assert set(gates) == {VALIDATION_OVERLAPPING_ATTENDANCE}
    item = gates[VALIDATION_OVERLAPPING_ATTENDANCE]["items"][0]
    assert item["user_id"] == str(api_ctx["user_id"])
    assert item["dates"] == ["2026-07-06"]
    assert "2026-07-06" in item["message"]
    assert "twice" in gates[VALIDATION_OVERLAPPING_ATTENDANCE]["message"]

    # 한쪽을 정정해 겹침을 없애면 통과 — 게이트가 sticky 하지 않다는 확인.
    async with async_session() as db:
        att = await db.get(Attendance, second)
        att.clock_in = datetime(2026, 7, 6, 14, 0, tzinfo=timezone.utc)
        await db.commit()

    resp2 = await _confirm(async_client, admin_headers, period_id)
    assert resp2.status_code == 200, resp2.text
    assert resp2.json()["period"]["status"] == "confirmed"


async def test_back_to_back_shifts_do_not_trip_the_overlap_gate(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """맞닿은 근무(한쪽 종료 == 다른 쪽 시작)는 겹침이 아니다.

    여기서 오탐이 나면 **멀쩡한 급여 확정이 매번 막힌다** — 이중 지급만큼이나 나쁜
    고장이라 경계를 따로 못 박는다.
    """
    await _mk_attendance(
        api_ctx, work_date=_MON, total_work_minutes=240,
        schedule_id=await _mk_schedule(
            api_ctx, work_date=_MON, start=time(9, 0), end=time(13, 0)
        ),
        clock_in=datetime(2026, 7, 6, 9, 0, tzinfo=timezone.utc),
        clock_out=datetime(2026, 7, 6, 13, 0, tzinfo=timezone.utc),
    )
    await _mk_attendance(
        api_ctx, work_date=_MON, total_work_minutes=240,
        schedule_id=await _mk_schedule(
            api_ctx, work_date=_MON, start=time(13, 0), end=time(17, 0)
        ),
        clock_in=datetime(2026, 7, 6, 13, 0, tzinfo=timezone.utc),
        clock_out=datetime(2026, 7, 6, 17, 0, tzinfo=timezone.utc),
    )
    await _mk_tip_period(api_ctx)
    period_id = await _ensure_period(api_ctx)

    resp = await _confirm(async_client, admin_headers, period_id)
    assert resp.status_code == 200, resp.text


async def test_overlap_gate_sees_pairs_across_the_period_boundary(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """★ 기간 경계를 사이에 둔 겹침 쌍도 **양쪽 기간 모두** 막는다.

    짝 찾기 창이 `work_date BETWEEN start AND end` 뿐이면, 겹치는 두 row 의 work_date
    가 경계를 사이에 두고 갈릴 때 **어느 기간의 게이트도 그 쌍을 보지 못한다** —
    두 기간이 각각 confirm 되면서 겹친 시간이 두 번 지급된다. 이번 트랙이 연 D4
    (어제 영업일 shift 로 clock-in)와 야간조가 정확히 (어제, 오늘) 쌍을 만들기 때문에
    반월 급여의 15/16일 경계에서 그대로 성립했다.
    """
    # 07-15 22:00 ~ 07-16 06:00 (앞 기간 라벨) / 07-16 05:00 ~ 09:00 (뒤 기간 라벨)
    # → 05:00~06:00 한 시간이 겹친다.
    await _mk_attendance(
        api_ctx, work_date=date(2026, 7, 15), total_work_minutes=480,
        clock_in=datetime(2026, 7, 15, 22, 0, tzinfo=timezone.utc),
        clock_out=datetime(2026, 7, 16, 6, 0, tzinfo=timezone.utc),
    )
    second = await _mk_attendance(
        api_ctx, work_date=date(2026, 7, 16), total_work_minutes=240,
        clock_in=datetime(2026, 7, 16, 5, 0, tzinfo=timezone.utc),
        clock_out=datetime(2026, 7, 16, 9, 0, tzinfo=timezone.utc),
    )
    await _mk_tip_period(api_ctx)
    await _mk_tip_period(api_ctx, start=date(2026, 7, 16), end=date(2026, 7, 31))
    first_period = await _ensure_period(api_ctx)
    second_period = await _ensure_period(api_ctx, date_in_period=date(2026, 7, 20))
    assert first_period != second_period

    # 앞 기간 — 짝이 뒤 기간에 있어도 잡는다.
    resp = await _confirm(async_client, admin_headers, first_period)
    assert resp.status_code == 409, resp.text
    gates = _gate_map(resp.json())
    assert VALIDATION_OVERLAPPING_ATTENDANCE in gates
    # 보고는 **이 기간 안의 날짜만** — 기간 밖 날짜를 내밀면 매니저가 열 화면이 없다.
    assert gates[VALIDATION_OVERLAPPING_ATTENDANCE]["items"][0]["dates"] == [
        "2026-07-15"
    ]

    # 뒤 기간 — 반대 방향도 똑같이 잡는다.
    resp2 = await _confirm(async_client, admin_headers, second_period)
    assert resp2.status_code == 409, resp2.text
    gates2 = _gate_map(resp2.json())
    assert VALIDATION_OVERLAPPING_ATTENDANCE in gates2
    assert gates2[VALIDATION_OVERLAPPING_ATTENDANCE]["items"][0]["dates"] == [
        "2026-07-16"
    ]

    # 겹침을 없애면 양쪽 다 통과 — sticky 하지 않다.
    async with async_session() as db:
        att = await db.get(Attendance, second)
        att.clock_in = datetime(2026, 7, 16, 6, 0, tzinfo=timezone.utc)
        await db.commit()

    assert (
        await _confirm(async_client, admin_headers, first_period)
    ).status_code == 200
    assert (
        await _confirm(async_client, admin_headers, second_period)
    ).status_code == 200


async def test_open_shift_at_another_store_does_not_block_this_store(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """열린 근무는 겹침 게이트(⑦) 대상이 아니다 — 지급되지 않기 때문.

    `compute_net_work_minutes` 는 clock_out 이 없으면 None 을 돌려 **애초에 지급하지
    않는다.** 그런 row 를 겹침으로 세면 이중 지급을 막는 게 아니라 확정만 막는데,
    이 게이트는 org 전역이라 **다른 매장의 잊힌 clock-out 하나가 이 매장 급여를
    영구히 막고** 그 row 는 이 콘솔 스코프 밖이라 고칠 수도 없다(게이트 item 에는
    날짜만 실린다). 같은 매장 것은 게이트 ②(store 스코프)가 잡는다.
    """
    await _mk_attendance(
        api_ctx, work_date=_MON, total_work_minutes=300,
        schedule_id=await _mk_schedule(
            api_ctx, work_date=_MON, start=time(9, 0), end=time(14, 0)
        ),
        clock_in=datetime(2026, 7, 6, 9, 0, tzinfo=timezone.utc),
        clock_out=datetime(2026, 7, 6, 14, 0, tzinfo=timezone.utc),
    )
    # 매장 B 의 열린 근무 — 시간상 위 근무와 겹친다.
    await _mk_attendance(
        api_ctx, work_date=_MON, total_work_minutes=None, status="working",
        store_id=api_ctx["store_b_id"],
        clock_in=datetime(2026, 7, 6, 8, 0, tzinfo=timezone.utc),
    )
    await _mk_tip_period(api_ctx)
    period_id = await _ensure_period(api_ctx)

    resp = await _confirm(async_client, admin_headers, period_id)
    assert resp.status_code == 200, resp.text
    assert resp.json()["period"]["status"] == "confirmed"


async def test_gate_rate_missing(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """게이트 ③a — rate_at 결과 0 (org_members.hourly_rate=0, 계산 규칙 5)."""
    async with async_session() as db:
        member = await db.get(OrgMember, api_ctx["member_id"])
        member.hourly_rate = Decimal("0.00")
        await db.commit()
    await _mk_attendance(api_ctx, work_date=_MON, total_work_minutes=240)
    await _mk_tip_period(api_ctx)
    period_id = await _ensure_period(api_ctx)

    resp = await _confirm(async_client, admin_headers, period_id)
    assert resp.status_code == 409, resp.text
    gates = _gate_map(resp.json())
    assert set(gates) == {VALIDATION_RATE_MISSING}
    item = gates[VALIDATION_RATE_MISSING]["items"][0]
    assert item["user_id"] == str(api_ctx["user_id"])
    assert item["dates"] == ["2026-07-06"]
    assert "rate" in gates[VALIDATION_RATE_MISSING]["message"].lower()


async def test_gate_below_minimum_wage(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """게이트 ③b — 적용 rate 가 법정 최저 미만 (C8)."""
    async with async_session() as db:
        member = await db.get(OrgMember, api_ctx["member_id"])
        member.hourly_rate = Decimal("10.00")
        await db.commit()
    await _mk_attendance(api_ctx, work_date=_MON, total_work_minutes=240)
    await _mk_tip_period(api_ctx)
    period_id = await _ensure_period(api_ctx)

    resp = await _confirm(async_client, admin_headers, period_id)
    assert resp.status_code == 409, resp.text
    gates = _gate_map(resp.json())
    assert set(gates) == {VALIDATION_BELOW_MINIMUM_WAGE}
    assert gates[VALIDATION_BELOW_MINIMUM_WAGE]["items"][0]["dates"] == ["2026-07-06"]
    assert "minimum wage" in gates[VALIDATION_BELOW_MINIMUM_WAGE]["message"]


async def test_gate_tip_period_missing_then_open_then_confirmed(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """게이트 ④ — tip period 미생성/미확정이면 막히고, confirmed 면 통과."""
    await _mk_attendance(api_ctx, work_date=_MON, total_work_minutes=240)
    period_id = await _ensure_period(api_ctx)

    # 미생성
    resp = await _confirm(async_client, admin_headers, period_id)
    assert resp.status_code == 409, resp.text
    gates = _gate_map(resp.json())
    assert set(gates) == {VALIDATION_TIP_PERIOD_NOT_CONFIRMED}
    assert "not created yet" in gates[VALIDATION_TIP_PERIOD_NOT_CONFIRMED]["message"]

    # open 상태
    tip_id = await _mk_tip_period(api_ctx, status="open")
    resp2 = await _confirm(async_client, admin_headers, period_id)
    assert resp2.status_code == 409
    gates2 = _gate_map(resp2.json())
    assert "open" in gates2[VALIDATION_TIP_PERIOD_NOT_CONFIRMED]["message"]

    # confirmed → 통과
    async with async_session() as db:
        tp = await db.get(TipPeriod, tip_id)
        tp.status = "confirmed"
        await db.commit()
    resp3 = await _confirm(async_client, admin_headers, period_id)
    assert resp3.status_code == 200, resp3.text


async def test_gates_all_reported_together(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """게이트는 waterfall 없이 전부 평가 — 복수 실패가 한 응답에 담긴다."""
    async with async_session() as db:
        member = await db.get(OrgMember, api_ctx["member_id"])
        member.hourly_rate = Decimal("0.00")  # rate_missing
        await db.commit()
    await _mk_attendance(api_ctx, work_date=_MON, total_work_minutes=240)
    open_user = await _mk_user(api_ctx, "Open Shift User")
    await _mk_attendance(
        api_ctx, user_id=open_user["user_id"], work_date=_TUE,
        total_work_minutes=None, status="working",
        clock_in=datetime(2026, 7, 7, 9, 0, tzinfo=timezone.utc),
    )
    # tip period 없음 → 게이트 ④ 도 실패
    period_id = await _ensure_period(api_ctx)

    resp = await _confirm(async_client, admin_headers, period_id)
    assert resp.status_code == 409, resp.text
    gates = _gate_map(resp.json())
    assert set(gates) == {
        VALIDATION_RATE_MISSING,
        VALIDATION_OPEN_SHIFT,
        VALIDATION_TIP_PERIOD_NOT_CONFIRMED,
    }
    # 동결/확정은 일어나지 않았다
    async with async_session() as db:
        period = await db.get(PayPeriod, UUID(period_id))
        assert period.status == "open"
        entry_count = len(
            (
                await db.execute(
                    select(PayrollEntry).where(
                        PayrollEntry.pay_period_id == UUID(period_id)
                    )
                )
            ).scalars().all()
        )
        assert entry_count == 0


async def test_gate_multi_store_week_over_40h(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """멀티스토어 게이트 (계산 규칙 2) — 두 매장 합산 45h 주는 확정 거부."""
    for i in range(3):  # Mon-Wed 8h @ A = 24h
        await _mk_attendance(api_ctx, work_date=_MON + timedelta(days=i), total_work_minutes=480)
    await _mk_attendance(  # Thu 11h @ B
        api_ctx, work_date=_THU, total_work_minutes=660,
        store_id=api_ctx["store_b_id"],
    )
    await _mk_attendance(  # Fri 10h @ B
        api_ctx, work_date=_FRI, total_work_minutes=600,
        store_id=api_ctx["store_b_id"],
    )
    await _mk_tip_period(api_ctx)
    period_id = await _ensure_period(api_ctx)

    resp = await _confirm(async_client, admin_headers, period_id)
    assert resp.status_code == 409, resp.text
    gates = _gate_map(resp.json())
    assert GATE_MULTI_STORE_WEEK in gates
    item = gates[GATE_MULTI_STORE_WEEK]["items"][0]
    assert item["user_id"] == str(api_ctx["user_id"])
    assert item["dates"] == ["2026-07-05"]  # 해당 주 시작 일요일
    assert "45.0" in item["message"]
    assert str(_SUN) in item["message"]


async def test_multi_store_under_40h_confirms(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """두 매장이라도 합산이 임계 미만(겹침 없음)이면 정상 확정 — 지급은 A 근무분만."""
    await _mk_attendance(api_ctx, work_date=_MON, total_work_minutes=480)
    await _mk_attendance(api_ctx, work_date=_TUE, total_work_minutes=480)
    await _mk_attendance(
        api_ctx, work_date=_WED, total_work_minutes=480,
        store_id=api_ctx["store_b_id"],
    )
    await _mk_tip_period(api_ctx)
    period_id = await _ensure_period(api_ctx)

    resp = await _confirm(async_client, admin_headers, period_id)
    assert resp.status_code == 200, resp.text
    entries = resp.json()["entries"]
    assert len(entries) == 1
    assert entries[0]["regular_minutes"] == 960  # A 매장 2일만


# ---------------------------------------------------------------------------
# All-green confirm — 동결 검증
# ---------------------------------------------------------------------------


async def test_confirm_freezes_entries_events_and_period(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict,
    test_users: dict,
) -> None:
    """정상 확정: entries 스냅샷(올바른 매장 empid), events 스탬프, 기간 확정.

    손계산: Mon 10h(8h reg + 2h OT) + Tue 4h — reg 12h×$20=$240.00,
    OT 2h×1.5×$20=$60.00, penalty Mon meal+rest $40 + Tue rest $20 = $60.00,
    card tips $50.00 → gross $410.00.
    """
    await _mk_attendance(api_ctx, work_date=_MON, total_work_minutes=600)
    await _mk_attendance(api_ctx, work_date=_TUE, total_work_minutes=240)
    tip_id = await _mk_tip_period(api_ctx)
    async with async_session() as db:
        db.add(
            TipEntry(
                store_id=api_ctx["store_id"], employee_id=api_ctx["user_id"],
                date=_MON, card_tips=Decimal("50.00"),
            )
        )
        await db.commit()
    period_id = await _ensure_period(api_ctx)

    resp = await _confirm(async_client, admin_headers, period_id)
    assert resp.status_code == 200, resp.text
    body = resp.json()

    assert body["period"]["status"] == "confirmed"
    assert body["period"]["confirmed_by"] == str(test_users["testadmin"]["id"])
    assert body["period"]["confirmed_at"] is not None
    assert body["period"]["tip_period_status"] == "confirmed"
    assert body["events_frozen"] == 4  # Mon daily_ot/meal/rest + Tue rest

    assert len(body["entries"]) == 1
    entry = body["entries"][0]
    assert entry["member_name"] == "Payroll Main"
    assert entry["empid"] == 7  # A 매장 empid — B(99) 아님 (store-scope 스냅샷)
    assert entry["crewid"] == api_ctx["crewid"]
    assert entry["revision"] == 0
    assert entry["calc_version"] == CALC_VERSION
    assert (entry["regular_minutes"], entry["ot_minutes"], entry["dt_minutes"]) == (
        720, 120, 0,
    )
    assert _money(entry["regular_pay"]) == Decimal("240.00")
    assert _money(entry["ot_pay"]) == Decimal("60.00")
    assert _money(entry["penalty_pay"]) == Decimal("60.00")
    assert _money(entry["card_tips"]) == Decimal("50.00")
    assert _money(entry["gross_pay"]) == Decimal("410.00")
    assert entry["breakdown"]["tip_period_id"] == str(tip_id)
    assert len(entry["breakdown"]["days"]) == 2
    # 일별 금액도 함께 동결 — 월 10h(reg $160 + OT $60), 화 4h(reg $80)
    days_by_date = {d["work_date"]: d for d in entry["breakdown"]["days"]}
    mon_day = days_by_date[_MON.isoformat()]
    assert _money(mon_day["regular_amount"]) == Decimal("160.00")
    assert _money(mon_day["ot_amount"]) == Decimal("60.00")
    assert _money(mon_day["total_amount"]) == Decimal("220.00")
    assert _money(days_by_date[_TUE.isoformat()]["total_amount"]) == Decimal("80.00")

    # DB 동결 상태 — breakdown 합계 == 스칼라, org_member 링크, 이벤트 스탬프
    async with async_session() as db:
        db_entry = (
            await db.execute(
                select(PayrollEntry).where(
                    PayrollEntry.pay_period_id == UUID(period_id)
                )
            )
        ).scalar_one()
        assert db_entry.user_id == api_ctx["user_id"]
        assert db_entry.org_member_id == api_ctx["member_id"]
        assert db_entry.gross_pay == Decimal("410.00")
        seg_total = sum(
            Decimal(str(s["amount"])) for s in db_entry.breakdown["segments"]
        )
        assert seg_total == db_entry.regular_pay + db_entry.ot_pay + db_entry.dt_pay
        pen_total = sum(
            Decimal(str(p["amount"])) for p in db_entry.breakdown["penalties"]
        )
        assert pen_total == db_entry.penalty_pay

    events = await _events_in_db(api_ctx)
    assert len(events) == 4
    assert all(e.pay_period_id == UUID(period_id) for e in events)
    assert {e.kind for e in events} == {
        EVENT_KIND_DAILY_OT, EVENT_KIND_MEAL_PENALTY, EVENT_KIND_REST_PENALTY,
    }


async def test_double_confirm_409(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    await _mk_attendance(api_ctx, work_date=_MON, total_work_minutes=240)
    await _mk_tip_period(api_ctx)
    period_id = await _ensure_period(api_ctx)

    resp = await _confirm(async_client, admin_headers, period_id)
    assert resp.status_code == 200, resp.text

    resp2 = await _confirm(async_client, admin_headers, period_id)
    assert resp2.status_code == 409, resp2.text
    detail = resp2.json()["detail"]
    assert detail["code"] == CODE_ALREADY_CONFIRMED
    assert "already confirmed" in detail["message"]


async def test_preview_after_confirm_409(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """confirmed 기간의 preview 는 409 — 동결본(entries)이 원천."""
    await _mk_attendance(api_ctx, work_date=_MON, total_work_minutes=240)
    await _mk_tip_period(api_ctx)
    period_id = await _ensure_period(api_ctx)
    assert (await _confirm(async_client, admin_headers, period_id)).status_code == 200

    resp = await async_client.get(
        f"{_BASE}/periods/{period_id}/preview", headers=admin_headers
    )
    assert resp.status_code == 409
    assert "frozen entries" in resp.json()["detail"]


# ---------------------------------------------------------------------------
# Entries 조회
# ---------------------------------------------------------------------------


async def test_entries_empty_before_confirm_then_frozen_after(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    await _mk_attendance(api_ctx, work_date=_MON, total_work_minutes=240)
    await _mk_tip_period(api_ctx)
    period_id = await _ensure_period(api_ctx)

    resp = await async_client.get(
        f"{_BASE}/periods/{period_id}/entries", headers=admin_headers
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["entries"] == []
    assert resp.json()["period"]["status"] == "open"

    assert (await _confirm(async_client, admin_headers, period_id)).status_code == 200

    resp2 = await async_client.get(
        f"{_BASE}/periods/{period_id}/entries", headers=admin_headers
    )
    assert resp2.status_code == 200
    body = resp2.json()
    assert body["period"]["status"] == "confirmed"
    assert len(body["entries"]) == 1
    entry = body["entries"][0]
    assert entry["member_name"] == "Payroll Main"
    assert _money(entry["regular_pay"]) == Decimal("80.00")  # 4h × $20
    assert entry["breakdown"]["calc_version"] == CALC_VERSION


# ---------------------------------------------------------------------------
# 권한 — GM 은 payroll:* 없음 (Owner 전용), staff 는 schedules:update 없음
# ---------------------------------------------------------------------------


async def test_gm_denied_on_payroll_endpoints(
    async_client: AsyncClient, api_ctx: dict, test_users: dict, _clean_state
) -> None:
    """GM — payroll:read/confirm 미보유 → 목록/미리보기/확정/entries 전부 403.

    Owner 는 require_permission bypass 라 GM 으로 거부 경로를 검증한다.
    """
    headers = await _login("testgm")
    period_id = await _ensure_period(api_ctx)

    resp = await async_client.get(
        f"{_BASE}/periods/",
        params={"store_id": str(api_ctx["store_id"])},
        headers=headers,
    )
    assert resp.status_code == 403
    assert "payroll:read" in resp.json()["detail"]

    resp = await async_client.get(
        f"{_BASE}/periods/{period_id}/preview", headers=headers
    )
    assert resp.status_code == 403

    resp = await _confirm(async_client, headers, period_id)
    assert resp.status_code == 403
    assert "payroll:confirm" in resp.json()["detail"]

    resp = await async_client.get(
        f"{_BASE}/periods/{period_id}/entries", headers=headers
    )
    assert resp.status_code == 403


async def test_staff_denied_on_event_endpoints(
    async_client: AsyncClient, api_ctx: dict, _clean_state
) -> None:
    """staff — schedules:update 미보유 → events 목록/태깅 403."""
    headers = await _login("teststaff")

    resp = await async_client.get(
        f"{_BASE}/events/",
        params={
            "store_id": str(api_ctx["store_id"]),
            "start": "2026-07-01",
            "end": "2026-07-15",
        },
        headers=headers,
    )
    assert resp.status_code == 403
    assert "schedules:update" in resp.json()["detail"]

    resp = await async_client.post(
        f"{_BASE}/events/{uuid_mod.uuid4()}/tag",
        json={"attribution": "staff"},
        headers=headers,
    )
    assert resp.status_code == 403


async def test_gm_without_store_access_denied_on_events(
    async_client: AsyncClient, api_ctx: dict, _clean_state
) -> None:
    """GM — schedules:update 는 있지만 관리 매장이 아니면 store 접근 403."""
    headers = await _login("testgm")
    resp = await async_client.get(
        f"{_BASE}/events/",
        params={
            "store_id": str(api_ctx["store_id"]),
            "start": "2026-07-01",
            "end": "2026-07-15",
        },
        headers=headers,
    )
    assert resp.status_code == 403
    assert "store" in resp.json()["detail"].lower()


# ---------------------------------------------------------------------------
# Events — 목록/필터/태깅
# ---------------------------------------------------------------------------


async def _seed_events(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> list[dict]:
    """Mon 10h 근무 → preview 로 daily_ot + meal + rest 이벤트 생성 후 목록 반환."""
    await _mk_attendance(api_ctx, work_date=_MON, total_work_minutes=600)
    period_id = await _ensure_period(api_ctx)
    resp = await async_client.get(
        f"{_BASE}/periods/{period_id}/preview", headers=admin_headers
    )
    assert resp.status_code == 200, resp.text
    resp2 = await async_client.get(
        f"{_BASE}/events/",
        params={
            "store_id": str(api_ctx["store_id"]),
            "start": "2026-07-01",
            "end": "2026-07-15",
        },
        headers=admin_headers,
    )
    assert resp2.status_code == 200, resp2.text
    return resp2.json()


async def test_events_list_money_free_payload(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    events = await _seed_events(async_client, admin_headers, api_ctx)
    assert {e["kind"] for e in events} == {
        EVENT_KIND_DAILY_OT, EVENT_KIND_MEAL_PENALTY, EVENT_KIND_REST_PENALTY,
    }
    event = events[0]
    assert event["member_name"] == "Payroll Main"
    assert event["frozen"] is False
    assert event["attribution"] is None
    # 금액 필드가 전혀 없는 운영 페이로드
    money_keys = {k for k in event if "pay" in k or "amount" in k or "tip" in k}
    assert money_keys == {"pay_period_id"}


async def test_event_tag_flow_and_filters(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict,
    test_users: dict,
) -> None:
    events = await _seed_events(async_client, admin_headers, api_ctx)
    daily = next(e for e in events if e["kind"] == EVENT_KIND_DAILY_OT)

    # 태깅 — staff 귀책
    resp = await async_client.post(
        f"{_BASE}/events/{daily['id']}/tag",
        json={"attribution": "staff"},
        headers=admin_headers,
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["attribution"] == "staff"
    assert body["tagged_by"] == str(test_users["testadmin"]["id"])
    assert body["tagged_at"] is not None

    # 재태깅 — management 로 변경 가능 (frozen 전)
    resp2 = await async_client.post(
        f"{_BASE}/events/{daily['id']}/tag",
        json={"attribution": "management"},
        headers=admin_headers,
    )
    assert resp2.status_code == 200
    assert resp2.json()["attribution"] == "management"

    # attribution 필터
    params = {
        "store_id": str(api_ctx["store_id"]),
        "start": "2026-07-01",
        "end": "2026-07-15",
    }
    resp3 = await async_client.get(
        f"{_BASE}/events/", params={**params, "attribution": "management"},
        headers=admin_headers,
    )
    assert resp3.status_code == 200
    assert [e["id"] for e in resp3.json()] == [daily["id"]]

    resp4 = await async_client.get(
        f"{_BASE}/events/", params={**params, "attribution": "untagged"},
        headers=admin_headers,
    )
    assert resp4.status_code == 200
    assert daily["id"] not in [e["id"] for e in resp4.json()]
    assert len(resp4.json()) == 2  # meal + rest

    # 잘못된 필터 값
    resp5 = await async_client.get(
        f"{_BASE}/events/", params={**params, "attribution": "nobody"},
        headers=admin_headers,
    )
    assert resp5.status_code == 400
    assert "attribution" in resp5.json()["detail"]


async def test_event_tag_invalid_value_400(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    events = await _seed_events(async_client, admin_headers, api_ctx)
    resp = await async_client.post(
        f"{_BASE}/events/{events[0]['id']}/tag",
        json={"attribution": "somebody"},
        headers=admin_headers,
    )
    assert resp.status_code == 400
    assert "Invalid attribution" in resp.json()["detail"]


async def test_event_tag_unknown_404(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    resp = await async_client.post(
        f"{_BASE}/events/{uuid_mod.uuid4()}/tag",
        json={"attribution": "staff"},
        headers=admin_headers,
    )
    assert resp.status_code == 404


async def test_event_tag_frozen_400(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """confirm 으로 동결된 이벤트는 태깅 불가 (amendment 전용)."""
    await _mk_attendance(api_ctx, work_date=_MON, total_work_minutes=240)  # rest 1건
    await _mk_tip_period(api_ctx)
    period_id = await _ensure_period(api_ctx)
    assert (await _confirm(async_client, admin_headers, period_id)).status_code == 200

    events = await _events_in_db(api_ctx)
    assert events and all(e.pay_period_id is not None for e in events)

    resp = await async_client.post(
        f"{_BASE}/events/{events[0].id}/tag",
        json={"attribution": "staff"},
        headers=admin_headers,
    )
    assert resp.status_code == 400
    assert "confirmed pay period" in resp.json()["detail"]


async def test_events_include_voided_filter(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """조건 소멸로 voided 된 이벤트는 기본 목록에서 빠지고, 플래그로만 보인다."""
    att_id = await _mk_attendance(api_ctx, work_date=_MON, total_work_minutes=600)
    period_id = await _ensure_period(api_ctx)
    resp = await async_client.get(
        f"{_BASE}/periods/{period_id}/preview", headers=admin_headers
    )
    assert resp.status_code == 200

    # 10h → 6h 로 정정 — daily_ot 조건 소멸 (meal/rest 는 유지)
    async with async_session() as db:
        att = await db.get(Attendance, att_id)
        att.total_work_minutes = 360
        await db.commit()
    resp2 = await async_client.get(
        f"{_BASE}/periods/{period_id}/preview", headers=admin_headers
    )
    assert resp2.status_code == 200

    params = {
        "store_id": str(api_ctx["store_id"]),
        "start": "2026-07-01",
        "end": "2026-07-15",
    }
    resp3 = await async_client.get(
        f"{_BASE}/events/", params=params, headers=admin_headers
    )
    assert resp3.status_code == 200
    assert EVENT_KIND_DAILY_OT not in {e["kind"] for e in resp3.json()}

    resp4 = await async_client.get(
        f"{_BASE}/events/", params={**params, "include_voided": "true"},
        headers=admin_headers,
    )
    assert resp4.status_code == 200
    voided = [e for e in resp4.json() if e["kind"] == EVENT_KIND_DAILY_OT]
    assert len(voided) == 1
    assert voided[0]["voided_at"] is not None


# ---------------------------------------------------------------------------
# Export — 기간 xlsx (Phase 4, DUMMY 포맷 v1): 확정=동결본 / 미확정=DRAFT
# ---------------------------------------------------------------------------


async def _mk_frozen_entry(
    ctx: dict,
    period_id: str,
    *,
    member_name: str,
    empid: int | None,
    crewid: int | None,
    segments: list[RateSegment],
    regular_minutes: int,
    regular_pay: Decimal,
) -> None:
    """확정 기간에 동결 entry 를 직접 추가 — export 전용 케이스 조립용.

    (멀티 rate / ID 미매칭은 실계산 픽스처로 만들기 번거로워 동결 스냅샷을
    직접 넣는다 — export 는 frozen rows 만 읽으므로 충분한 재현.)
    """
    breakdown = EntryBreakdown(
        segments=segments,
        days=[DayDetail(work_date=_MON, regular_minutes=regular_minutes)],
    )
    async with async_session() as db:
        db.add(
            PayrollEntry(
                pay_period_id=UUID(period_id),
                organization_id=ctx["org_id"],
                store_id=ctx["store_id"],
                user_id=None,
                org_member_id=None,
                empid=empid,
                crewid=crewid,
                member_name=member_name,
                revision=0,
                regular_minutes=regular_minutes,
                regular_pay=regular_pay,
                gross_pay=regular_pay,
                calc_version=CALC_VERSION,
                breakdown=breakdown.model_dump(mode="json"),
            )
        )
        await db.commit()


async def _export(client: AsyncClient, headers: dict, period_id: str):
    return await client.get(f"{_BASE}/periods/{period_id}/export", headers=headers)


def _numeric_cells(cells: list) -> list:
    """셀 값 비교용 정규화 — 숫자는 Decimal 로 (openpyxl 은 float 로 돌려준다)."""
    return [
        _money(v) if isinstance(v, (int, float, Decimal)) else v for v in cells
    ]


async def test_export_unconfirmed_period_draft(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """open 기간 export 는 DRAFT 워크북 — 확인용이라 차단하지 않는다.

    배너 행 + payroll_draft_ 파일명으로 확정본과 구분되고, 값은 live preview
    와 같아야 한다. ID 없는 직원의 Warnings 시트 규칙도 동결본과 동일.
    """
    await _mk_attendance(api_ctx, work_date=_MON, total_work_minutes=600)
    await _mk_attendance(api_ctx, work_date=_TUE, total_work_minutes=240)
    await _mk_tip_period(api_ctx)
    async with async_session() as db:
        db.add(
            TipEntry(
                store_id=api_ctx["store_id"], employee_id=api_ctx["user_id"],
                date=_MON, card_tips=Decimal("50.00"),
            )
        )
        await db.commit()
    no_ids = await _mk_user(api_ctx, "ZZ No Ids")  # empid/crewid 둘 다 없음
    await _mk_attendance(
        api_ctx, work_date=_WED, total_work_minutes=300,
        user_id=no_ids["user_id"],
    )
    period_id = await _ensure_period(api_ctx)

    preview = await async_client.get(
        f"{_BASE}/periods/{period_id}/preview", headers=admin_headers
    )
    assert preview.status_code == 200, preview.text
    expected = [
        preview_export_row(PayrollPreviewRow.model_validate(row))
        for row in preview.json()["rows"]
    ]
    assert len(expected) == 2  # Payroll Main + ZZ No Ids

    resp = await _export(async_client, admin_headers, period_id)
    assert resp.status_code == 200, resp.text
    disposition = resp.headers["content-disposition"]
    assert "_DRAFT.xlsx" in disposition  # 저장된 파일만 봐도 미확정본
    assert "Payroll_" in disposition
    assert "2026-07-01~2026-07-15_DRAFT.xlsx" in disposition

    wb = load_workbook(BytesIO(resp.content))
    ws = wb[EXPORT_SHEET_TITLE]
    assert ws.cell(row=1, column=1).value == DRAFT_BANNER  # 배너가 헤더 위 1행
    assert [c.value for c in ws[2]] == EXPORT_COLUMNS
    rows = [[c.value for c in row] for row in ws.iter_rows(min_row=3)]
    assert [_numeric_cells(r) for r in rows] == [
        _numeric_cells(e) for e in expected
    ]

    # ID 미매칭 — draft 에서도 Warnings 시트 (동결본과 같은 규칙)
    assert wb.sheetnames == [EXPORT_SHEET_TITLE, WARNINGS_SHEET_TITLE]
    warn = wb[WARNINGS_SHEET_TITLE]
    warn_rows = [[c.value for c in row] for row in warn.iter_rows(min_row=2)]
    assert [r[0] for r in warn_rows] == ["ZZ No Ids"]
    assert "EMPID" in warn_rows[0][1]

    # export 는 읽기 동작 — 기간을 확정시키지 않는다
    async with async_session() as db:
        period = await db.get(PayPeriod, UUID(period_id))
        assert period.status == "open"


async def test_export_unknown_period_404(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    resp = await _export(async_client, admin_headers, str(uuid_mod.uuid4()))
    assert resp.status_code == 404


async def test_export_includes_rate_changes_sheet(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """기간 내 effective_date 인 시급 변경(메모 포함)이 Rate Changes 시트로 나간다.

    payroll 화면 인라인 rate 등록의 감사 흔적 — 없으면 시트 자체가 없어야 한다.
    """
    await _mk_attendance(api_ctx, work_date=_MON, total_work_minutes=600)
    await _mk_tip_period(api_ctx)
    period_id = await _ensure_period(api_ctx)

    # 변경 전 export — 시트 없음
    resp = await _export(async_client, admin_headers, period_id)
    assert resp.status_code == 200, resp.text
    assert RATE_CHANGES_SHEET_TITLE not in load_workbook(BytesIO(resp.content)).sheetnames

    # 기간 시작일 소급 시급 등록 (payroll 인라인 팝오버와 같은 호출)
    rc = await async_client.post(
        f"/api/v1/console/users/{api_ctx['user_id']}/rate-changes",
        headers=admin_headers,
        json={
            "new_rate": "21.50",
            "effective_date": "2026-07-01",
            "reason": "Set from payroll",
        },
    )
    assert rc.status_code == 200, rc.text

    resp = await _export(async_client, admin_headers, period_id)
    assert resp.status_code == 200, resp.text
    wb = load_workbook(BytesIO(resp.content))
    assert RATE_CHANGES_SHEET_TITLE in wb.sheetnames
    ws = wb[RATE_CHANGES_SHEET_TITLE]
    assert [c.value for c in ws[1]] == RATE_CHANGES_COLUMNS
    rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    assert len(rows) == 1
    name, empid, old_rate, new_rate, eff, memo, changed_by, changed_at = rows[0]
    assert name  # 본 시트와 같은 스냅샷 이름
    assert _money(new_rate) == Decimal("21.50")
    assert eff == "2026-07-01"
    assert memo == "Set from payroll"
    assert changed_by  # admin 이름
    assert changed_at  # UTC 타임스탬프

    # 기간 밖 effective_date 는 시트 대상 아님
    rc2 = await async_client.post(
        f"/api/v1/console/users/{api_ctx['user_id']}/rate-changes",
        headers=admin_headers,
        json={"new_rate": "22.00", "effective_date": "2026-08-01"},
    )
    assert rc2.status_code == 200, rc2.text
    resp = await _export(async_client, admin_headers, period_id)
    ws = load_workbook(BytesIO(resp.content))[RATE_CHANGES_SHEET_TITLE]
    assert ws.max_row == 2  # 여전히 7/1 변경 1건뿐


async def test_export_confirmed_period_workbook(
    async_client: AsyncClient, admin_headers: dict, api_ctx: dict
) -> None:
    """확정 기간 export — 헤더/행 값, 멀티 rate 셀, EMPID fallback, Warnings 시트.

    행 1 (실계산 동결): Payroll Main — Mon 10h + Tue 4h @ $20, tips $50
        → 12.00h reg / 2.00h OT, $240/$60/$60/$50/gross $410 (confirm 손계산).
    행 2 (조립): ZZ Multi Rate — rate 18/22 두 구간, empid 없음 → EMPID=crewid.
    행 3 (조립): ZZ Unmatched — empid/crewid 둘 다 없음 → 빈칸 + Warnings.
    """
    await _mk_attendance(api_ctx, work_date=_MON, total_work_minutes=600)
    await _mk_attendance(api_ctx, work_date=_TUE, total_work_minutes=240)
    tip_id = await _mk_tip_period(api_ctx)
    async with async_session() as db:
        db.add(
            TipEntry(
                store_id=api_ctx["store_id"], employee_id=api_ctx["user_id"],
                date=_MON, card_tips=Decimal("50.00"),
            )
        )
        await db.commit()
    period_id = await _ensure_period(api_ctx)
    assert (await _confirm(async_client, admin_headers, period_id)).status_code == 200
    del tip_id  # gross 검증은 confirm 테스트가 담당 — 여기선 export 형태만

    await _mk_frozen_entry(
        api_ctx, period_id,
        member_name="ZZ Multi Rate", empid=None, crewid=555,
        segments=[
            RateSegment(
                rate=Decimal("18.00"), regular_minutes=300, amount=Decimal("90.00")
            ),
            RateSegment(
                rate=Decimal("22.00"), regular_minutes=300, amount=Decimal("110.00")
            ),
        ],
        regular_minutes=600, regular_pay=Decimal("200.00"),
    )
    await _mk_frozen_entry(
        api_ctx, period_id,
        member_name="ZZ Unmatched", empid=None, crewid=None,
        segments=[
            RateSegment(
                rate=Decimal("20.00"), regular_minutes=240, amount=Decimal("80.00")
            )
        ],
        regular_minutes=240, regular_pay=Decimal("80.00"),
    )

    resp = await _export(async_client, admin_headers, period_id)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disposition = resp.headers["content-disposition"]
    assert "Payroll_" in disposition
    assert "2026-07-01~2026-07-15.xlsx" in disposition
    assert "_DRAFT" not in disposition  # 확정본에는 draft 표식 없음
    assert "filename*=UTF-8''" in disposition  # 유니코드 파라미터 항상 동반

    wb = load_workbook(BytesIO(resp.content))
    assert wb.sheetnames == [EXPORT_SHEET_TITLE, WARNINGS_SHEET_TITLE]
    ws = wb[EXPORT_SHEET_TITLE]
    assert [c.value for c in ws[1]] == EXPORT_COLUMNS

    rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    assert [r[2] for r in rows] == [  # member_name 오름차순
        "Payroll Main", "ZZ Multi Rate", "ZZ Unmatched",
    ]

    main, multi, unmatched = rows
    # 실계산 동결 행 — 올바른 매장 empid(7) + 손계산 금액/시간
    assert main[0] == 7
    assert main[1] == api_ctx["crewid"]
    assert (main[3], main[4], main[5]) == (12.0, 2.0, 0.0)  # 시간 (분/60, 2dp)
    assert main[6] == "20.00"
    assert [_money(v) for v in main[7:]] == [
        Decimal("240.00"), Decimal("60.00"), Decimal("0.00"),
        Decimal("60.00"), Decimal("50.00"), Decimal("410.00"),
    ]
    # 멀티 rate — 콤마 목록 + EMPID fallback(crewid)
    assert multi[0] == 555
    assert multi[1] == 555
    assert multi[6] == "18.00, 22.00"
    assert multi[3] == 10.0
    # ID 미매칭 — 빈칸 + Warnings 시트에 나열
    assert unmatched[0] is None
    assert unmatched[1] is None
    warn = wb[WARNINGS_SHEET_TITLE]
    warn_rows = [[c.value for c in row] for row in warn.iter_rows(min_row=2)]
    assert len(warn_rows) == 1
    assert warn_rows[0][0] == "ZZ Unmatched"
    assert "EMPID" in warn_rows[0][1]


async def test_export_gm_denied_403(
    async_client: AsyncClient, api_ctx: dict, _clean_state
) -> None:
    """GM — payroll:export 미보유 (Owner 전용) → 403."""
    headers = await _login("testgm")
    period_id = await _ensure_period(api_ctx)

    resp = await _export(async_client, headers, period_id)
    assert resp.status_code == 403
    assert "payroll:export" in resp.json()["detail"]
