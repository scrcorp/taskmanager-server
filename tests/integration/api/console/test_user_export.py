"""Integration tests — Staff xlsx export (GET /api/v1/console/users/export).

대상:
    - 200 + xlsx 파싱: 헤더 셋(CREWID / Store EMPID 포함) + 시급/급여 컬럼 부재
    - 행 = 사람 × 매장배정 (배정 없는 사람은 매장 칸 빈 1행)
    - 미가입(유령) → Account Status "Not signed up"
    - store_ids 필터 — 사람 자체가 배정 기준으로 포함/제외
    - Manager Y/N + Employment Status / Hire / Termination (org_members 소스)
    - 라우트 순서 회귀: /export 가 GET /{user_id} 뒤로 가면 "export" 가 UUID
      파싱에 걸려 422 — 모든 테스트의 200 assert 가 그 회귀를 잡는다.

주의: DB 를 다른 세션과 공유하므로 고유 username 으로 생성하고 teardown 에서
hard delete 한다. 시드 데이터(test_users / 테스트 매장)는 건드리지 않는다.

매장 배정은 API(PUT /{id}/stores)가 아니라 org_member_stores 직접 INSERT 로
시딩한다 — 같은 브랜치의 empid_changes 이력 트랙이 모델만 있고 테이블은
아직 없어서(마이그레이션 일괄 생성 대기) 배정 API 가 일시적으로 500 이다.
export 는 org_member_stores 를 읽기만 하므로 시딩 경로와 무관하게 검증된다.
마이그레이션이 붙은 뒤에도 이 테스트는 그대로 유효하다.
"""

from __future__ import annotations

import random
import uuid
from datetime import date
from io import BytesIO
from typing import AsyncIterator

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.database import async_session
from app.models.org_member import OrgMember, OrgMemberStore
from app.models.user import User

pytestmark = pytest.mark.asyncio

EXPORT_URL = "/api/v1/console/users/export"

# 서비스와 독립적으로 못박은 기대 헤더 — 컬럼이 바뀌면 여기서 잡힌다.
EXPECTED_HEADERS: list[str] = [
    "Username",
    "Name",
    "CREWID",
    "Role",
    "Account Status",
    "Employment Status",
    "Department",
    "Email",
    "Hire Date",
    "Termination Date",
    "Store",
    "Store EMPID",
    "Manager",
]

# cost 마스킹 정책 — 헤더에 절대 나오면 안 되는 단어 (권한 무관 전면 제외)
FORBIDDEN_HEADER_WORDS: list[str] = ["rate", "salary", "wage", "cost", "pay"]


# ---------------------------------------------------------------------------
# 헬퍼
# ---------------------------------------------------------------------------


def _blank(v: object) -> bool:
    """빈 셀 판정 — openpyxl 은 빈 문자열 셀을 None 으로 읽을 수 있다."""
    return v in ("", None)


def _load_sheet(content: bytes) -> tuple[list[str], list[dict]]:
    """응답 xlsx 바이트 → (헤더, 행 dict 목록)."""
    wb = load_workbook(BytesIO(content), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))
    rows = [dict(zip(headers, r)) for r in rows_iter]
    wb.close()
    return headers, rows


def _rows_for(rows: list[dict], username: str) -> list[dict]:
    return [r for r in rows if r["Username"] == username]


def _new_username(prefix: str = "exp") -> str:
    return f"{prefix}_{uuid.uuid4().hex[:8]}"


async def _create_user(
    client: AsyncClient,
    headers: dict,
    role_id,
    tracked: list[str],
    **extra,
) -> dict:
    """콘솔 API 로 user 생성 (org_member + crewid 동반 생성 경로). username 추적."""
    username = _new_username()
    tracked.append(username)
    resp = await client.post(
        "/api/v1/console/users",
        headers=headers,
        json={
            "username": username,
            "password": "test1234",
            "full_name": f"Export Test {username[-4:]}",
            "role_id": str(role_id),
            **extra,
        },
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


# empid 시딩용 — 실데이터(작은 번호)와 절대 안 겹치는 고유 대역 (store 내 unique 제약)
_EMPID_SEQ = iter(range(800000 + random.randint(0, 99999) * 100, 2_000_000_000))


async def _assign_stores(
    _client: AsyncClient,
    _headers: dict,
    user_id: str,
    assignments: list[dict],
) -> None:
    """org_member_stores 직접 시딩 — export 의 읽기 소스를 만든다.

    (모듈 docstring 참조 — 배정 API 는 empid_changes 마이그레이션 대기로 일시 500.
    export 검증에는 행의 존재만 필요하므로 직접 INSERT 로 시딩한다.)
    """
    async with async_session() as db:
        member_id = (
            await db.execute(
                select(OrgMember.id).where(OrgMember.user_id == uuid.UUID(user_id))
            )
        ).scalar_one()
        for a in assignments:
            db.add(
                OrgMemberStore(
                    org_member_id=member_id,
                    store_id=uuid.UUID(a["store_id"]),
                    is_manager=a["is_manager"],
                    is_work_assignment=a["is_work_assignment"],
                    empid=next(_EMPID_SEQ),
                )
            )
        await db.commit()


async def _export(
    client: AsyncClient, headers: dict, **params
) -> tuple[list[str], list[dict], object]:
    resp = await client.get(EXPORT_URL, headers=headers, params=params or None)
    assert resp.status_code == 200, resp.text
    headers_row, rows = _load_sheet(resp.content)
    return headers_row, rows, resp


@pytest_asyncio.fixture
async def cleanup_created_users() -> AsyncIterator[list[str]]:
    """테스트가 만든 user 를 username 으로 추적 → teardown 에서 hard delete.

    User 삭제가 org_members → org_member_stores 까지 CASCADE 로 정리한다.
    (이 테스트는 rate 변경을 안 하므로 hourly_rate_history 선삭제는 불필요.)
    """
    usernames: list[str] = []
    yield usernames
    if usernames:
        async with async_session() as db:
            await db.execute(delete(User).where(User.username.in_(usernames)))
            await db.commit()


# ---------------------------------------------------------------------------
# 헤더 / 파일 형식 (+ 라우트 순서 회귀)
# ---------------------------------------------------------------------------


async def test_export_headers_filename_and_no_cost_columns(
    async_client: AsyncClient, admin_headers: dict
) -> None:
    """200 xlsx + 고정 헤더 + 파일명 staff_{today} + 시급/급여 컬럼 부재.

    이 200 assert 자체가 라우트 순서 회귀 방어다 — /export 가 GET /{user_id}
    뒤로 밀리면 "export" 가 UUID 파싱에 걸려 422 가 난다.
    """
    resp = await async_client.get(EXPORT_URL, headers=admin_headers)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    dispo = resp.headers.get("content-disposition", "")
    assert f"staff_{date.today().isoformat()}.xlsx" in dispo

    headers_row, _rows = _load_sheet(resp.content)
    assert headers_row == EXPECTED_HEADERS

    # cost 마스킹 — 시급/급여류 컬럼은 권한 분기 없이 아예 없어야 한다
    for header in headers_row:
        lowered = header.lower()
        for word in FORBIDDEN_HEADER_WORDS:
            assert word not in lowered, f"cost column leaked: {header}"


async def test_export_invalid_store_ids_returns_400(
    async_client: AsyncClient, admin_headers: dict
) -> None:
    """store_ids 에 UUID 아닌 값 → 400 (raw ValueError 500 금지)."""
    resp = await async_client.get(
        EXPORT_URL, headers=admin_headers, params={"store_ids": "not-a-uuid"}
    )
    assert resp.status_code == 400, resp.text


# ---------------------------------------------------------------------------
# 행 = 사람 × 매장배정
# ---------------------------------------------------------------------------


async def test_export_rows_person_by_store_assignment(
    async_client: AsyncClient,
    admin_headers: dict,
    seed_roles: dict,
    test_store_id,
    second_store_id,
    cleanup_created_users: list[str],
) -> None:
    """배정 1개=1행(CREWID/EMPID/Manager N), 매니저=Y, 2배정=2행, 무배정=빈 1행."""
    tracked = cleanup_created_users

    # A: staff — 매장 1개 배정 (일반)
    a = await _create_user(async_client, admin_headers, seed_roles["staff"], tracked)
    await _assign_stores(
        async_client, admin_headers, a["id"],
        [{"store_id": str(test_store_id), "is_manager": False, "is_work_assignment": True}],
    )
    # B: supervisor — 매장 1개 매니저 배정
    b = await _create_user(
        async_client, admin_headers, seed_roles["supervisor"], tracked
    )
    await _assign_stores(
        async_client, admin_headers, b["id"],
        [{"store_id": str(test_store_id), "is_manager": True, "is_work_assignment": True}],
    )
    # C: staff — 매장 2개 배정 (행 2개 기대)
    c = await _create_user(async_client, admin_headers, seed_roles["staff"], tracked)
    await _assign_stores(
        async_client, admin_headers, c["id"],
        [
            {"store_id": str(test_store_id), "is_manager": False, "is_work_assignment": True},
            {"store_id": str(second_store_id), "is_manager": False, "is_work_assignment": True},
        ],
    )
    # D: staff — 배정 없음 (매장 칸 빈 1행 기대)
    d = await _create_user(async_client, admin_headers, seed_roles["staff"], tracked)

    _headers, rows, _resp = await _export(async_client, admin_headers)

    # A — 배정 1개 = 1행, org_member 소스 값 배선 확인
    a_rows = _rows_for(rows, a["username"])
    assert len(a_rows) == 1, a_rows
    row = a_rows[0]
    assert row["Name"] == a["full_name"]
    assert isinstance(row["CREWID"], int)  # org_members.crewid
    assert row["Role"] == "staff"
    assert row["Account Status"] == "Active"
    assert row["Employment Status"] == "active"  # org_members.status
    assert row["Store"] == "__attendance_test_store__"
    assert isinstance(row["Store EMPID"], int)  # org_member_stores.empid
    assert row["Manager"] == "N"

    # B — 매니저 배정 = Y
    b_rows = _rows_for(rows, b["username"])
    assert len(b_rows) == 1, b_rows
    assert b_rows[0]["Manager"] == "Y"
    assert b_rows[0]["Role"] == "supervisor"

    # C — 배정 2개 = 행 2개 (매장별 EMPID 독립)
    c_rows = _rows_for(rows, c["username"])
    assert len(c_rows) == 2, c_rows
    assert {r["Store"] for r in c_rows} == {
        "__attendance_test_store__",
        "__attendance_test_store_B__",
    }
    assert all(isinstance(r["Store EMPID"], int) for r in c_rows)
    # 같은 사람이면 CREWID 는 행마다 동일 (org 번호는 사람×org 당 1개)
    assert len({r["CREWID"] for r in c_rows}) == 1

    # D — 배정 없음 = 매장 칸 빈 1행
    d_rows = _rows_for(rows, d["username"])
    assert len(d_rows) == 1, d_rows
    assert _blank(d_rows[0]["Store"])
    assert _blank(d_rows[0]["Store EMPID"])
    assert _blank(d_rows[0]["Manager"])
    assert isinstance(d_rows[0]["CREWID"], int)  # 소속 자체는 있다


async def test_export_employment_fields_from_org_member(
    async_client: AsyncClient,
    admin_headers: dict,
    seed_roles: dict,
    test_store_id,
    cleanup_created_users: list[str],
) -> None:
    """Employment Status / Hire / Termination 은 org_members 값을 그대로 낸다."""
    tracked = cleanup_created_users
    u = await _create_user(async_client, admin_headers, seed_roles["staff"], tracked)
    await _assign_stores(
        async_client, admin_headers, u["id"],
        [{"store_id": str(test_store_id), "is_manager": False, "is_work_assignment": True}],
    )

    # 재직 이력 시뮬레이션 — org_members 컬럼 직접 세팅 (export 소스 검증 목적)
    async with async_session() as db:
        member = (
            await db.execute(
                select(OrgMember).where(OrgMember.user_id == uuid.UUID(u["id"]))
            )
        ).scalar_one()
        member.status = "terminated"
        member.hire_date = date(2024, 1, 15)
        member.termination_date = date(2026, 1, 31)
        await db.commit()

    _headers, rows, _resp = await _export(async_client, admin_headers)
    u_rows = _rows_for(rows, u["username"])
    assert len(u_rows) == 1, u_rows
    assert u_rows[0]["Employment Status"] == "terminated"
    assert u_rows[0]["Hire Date"] == "2024-01-15"
    assert u_rows[0]["Termination Date"] == "2026-01-31"


async def test_export_provisional_account_status(
    async_client: AsyncClient,
    admin_headers: dict,
    seed_roles: dict,
    test_store_id,
    cleanup_created_users: list[str],
) -> None:
    """미가입(유령) 계정은 Account Status 가 "Not signed up" (Inactive 아님)."""
    # store_ids 없이 생성 후 배정은 직접 시딩 (모듈 docstring — 배정 API 일시 500)
    resp = await async_client.post(
        "/api/v1/console/users/provisional",
        headers=admin_headers,
        json={
            "full_name": f"Ghost Export {uuid.uuid4().hex[:4]}",
            "role_id": str(seed_roles["staff"]),
        },
    )
    assert resp.status_code == 201, resp.text
    ghost = resp.json()
    cleanup_created_users.append(ghost["username"])
    await _assign_stores(
        async_client, admin_headers, ghost["id"],
        [{"store_id": str(test_store_id), "is_manager": False, "is_work_assignment": True}],
    )

    _headers, rows, _resp = await _export(async_client, admin_headers)
    g_rows = _rows_for(rows, ghost["username"])
    assert len(g_rows) == 1, g_rows
    assert g_rows[0]["Account Status"] == "Not signed up"
    assert isinstance(g_rows[0]["CREWID"], int)
    assert isinstance(g_rows[0]["Store EMPID"], int)


# ---------------------------------------------------------------------------
# store_ids 필터
# ---------------------------------------------------------------------------


async def test_export_store_ids_filter(
    async_client: AsyncClient,
    admin_headers: dict,
    seed_roles: dict,
    test_store_id,
    second_store_id,
    cleanup_created_users: list[str],
) -> None:
    """필터 매장 배정 행만 — 다른 매장만 배정된 사람/무배정자는 사람째 제외."""
    tracked = cleanup_created_users

    a = await _create_user(async_client, admin_headers, seed_roles["staff"], tracked)
    await _assign_stores(
        async_client, admin_headers, a["id"],
        [{"store_id": str(test_store_id), "is_manager": False, "is_work_assignment": True}],
    )
    b = await _create_user(async_client, admin_headers, seed_roles["staff"], tracked)
    await _assign_stores(
        async_client, admin_headers, b["id"],
        [{"store_id": str(second_store_id), "is_manager": False, "is_work_assignment": True}],
    )
    c = await _create_user(async_client, admin_headers, seed_roles["staff"], tracked)

    # 단일 매장 필터 — B 만
    _h, rows, _r = await _export(
        async_client, admin_headers, store_ids=str(second_store_id)
    )
    assert _rows_for(rows, b["username"]), "assigned user missing from filtered export"
    assert not _rows_for(rows, a["username"]), "other-store user leaked into filter"
    assert not _rows_for(rows, c["username"]), "unassigned user leaked into filter"
    # 필터 결과의 모든 행은 그 매장 행이어야 한다
    assert rows and all(r["Store"] == "__attendance_test_store_B__" for r in rows)

    # 복수(콤마) 필터 — A/B 둘 다, C(무배정)는 여전히 제외
    _h, rows, _r = await _export(
        async_client, admin_headers,
        store_ids=f"{test_store_id},{second_store_id}",
    )
    assert _rows_for(rows, a["username"])
    assert _rows_for(rows, b["username"])
    assert not _rows_for(rows, c["username"])
