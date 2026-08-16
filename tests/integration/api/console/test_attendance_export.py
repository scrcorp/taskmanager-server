"""Integration tests — attendance xlsx export.

GET /api/v1/console/attendances/export 검증:
    - 고정 경로가 /{attendance_id} 에 먹히지 않고 200 (등록 순서 회귀 방지)
    - 헤더가 정확히 시간 기록 스코프 — 급여/프리미엄/OT 컬럼 부재
    - EMPID(org_member_stores.empid) / 매장 tz 벽시계 HH:MM / breaks 세션
      (paid_10min 포함) / paid·unpaid 분 집계 / 진행 중 clock_out 빈칸
    - date_from > date_to → 400

주의: 테스트 데이터는 고정 과거 주(2026-04-05 ~ 04-11)에 생성하고,
테스트 유저의 해당 범위 attendance 를 전후로 purge 해 dev DB 잔재와 격리한다.
"""

from __future__ import annotations

from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from uuid import UUID

import pytest
import pytest_asyncio
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import delete, select

from app.database import async_session
from app.models.attendance import Attendance
from app.models.attendance_break import (
    BREAK_TYPE_PAID_10MIN,
    BREAK_TYPE_UNPAID_MEAL,
    AttendanceBreak,
)
from app.models.org_member import OrgMember, OrgMemberStore
from app.services.attendance_export_service import EXPORT_HEADERS


pytestmark = pytest.mark.asyncio

# 고정 과거 1주 — dashboard export 테스트 범위(5월)와 겹치지 않게 4월 사용.
RANGE_FROM = date(2026, 4, 5)   # Sun
RANGE_TO = date(2026, 4, 11)    # Sat

# 매장 내 unique 충돌이 거의 없을 높은 번호 (fixture 에서 선점 행 정리).
TEST_EMPID = 9917


async def _purge_range_attendance(user_ids: list[UUID]) -> None:
    """테스트 유저의 대상 범위 attendance 제거 (매장 무관 — dev DB 잔재 격리)."""
    async with async_session() as db:
        await db.execute(
            delete(Attendance).where(
                Attendance.user_id.in_(user_ids),
                Attendance.work_date >= RANGE_FROM,
                Attendance.work_date <= RANGE_TO,
            )
        )
        await db.commit()


@pytest_asyncio.fixture
async def clean_export_range(test_users: dict):
    """대상 주 attendance 전후 정리."""
    user_ids = [info["id"] for info in test_users.values()]
    await _purge_range_attendance(user_ids)
    yield
    await _purge_range_attendance(user_ids)


@pytest_asyncio.fixture
async def staff_empid(
    seed_organization: dict,
    seed_roles: dict[str, UUID],
    test_users: dict,
    test_store_id: UUID,
) -> int:
    """teststaff 의 테스트 매장 empid 를 TEST_EMPID 로 보장. 종료 시 원복.

    org_member / org_member_stores 행이 없으면 만들고, 있으면 empid 만 바꾼다.
    empid 는 (store, empid) partial unique 라 선점 행이 있으면 먼저 비운다.
    """
    staff_id: UUID = test_users["teststaff"]["id"]
    org_id: UUID = seed_organization["id"]

    async with async_session() as db:
        # 같은 매장에서 TEST_EMPID 를 이미 쓰는 행이 있으면 비운다 (unique 충돌 방지)
        holder = (
            await db.execute(
                select(OrgMemberStore).where(
                    OrgMemberStore.store_id == test_store_id,
                    OrgMemberStore.empid == TEST_EMPID,
                )
            )
        ).scalar_one_or_none()
        holder_id = None
        if holder is not None:
            holder_id = holder.id
            holder.empid = None
            await db.flush()

        member = (
            await db.execute(
                select(OrgMember).where(
                    OrgMember.user_id == staff_id,
                    OrgMember.organization_id == org_id,
                )
            )
        ).scalar_one_or_none()
        created_member = False
        if member is None:
            member = OrgMember(
                user_id=staff_id,
                organization_id=org_id,
                role_id=seed_roles["staff"],
            )
            db.add(member)
            await db.flush()
            created_member = True

        assignment = (
            await db.execute(
                select(OrgMemberStore).where(
                    OrgMemberStore.org_member_id == member.id,
                    OrgMemberStore.store_id == test_store_id,
                )
            )
        ).scalar_one_or_none()
        created_assignment = False
        prev_empid: int | None = None
        if assignment is None:
            assignment = OrgMemberStore(
                org_member_id=member.id,
                store_id=test_store_id,
                empid=TEST_EMPID,
            )
            db.add(assignment)
            created_assignment = True
        else:
            prev_empid = assignment.empid
            assignment.empid = TEST_EMPID
        member_id = member.id
        assignment_id = assignment.id if not created_assignment else None
        await db.commit()

    yield TEST_EMPID

    async with async_session() as db:
        if created_assignment:
            await db.execute(
                delete(OrgMemberStore).where(
                    OrgMemberStore.org_member_id == member_id,
                    OrgMemberStore.store_id == test_store_id,
                )
            )
        elif assignment_id is not None:
            restored = (
                await db.execute(
                    select(OrgMemberStore).where(OrgMemberStore.id == assignment_id)
                )
            ).scalar_one_or_none()
            if restored is not None:
                restored.empid = prev_empid
        if created_member:
            await db.execute(delete(OrgMember).where(OrgMember.id == member_id))
        if holder_id is not None:
            back = (
                await db.execute(
                    select(OrgMemberStore).where(OrgMemberStore.id == holder_id)
                )
            ).scalar_one_or_none()
            if back is not None:
                back.empid = TEST_EMPID
        await db.commit()


async def _create_attendance(
    *,
    org_id: UUID,
    store_id: UUID,
    user_id: UUID,
    work_date: date,
    clock_in_hm: tuple[int, int] | None,
    clock_out_hm: tuple[int, int] | None,
    status: str,
    breaks: list[tuple[str, tuple[int, int], tuple[int, int]]] | None = None,
) -> UUID:
    """attendance + break 세션 직접 생성.

    테스트 매장 tz=UTC 라 (hour, minute) UTC 가 곧 화면 벽시계 HH:MM.
    breaks: (break_type, (sh, sm), (eh, em)) 목록.
    """
    async with async_session() as db:
        att = Attendance(
            organization_id=org_id,
            store_id=store_id,
            user_id=user_id,
            work_date=work_date,
            status=status,
            clock_in=(
                datetime.combine(work_date, time(*clock_in_hm), tzinfo=timezone.utc)
                if clock_in_hm
                else None
            ),
            clock_out=(
                datetime.combine(work_date, time(*clock_out_hm), tzinfo=timezone.utc)
                if clock_out_hm
                else None
            ),
        )
        db.add(att)
        await db.flush()
        for break_type, (sh, sm), (eh, em) in breaks or []:
            started = datetime.combine(work_date, time(sh, sm), tzinfo=timezone.utc)
            ended = datetime.combine(work_date, time(eh, em), tzinfo=timezone.utc)
            db.add(
                AttendanceBreak(
                    attendance_id=att.id,
                    started_at=started,
                    ended_at=ended,
                    break_type=break_type,
                    duration_minutes=int((ended - started).total_seconds() // 60),
                )
            )
        await db.commit()
        return att.id


def _sheet_rows(content: bytes) -> tuple[list[str], list[dict]]:
    """응답 xlsx → (헤더, 데이터 row dict 목록)."""
    wb = load_workbook(BytesIO(content), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h) for h in next(rows_iter)]
    rows = [dict(zip(headers, r)) for r in rows_iter]
    wb.close()
    return headers, rows


async def test_export_columns_and_values(
    async_client: AsyncClient,
    admin_headers: dict,
    clean_export_range,
    staff_empid: int,
    seed_organization: dict,
    test_users: dict,
    test_store_id: UUID,
) -> None:
    """완료 근무 1건 — EMPID/벽시계 HH:MM/breaks(paid 10m 포함)/분 집계 검증.

    급여 스코프 회귀 방지: 헤더가 정확히 EXPORT_HEADERS 이고 pay 계열 단어 부재.
    """
    org_id = seed_organization["id"]
    staff = test_users["teststaff"]
    work_date = RANGE_FROM + timedelta(days=1)  # Mon 04-06

    await _create_attendance(
        org_id=org_id,
        store_id=test_store_id,
        user_id=staff["id"],
        work_date=work_date,
        clock_in_hm=(9, 0),
        clock_out_hm=(17, 0),
        status="clocked_out",
        breaks=[
            (BREAK_TYPE_PAID_10MIN, (12, 0), (12, 10)),
            (BREAK_TYPE_UNPAID_MEAL, (13, 0), (13, 30)),
        ],
    )

    resp = await async_client.get(
        "/api/v1/console/attendances/export",
        headers=admin_headers,
        params={
            "date_from": str(RANGE_FROM),
            "date_to": str(RANGE_TO),
            "store_id": str(test_store_id),
        },
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        f"attendance_{RANGE_FROM}_{RANGE_TO}.xlsx"
        in resp.headers["content-disposition"]
    )

    headers, rows = _sheet_rows(resp.content)
    assert headers == EXPORT_HEADERS
    # 급여/프리미엄/OT 컬럼 절대 금지 (사용자 요구)
    banned = ("pay", "rate", "wage", "premium", "overtime", "ot ", "amount", "tip")
    for h in headers:
        assert not any(b in h.lower() for b in banned), h

    mine = [r for r in rows if r["Staff"] == staff["full_name"]]
    assert len(mine) == 1, rows
    row = mine[0]
    assert row["EMPID"] == staff_empid
    assert row["Store"] == "__attendance_test_store__"
    assert row["Date"] == str(work_date)
    assert row["Clock In"] == "09:00"
    assert row["Clock Out"] == "17:00"
    assert row["Breaks"] == "12:00–12:10 paid 10m; 13:00–13:30 meal"
    assert row["Paid Break Min"] == 10
    assert row["Unpaid Break Min"] == 30
    assert row["Status"] == "clocked_out"


async def test_export_in_progress_blank_clock_out_and_sort(
    async_client: AsyncClient,
    admin_headers: dict,
    clean_export_range,
    seed_organization: dict,
    test_users: dict,
    test_store_id: UUID,
) -> None:
    """진행 중(clock_out 없음) 은 Clock Out 빈칸 + Date 오름차순 정렬."""
    org_id = seed_organization["id"]
    staff = test_users["teststaff"]
    gm = test_users["testgm"]

    # 화요일 — 진행 중 (working, clock_out 없음)
    await _create_attendance(
        org_id=org_id,
        store_id=test_store_id,
        user_id=staff["id"],
        work_date=RANGE_FROM + timedelta(days=2),
        clock_in_hm=(8, 30),
        clock_out_hm=None,
        status="working",
    )
    # 월요일 — 완료 (다른 유저, 정렬 확인용)
    await _create_attendance(
        org_id=org_id,
        store_id=test_store_id,
        user_id=gm["id"],
        work_date=RANGE_FROM + timedelta(days=1),
        clock_in_hm=(10, 0),
        clock_out_hm=(18, 0),
        status="clocked_out",
    )

    resp = await async_client.get(
        "/api/v1/console/attendances/export",
        headers=admin_headers,
        params={
            "date_from": str(RANGE_FROM),
            "date_to": str(RANGE_TO),
            "store_id": str(test_store_id),
        },
    )
    assert resp.status_code == 200, resp.text

    _, rows = _sheet_rows(resp.content)
    names = {staff["full_name"], gm["full_name"]}
    ours = [r for r in rows if r["Staff"] in names]
    assert len(ours) == 2, rows

    # Date 오름차순 — gm(월) 이 staff(화) 보다 먼저
    assert [r["Staff"] for r in ours] == [gm["full_name"], staff["full_name"]]

    in_progress = next(r for r in ours if r["Staff"] == staff["full_name"])
    assert in_progress["Clock In"] == "08:30"
    # openpyxl 은 빈 문자열 셀을 None 으로 읽는다
    assert in_progress["Clock Out"] in ("", None)
    assert in_progress["Breaks"] in ("", None)
    assert in_progress["Paid Break Min"] == 0
    assert in_progress["Unpaid Break Min"] == 0
    assert in_progress["Status"] == "working"


async def test_export_invalid_range_400(
    async_client: AsyncClient,
    admin_headers: dict,
) -> None:
    """date_from > date_to → 400 (UUID 파싱 422 가 아니어야 함 — 경로 순서 회귀 방지)."""
    resp = await async_client.get(
        "/api/v1/console/attendances/export",
        headers=admin_headers,
        params={"date_from": "2026-04-11", "date_to": "2026-04-05"},
    )
    assert resp.status_code == 400, resp.text
