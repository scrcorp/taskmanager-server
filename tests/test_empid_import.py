"""empid_import_service 테스트 — preview 분류 + commit 3-phase 재기입.

각 테스트는 uuid4 suffix 로 고유한 org/store/user 를 새로 만들어 격리한다
(워크트리 DB 에 실데이터가 있어도 기존 데이터에 의존/영향 없음).
xlsx 는 openpyxl 인메모리 생성(tests/test_empid_reconcile.py 패턴).
"""

from __future__ import annotations

import io
from typing import AsyncIterator
from uuid import UUID, uuid4

import openpyxl
import pytest_asyncio
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import async_session
from app.models.org_member import OrgMember, OrgMemberStore
from app.models.organization import Organization, Store, StoreGroup
from app.models.user import Role, User
from app.services import empid_import_service as svc
from app.services.empid_import_service import (
    ACTION_INVALID,
    ACTION_NEW_ASSIGNMENT,
    ACTION_REBIND,
    ACTION_SAME,
    ACTION_UNMATCHED_STORE,
    build_store_index,
    match_store,
)

# asyncio_mode=auto — async 테스트 자동 마킹 (파일 전역 마크는 sync 테스트에 경고 유발)


# ---------------------------------------------------------------------------
# 격리 시드 헬퍼 — 테스트마다 고유 org/role/store/user 직접 INSERT
# ---------------------------------------------------------------------------


class Ctx:
    """테스트 1개 전용 org 컨텍스트."""

    def __init__(self, org_id: UUID, role_id: UUID, sfx: str):
        self.org_id = org_id
        self.role_id = role_id
        self.sfx = sfx


@pytest_asyncio.fixture
async def ctx(db: AsyncSession) -> AsyncIterator[Ctx]:
    """고유 org + staff role 생성, 종료 시 org 삭제(CASCADE 정리)."""
    sfx = uuid4().hex[:8]
    org = Organization(name=f"__imptest_org_{sfx}__")
    db.add(org)
    await db.flush()
    role = Role(organization_id=org.id, name="staff", priority=40)
    db.add(role)
    await db.flush()
    org_id, role_id = org.id, role.id
    await db.commit()
    try:
        yield Ctx(org_id, role_id, sfx)
    finally:
        async with async_session() as s:
            await s.execute(delete(Organization).where(Organization.id == org_id))
            await s.commit()


async def _make_user(db: AsyncSession, ctx: Ctx, tag: str, email: str) -> UUID:
    """user + org_member 생성 (username 전역 unique — sfx 필수)."""
    user = User(
        organization_id=ctx.org_id,
        role_id=ctx.role_id,
        username=f"__imptest_{tag}_{ctx.sfx}",
        full_name=f"Imp Test {tag}",
        password_hash="x",
        email=email,
        is_active=True,
    )
    db.add(user)
    await db.flush()
    db.add(OrgMember(user_id=user.id, organization_id=ctx.org_id, role_id=ctx.role_id))
    await db.flush()
    return user.id


async def _make_store(
    db: AsyncSession, ctx: Ctx, tag: str, group_id: UUID | None = None
) -> UUID:
    """이름 = "IMPTEST STORE {tag} {sfx}" — COMPANY 정규화 매칭 대상."""
    store = Store(
        organization_id=ctx.org_id,
        name=f"IMPTEST STORE {tag} {ctx.sfx}",
        timezone="UTC",
        group_id=group_id,
    )
    db.add(store)
    await db.flush()
    return store.id


async def _member_id(db: AsyncSession, ctx: Ctx, user_id: UUID) -> UUID:
    return (
        await db.execute(
            select(OrgMember.id).where(
                OrgMember.user_id == user_id, OrgMember.organization_id == ctx.org_id
            )
        )
    ).scalar_one()


async def _give_empid(
    db: AsyncSession, ctx: Ctx, user_id: UUID, store_id: UUID, empid: int
) -> None:
    """기존 배정+번호 상태 시드 — org_member_stores 직접 INSERT."""
    db.add(
        OrgMemberStore(
            org_member_id=await _member_id(db, ctx, user_id),
            store_id=store_id,
            empid=empid,
        )
    )
    await db.flush()


async def _empids_in_store(db: AsyncSession, store_id: UUID) -> dict[UUID, int | None]:
    """{user_id: empid} — 검증용."""
    rows = (
        await db.execute(
            select(OrgMember.user_id, OrgMemberStore.empid)
            .join(OrgMember, OrgMember.id == OrgMemberStore.org_member_id)
            .where(OrgMemberStore.store_id == store_id)
        )
    ).all()
    return {r.user_id: r.empid for r in rows}


def _xlsx(rows: list[list]) -> bytes:
    """[[COMPANY, CORP_ABR_3, Name, emp_id, Email], ...] → xlsx bytes (인메모리)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["COMPANY", "CORP_ABR_3", "Name", "emp_id", "Email"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _entries_by_action(person) -> dict[str, list]:
    out: dict[str, list] = {}
    for e in person.entries:
        out.setdefault(e.action, []).append(e)
    return out


# ---------------------------------------------------------------------------
# preview — (a) same / (b) rebind / (c) new_assignment / (d) unmatched / (e) invalid
# ---------------------------------------------------------------------------


async def test_preview_actions_same_rebind_new_unmatched_invalid(
    db: AsyncSession, ctx: Ctx
) -> None:
    store_a = await _make_store(db, ctx, "A")
    store_b = await _make_store(db, ctx, "B")
    email1 = f"alice.{ctx.sfx}@example.com"
    email2 = f"bob.{ctx.sfx}@example.com"
    email3 = f"carol.{ctx.sfx}@example.com"
    u1 = await _make_user(db, ctx, "alice", email1)
    u2 = await _make_user(db, ctx, "bob", email2)
    u3 = await _make_user(db, ctx, "carol", email3)
    await _give_empid(db, ctx, u1, store_a, 5)   # (a) 업로드값 == 현재값
    await _give_empid(db, ctx, u2, store_a, 6)   # (b) 업로드값 != 현재값
    await db.commit()

    company_a = f"imptest-store-a-{ctx.sfx}"     # 정규화 매칭 (소문자/구분자 무시)
    company_b = f"IMPTEST STORE B {ctx.sfx}"
    content = _xlsx([
        [company_a, None, "ALICE KIM", "5", email1],                  # (a) same
        [f"ZZZ NOWHERE {ctx.sfx}", None, "ALICE KIM", "7", email1],   # (d) unmatched_store
        [company_a, None, "BOB LEE", "9", email2],                    # (b) rebind (6→9)
        [company_b, None, "BOB LEE", "ABC", email2],                  # (e) invalid (비정수)
        [company_b, None, "CAROL PARK", "4", email3],                 # (c) new_assignment
    ])

    res = await svc.preview(db, ctx.org_id, content, "list.xlsx")

    counts = res.counts()
    assert counts["people"] == 3
    assert counts["same"] == 1
    assert counts["rebind"] == 1
    assert counts["new_assignment"] == 1
    assert counts["unmatched_store"] == 1
    assert counts["invalid"] == 1
    assert not res.placeholder and not res.deferred

    by_email = {p.email: p for p in res.people}

    # (a) same — 현재값 5 == 업로드 5
    alice = _entries_by_action(by_email[email1])
    same = alice[ACTION_SAME][0]
    assert same.store_id == str(store_a)
    assert same.emp_id == 5 and same.current_empid == 5 and same.has_assignment

    # (d) unmatched_store — 매장 미매칭, store_id 없음
    unmatched = alice[ACTION_UNMATCHED_STORE][0]
    assert unmatched.store_id is None and unmatched.emp_id == 7

    # (b) rebind — current_empid 병기
    bob = _entries_by_action(by_email[email2])
    rebind = bob[ACTION_REBIND][0]
    assert rebind.store_id == str(store_a)
    assert rebind.emp_id == 9 and rebind.current_empid == 6 and rebind.has_assignment

    # (e) invalid — 비정수 emp_id (매장은 매칭됨)
    invalid = bob[ACTION_INVALID][0]
    assert invalid.store_id == str(store_b)
    assert invalid.emp_id is None and invalid.emp_id_raw == "ABC"

    # (c) new_assignment — 배정 행 없음
    carol = _entries_by_action(by_email[email3])
    new = carol[ACTION_NEW_ASSIGNMENT][0]
    assert new.store_id == str(store_b)
    assert new.emp_id == 4 and new.current_empid is None and not new.has_assignment


# ---------------------------------------------------------------------------
# (f) match_store — name/code 정규화 매칭 + 모호(동일 키 매장 2개) → None
# ---------------------------------------------------------------------------


def test_match_store_normalization_and_ambiguity() -> None:
    org_id = uuid4()
    # 미저장 Store 는 id 기본값이 안 채워지므로 명시 (모호 판정이 id 유니크화 기반)
    il_fiora = Store(id=uuid4(), organization_id=org_id, name="IL FIORA", code="IFO")
    bbq = Store(id=uuid4(), organization_id=org_id, name="M KOREAN BBQ", code="MKB")
    index = build_store_index([il_fiora, bbq])

    # name 정규화 매칭 — 대소문자/구분자 무시
    assert match_store(index, "il-fiora!!", None) is il_fiora
    # COMPANY 미매칭이면 CORP_ABR_3(code) 폴백
    assert match_store(index, "no such company", "ifo") is il_fiora
    # name == code 로 이중 등록된 자기 자신은 모호 아님 (id 유니크화)
    solo = Store(id=uuid4(), organization_id=org_id, name="SOLO", code="SOLO")
    assert match_store(build_store_index([solo]), "solo", None) is solo

    # 모호 — 정규화 키가 같은 매장 2개 → None
    twin1 = Store(id=uuid4(), organization_id=org_id, name="TWIN", code=None)
    twin2 = Store(id=uuid4(), organization_id=org_id, name="T-W-I-N", code=None)
    assert match_store(build_store_index([twin1, twin2]), "twin", None) is None
    # 전혀 없는 키 → None
    assert match_store(index, "ghost mart", None) is None


# ---------------------------------------------------------------------------
# commit — (g) rebind 반영 + 멱등
# ---------------------------------------------------------------------------


async def test_commit_rebind_applies_then_idempotent_skip(db: AsyncSession, ctx: Ctx) -> None:
    store = await _make_store(db, ctx, "A")
    u = await _make_user(db, ctx, "u", f"u.{ctx.sfx}@example.com")
    await _give_empid(db, ctx, u, store, 5)
    await db.commit()

    r1 = await svc.commit(db, ctx.org_id, [(u, store, 7)])
    assert len(r1.applied) == 1 and not r1.skipped and not r1.rejected and not r1.renumbered
    assert r1.applied[0]["empid"] == 7 and r1.applied[0]["created"] is False
    assert (await _empids_in_store(db, store))[u] == 7

    # 같은 요청 재실행 → skipped(already set), DB 불변 (멱등)
    r2 = await svc.commit(db, ctx.org_id, [(u, store, 7)])
    assert not r2.applied and not r2.renumbered
    assert len(r2.skipped) == 1 and r2.skipped[0]["reason"] == "already set"
    assert (await _empids_in_store(db, store))[u] == 7


# ---------------------------------------------------------------------------
# commit — (h) 순열 교환 (A:5→3, B:3→5) 이 IntegrityError 없이 통과
# ---------------------------------------------------------------------------


async def test_commit_swap_passes_partial_unique(db: AsyncSession, ctx: Ctx) -> None:
    store = await _make_store(db, ctx, "A")
    ua = await _make_user(db, ctx, "ua", f"ua.{ctx.sfx}@example.com")
    ub = await _make_user(db, ctx, "ub", f"ub.{ctx.sfx}@example.com")
    await _give_empid(db, ctx, ua, store, 5)
    await _give_empid(db, ctx, ub, store, 3)
    await db.commit()

    # 3-phase(비우기→기입) 덕에 (store, empid) partial unique 에 안 걸려야 한다
    r = await svc.commit(db, ctx.org_id, [(ua, store, 3), (ub, store, 5)])
    assert len(r.applied) == 2 and not r.rejected and not r.renumbered

    final = await _empids_in_store(db, store)
    assert final[ua] == 3 and final[ub] == 5


# ---------------------------------------------------------------------------
# commit — (i) 번호 뺏긴 제3자 재채번(renumbered) + 매장 내 unique 유지
# ---------------------------------------------------------------------------


async def test_commit_renumbers_displaced_third_party(db: AsyncSession, ctx: Ctx) -> None:
    store = await _make_store(db, ctx, "A")
    ux = await _make_user(db, ctx, "ux", f"ux.{ctx.sfx}@example.com")
    uy = await _make_user(db, ctx, "uy", f"uy.{ctx.sfx}@example.com")
    uz = await _make_user(db, ctx, "uz", f"uz.{ctx.sfx}@example.com")
    await _give_empid(db, ctx, ux, store, 1)
    await _give_empid(db, ctx, uy, store, 2)
    await _give_empid(db, ctx, uz, store, 3)
    await db.commit()

    # x 가 3 을 가져감 → 기존 3 소유자 z 는 next_empid 로 재채번
    r = await svc.commit(db, ctx.org_id, [(ux, store, 3)])
    assert len(r.applied) == 1 and r.applied[0]["empid"] == 3
    assert len(r.renumbered) == 1
    assert r.renumbered[0]["old"] == 3 and r.renumbered[0]["new"] == 4

    final = await _empids_in_store(db, store)
    assert final[ux] == 3 and final[uy] == 2 and final[uz] == 4
    # 매장 내 unique 유지
    values = [v for v in final.values() if v is not None]
    assert len(values) == len(set(values))


# ---------------------------------------------------------------------------
# commit — (j) new_assignment: OrgMemberStore 행 생성 (is_work_assignment=True)
# ---------------------------------------------------------------------------


async def test_commit_new_assignment_creates_member_store_row(
    db: AsyncSession, ctx: Ctx
) -> None:
    store = await _make_store(db, ctx, "A")
    u = await _make_user(db, ctx, "new", f"new.{ctx.sfx}@example.com")  # 배정 행 없음
    await db.commit()

    r = await svc.commit(db, ctx.org_id, [(u, store, 9)])
    assert len(r.applied) == 1 and not r.rejected
    assert r.applied[0]["empid"] == 9 and r.applied[0]["created"] is True

    row = (
        await db.execute(
            select(OrgMemberStore).where(
                OrgMemberStore.org_member_id == await _member_id(db, ctx, u),
                OrgMemberStore.store_id == store,
            )
        )
    ).scalar_one()
    assert row.empid == 9
    assert row.is_work_assignment is True and row.is_manager is False


# ---------------------------------------------------------------------------
# commit — (k) 같은 매장에 두 사람이 같은 신규 값 요청 → 1 applied + 1 rejected
# ---------------------------------------------------------------------------


async def test_commit_duplicate_value_in_request_rejects_one(
    db: AsyncSession, ctx: Ctx
) -> None:
    store = await _make_store(db, ctx, "A")
    ua = await _make_user(db, ctx, "da", f"da.{ctx.sfx}@example.com")
    ub = await _make_user(db, ctx, "db", f"db.{ctx.sfx}@example.com")
    await _give_empid(db, ctx, ua, store, 1)
    await _give_empid(db, ctx, ub, store, 2)
    await db.commit()

    # 둘 다 9 를 요청 — 전체 롤백(DuplicateError) 없이 첫 요청만 반영, 나머지는 거절
    r = await svc.commit(db, ctx.org_id, [(ua, store, 9), (ub, store, 9)])
    assert len(r.applied) == 1
    assert len(r.rejected) == 1
    assert r.rejected[0]["reason"] == "duplicate empid for this store in request"
    assert not r.renumbered

    final = await _empids_in_store(db, store)
    assert final[ua] == 9 and final[ub] == 2  # ub 는 기존 값 유지


# ---------------------------------------------------------------------------
# commit — (l) 보유자 우선: 현재 보유자와 타인이 같은 값 요청 시 보유자가 이긴다
# ---------------------------------------------------------------------------


async def test_commit_duplicate_value_holder_wins_regardless_of_order(
    db: AsyncSession, ctx: Ctx
) -> None:
    store = await _make_store(db, ctx, "A")
    holder = await _make_user(db, ctx, "hold", f"hold.{ctx.sfx}@example.com")
    other = await _make_user(db, ctx, "oth", f"oth.{ctx.sfx}@example.com")
    await _give_empid(db, ctx, holder, store, 3)
    await _give_empid(db, ctx, other, store, 1)
    await db.commit()

    # [타인 먼저, 보유자 나중] 순서로 같은 3 을 요청 — 보유자가 살아남아야 함
    r = await svc.commit(db, ctx.org_id, [(other, store, 3), (holder, store, 3)])
    assert len(r.skipped) == 1 and r.skipped[0]["reason"] == "already set"
    assert len(r.rejected) == 1
    assert r.rejected[0]["user_id"] == str(other)
    assert not r.applied and not r.renumbered

    final = await _empids_in_store(db, store)
    assert final[holder] == 3 and final[other] == 1  # 아무도 안 움직임


# ---------------------------------------------------------------------------
# commit — (m) 그룹 공유 스코프 2-pass: 재채번이 형제 매장 기입값과 충돌하지 않음
# ---------------------------------------------------------------------------


async def test_commit_group_scope_renumber_after_sibling_writes(
    db: AsyncSession, ctx: Ctx
) -> None:
    group = StoreGroup(
        organization_id=ctx.org_id, name=f"__imptest_grp_{ctx.sfx}__",
        numbering_mode="group",
    )
    db.add(group)
    await db.flush()
    store_a = await _make_store(db, ctx, "A", group_id=group.id)
    store_b = await _make_store(db, ctx, "B", group_id=group.id)
    u1 = await _make_user(db, ctx, "g1", f"g1.{ctx.sfx}@example.com")
    u2 = await _make_user(db, ctx, "g2", f"g2.{ctx.sfx}@example.com")
    u3 = await _make_user(db, ctx, "g3", f"g3.{ctx.sfx}@example.com")
    await _give_empid(db, ctx, u1, store_a, 1)
    await _give_empid(db, ctx, u2, store_a, 2)
    await _give_empid(db, ctx, u3, store_b, 3)  # 그룹 공유 max = 3
    await db.commit()

    # A: u2 가 1 을 가져감 → u1 재채번 필요. B: u3 → 4 로 rebind.
    # 1-pass 순차 처리라면 u1 재채번이 4 를 받아 B 의 기입값 4 와 그룹 스코프 중복이 된다.
    # 2-pass 에서는 전 매장 기입(4 포함) 후 재채번하므로 u1 은 5 를 받아야 한다.
    r = await svc.commit(
        db, ctx.org_id, [(u2, store_a, 1), (u3, store_b, 4)]
    )
    assert len(r.applied) == 2 and not r.rejected
    assert len(r.renumbered) == 1
    assert r.renumbered[0]["old"] == 1 and r.renumbered[0]["new"] == 5

    a = await _empids_in_store(db, store_a)
    b = await _empids_in_store(db, store_b)
    assert a[u2] == 1 and a[u1] == 5 and b[u3] == 4
    # 그룹 스코프 전체에서 중복 없음
    all_values = [v for v in list(a.values()) + list(b.values()) if v is not None]
    assert len(all_values) == len(set(all_values))


# ---------------------------------------------------------------------------
# commit — (n) empid=None 은 번호 삭제 (배정 행 유지, 재채번 없음, 멱등)
# ---------------------------------------------------------------------------


async def test_commit_null_clears_number(db: AsyncSession, ctx: Ctx) -> None:
    store = await _make_store(db, ctx, "A")
    u = await _make_user(db, ctx, "cl", f"cl.{ctx.sfx}@example.com")
    await _give_empid(db, ctx, u, store, 5)
    await db.commit()

    r = await svc.commit(db, ctx.org_id, [(u, store, None)])
    assert len(r.applied) == 1 and r.applied[0]["empid"] is None
    assert not r.renumbered and not r.rejected
    assert (await _empids_in_store(db, store))[u] is None  # 행 유지, 번호만 해제

    # 재실행 — nothing to clear 로 스킵 (멱등)
    r2 = await svc.commit(db, ctx.org_id, [(u, store, None)])
    assert not r2.applied
    assert len(r2.skipped) == 1 and r2.skipped[0]["reason"] == "nothing to clear"


# ---------------------------------------------------------------------------
# preview — (o) placeholder/deferred 행이 needs_user 로 등록 가능해짐
# ---------------------------------------------------------------------------


async def test_preview_needs_user_for_placeholder_and_deferred(
    db: AsyncSession, ctx: Ctx
) -> None:
    store = await _make_store(db, ctx, "A")
    # 이름 유사 매칭 대상 유저 (파일과 이메일은 다름)
    await _make_user(db, ctx, "sim", f"sim.{ctx.sfx}@example.com")
    await db.commit()

    company = f"IMPTEST STORE A {ctx.sfx}"
    content = _xlsx([
        # 공유 이메일(서로 다른 두 사람) → placeholder, 행별 needs_user
        [company, None, "PERSON ONE", "11", f"shared.{ctx.sfx}@example.com"],
        [company, None, "OTHER TWO", "12", f"shared.{ctx.sfx}@example.com"],
        # DB 에 없는 이메일 → deferred, needs_user + similar_users 구조화
        [company, None, "Imp Test sim", "13", f"nobody.{ctx.sfx}@example.com"],
    ])
    res = await svc.preview(db, ctx.org_id, content, "list.xlsx")

    assert len(res.placeholder) == 1
    ph = res.placeholder[0]
    assert len(ph.entries) == 2
    assert all(e.action == "needs_user" for e in ph.entries)
    assert {e.person_name for e in ph.entries} == {"PERSON ONE", "OTHER TWO"}
    assert all(e.store_id == str(store) and e.emp_id in (11, 12) for e in ph.entries)

    assert len(res.deferred) == 1
    df = res.deferred[0]
    assert df.entries[0].action == "needs_user" and df.entries[0].emp_id == 13
    # 이름 유사 유저가 구조화 후보로 제공됨 (picker 프리필용)
    assert any(su["full_name"] == "Imp Test sim" for su in df.similar_users)

    assert res.counts()["needs_user"] == 3


# ---------------------------------------------------------------------------
# roster — (p) 매장별 배정·empid 현황 (empid 오름차순, 없는 사람 뒤)
# ---------------------------------------------------------------------------


async def test_template_export_roundtrip(db: AsyncSession, ctx: Ctx) -> None:
    """current export 를 그대로 재업로드하면 전부 same — 왕복 무손실."""
    store = await _make_store(db, ctx, "A")
    u1 = await _make_user(db, ctx, "t1", f"t1.{ctx.sfx}@example.com")
    u2 = await _make_user(db, ctx, "t2", f"t2.{ctx.sfx}@example.com")
    await _give_empid(db, ctx, u1, store, 5)
    await _give_empid(db, ctx, u2, store, 7)
    await db.commit()

    content = await svc.build_template_xlsx(db, ctx.org_id, prefill=True)
    res = await svc.preview(db, ctx.org_id, content, "empid_export.xlsx")
    mine = [p for p in res.people if p.email and f".{ctx.sfx}@" in p.email]
    assert len(mine) == 2
    for p in mine:
        assert all(e.action == ACTION_SAME for e in p.entries)

    # 빈 템플릿 — 데이터 0행 (Instructions 시트는 파싱에 안 잡힘)
    blank = await svc.build_template_xlsx(db, ctx.org_id, prefill=False)
    res2 = await svc.preview(db, ctx.org_id, blank, "empid_import_template.xlsx")
    assert res2.total_rows == 0


async def test_template_export_filters(db: AsyncSession, ctx: Ctx) -> None:
    """export 필터 — 매장 부분집합 / 번호 유무 / 휴면 제외 / 이메일·번호 비우기."""
    from openpyxl import load_workbook

    store_a = await _make_store(db, ctx, "A")
    store_b = await _make_store(db, ctx, "B")
    u1 = await _make_user(db, ctx, "f1", f"f1.{ctx.sfx}@example.com")  # A, 번호 5
    u2 = await _make_user(db, ctx, "f2", f"f2.{ctx.sfx}@example.com")  # A, 번호 없음
    u3 = await _make_user(db, ctx, "f3", f"f3.{ctx.sfx}@example.com")  # A, 휴면 + 번호 9
    u4 = await _make_user(db, ctx, "f4", f"f4.{ctx.sfx}@example.com")  # B, 번호 7
    await _give_empid(db, ctx, u1, store_a, 5)
    db.add(OrgMemberStore(org_member_id=await _member_id(db, ctx, u2), store_id=store_a))
    db.add(OrgMemberStore(
        org_member_id=await _member_id(db, ctx, u3), store_id=store_a,
        empid=9, is_work_assignment=False,
    ))
    await _give_empid(db, ctx, u4, store_b, 7)
    await db.commit()

    def names(content: bytes) -> list[list]:
        ws = load_workbook(io.BytesIO(content)).worksheets[0]
        return [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]

    # 매장 A만 + 휴면 제외 → u1, u2 (u3 휴면 제외, u4 타매장)
    rows = names(await svc.build_template_xlsx(
        db, ctx.org_id, True, store_ids={store_a}, include_dormant=False))
    assert {r[2] for r in rows} == {"Imp Test f1", "Imp Test f2"}

    # 번호 없는 사람만 (A+B 전체) → u2 뿐
    rows = names(await svc.build_template_xlsx(
        db, ctx.org_id, True, store_ids={store_a, store_b}, people="unnumbered"))
    assert {r[2] for r in rows} == {"Imp Test f2"}

    # 번호 있는 사람만 + 이메일·번호 비우기 → u1/u3/u4, emp_id·Email 셀 공란
    rows = names(await svc.build_template_xlsx(
        db, ctx.org_id, True, store_ids={store_a, store_b}, people="numbered",
        include_email=False, include_numbers=False))
    assert {r[2] for r in rows} == {"Imp Test f1", "Imp Test f3", "Imp Test f4"}
    assert all((r[3] in (None, "")) and (r[4] in (None, "")) for r in rows)


async def _set_crewid(db: AsyncSession, ctx: Ctx, user_id: UUID, crewid: int) -> None:
    from app.models.org_member import OrgMember as _OM
    member = (
        await db.execute(
            select(_OM).where(_OM.user_id == user_id, _OM.organization_id == ctx.org_id)
        )
    ).scalar_one()
    member.crewid = crewid
    await db.flush()


def _xlsx6(rows: list[list]) -> bytes:
    """crewid 컬럼 포함 6열 xlsx."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["COMPANY", "CORP_ABR_3", "Name", "emp_id", "Email", "crewid"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def test_crewid_exact_match_beats_email(db: AsyncSession, ctx: Ctx) -> None:
    """CREWID 가 있으면 email/이름이 달라도 정확 매칭 (개별 가입 시나리오)."""
    store = await _make_store(db, ctx, "A")
    u = await _make_user(db, ctx, "cw", f"new.signup.{ctx.sfx}@gmail.com")  # 새로 가입한 이메일
    await _set_crewid(db, ctx, u, 42)
    await _give_empid(db, ctx, u, store, 3)
    await db.commit()

    company = f"IMPTEST STORE A {ctx.sfx}"
    # 파일엔 레거시 이메일/다른 이름 — crewid 42 로만 매칭 가능
    content = _xlsx6([
        [company, None, "Legacy Name", "310", f"old.legacy.{ctx.sfx}@corp.example", "42"],
    ])
    res = await svc.preview(db, ctx.org_id, content, "list.xlsx")

    assert not res.deferred and not res.placeholder
    assert len(res.people) == 1
    p = res.people[0]
    assert p.matched_by == "crewid" and p.user_id == str(u)
    e = p.entries[0]
    assert e.action == ACTION_REBIND and e.current_empid == 3 and e.emp_id == 310
    assert "matched by CREWID" in (e.warning or "")  # 이메일 상이 표기


async def test_crewid_disambiguates_shared_email_and_no_email(
    db: AsyncSession, ctx: Ctx
) -> None:
    """공유 이메일·무이메일 행도 CREWID 만 있으면 확정 매칭 — placeholder/deferred 로 안 빠짐."""
    store = await _make_store(db, ctx, "A")
    u1 = await _make_user(db, ctx, "sh1", f"sh1.{ctx.sfx}@example.com")
    u2 = await _make_user(db, ctx, "sh2", f"sh2.{ctx.sfx}@example.com")
    await _set_crewid(db, ctx, u1, 101)
    await _set_crewid(db, ctx, u2, 102)
    await db.commit()

    company = f"IMPTEST STORE A {ctx.sfx}"
    shared = f"shared.{ctx.sfx}@example.com"
    content = _xlsx6([
        # 서로 다른 두 사람이 같은 이메일 — 원래 placeholder 감이지만 crewid 로 각각 확정
        [company, None, "PERSON ONE", "11", shared, "101"],
        [company, None, "OTHER TWO", "12", shared, "102"],
    ])
    res = await svc.preview(db, ctx.org_id, content, "list.xlsx")
    assert not res.placeholder
    assert {p.user_id for p in res.people} == {str(u1), str(u2)}
    assert all(p.matched_by == "crewid" for p in res.people)

    # 무이메일 + crewid → deferred 아님
    content2 = _xlsx6([[company, None, "No Mail", "13", None, "101"]])
    res2 = await svc.preview(db, ctx.org_id, content2, "list.xlsx")
    assert not res2.deferred and len(res2.people) == 1


async def test_crewid_not_found_requires_confirmation(db: AsyncSession, ctx: Ctx) -> None:
    """CREWID 미해결(오타 가능성)이면 email 이 맞아도 자동 등록하지 않는다 —
    email 유저를 프리필 후보로 제시하는 확인 대기(needs_user)로 분리."""
    store = await _make_store(db, ctx, "A")
    u = await _make_user(db, ctx, "fb", f"fb.{ctx.sfx}@example.com")
    await _give_empid(db, ctx, u, store, 4)
    await db.commit()

    company = f"IMPTEST STORE A {ctx.sfx}"
    content = _xlsx6([
        [company, None, "Imp Test fb", "320", f"fb.{ctx.sfx}@example.com", "99999"],
    ])
    res = await svc.preview(db, ctx.org_id, content, "list.xlsx")
    assert not res.people  # 자동 매칭 카드 없음 (빈 카드도 없음)
    assert len(res.deferred) == 1
    p = res.deferred[0]
    assert "CREWID mismatch" in p.note
    assert p.similar_users and p.similar_users[0]["user_id"] == str(u)  # 프리필 후보
    e = p.entries[0]
    assert e.action == "needs_user" and "not found" in (e.warning or "")

    # crewid 없는 행이 섞여 있으면 그 행은 정상 자동 매칭 유지
    content2 = _xlsx6([
        [company, None, "Imp Test fb", "320", f"fb.{ctx.sfx}@example.com", "99999"],
        [company, None, "Imp Test fb", "321", f"fb.{ctx.sfx}@example.com", None],
    ])
    res2 = await svc.preview(db, ctx.org_id, content2, "list.xlsx")
    assert len(res2.people) == 1 and len(res2.people[0].entries) == 1
    assert res2.people[0].entries[0].emp_id == 321
    assert len(res2.deferred) == 1  # 99999 행은 확인 대기


async def test_export_includes_crewid_column(db: AsyncSession, ctx: Ctx) -> None:
    """export 6번째 컬럼에 crewid — 재업로드 시 정확 매칭 키가 된다."""
    from openpyxl import load_workbook

    store = await _make_store(db, ctx, "A")
    u = await _make_user(db, ctx, "ex", f"ex.{ctx.sfx}@example.com")
    await _set_crewid(db, ctx, u, 77)
    await _give_empid(db, ctx, u, store, 5)
    await db.commit()

    content = await svc.build_selected_export_xlsx(db, ctx.org_id, [(u, store)])
    ws = load_workbook(io.BytesIO(content)).worksheets[0]
    header = [c.value for c in ws[1]]
    assert header == ["COMPANY", "CORP_ABR_3", "Name", "emp_id", "Email", "crewid"]
    row = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)][0]
    assert row[5] == 77

    # 왕복: export 재업로드 → crewid 정확 매칭 + same
    res = await svc.preview(db, ctx.org_id, content, "empid_export.xlsx")
    mine = [p for p in res.people if p.user_id == str(u)]
    assert len(mine) == 1 and mine[0].matched_by == "crewid"
    assert mine[0].entries[0].action == ACTION_SAME


async def test_selected_export_and_split_sheets(db: AsyncSession, ctx: Ctx) -> None:
    """사람 단위 선택 export — 선택된 (user,store)만 + split_by 시트 구분."""
    from openpyxl import load_workbook

    store_a = await _make_store(db, ctx, "A")
    store_b = await _make_store(db, ctx, "B")
    u1 = await _make_user(db, ctx, "s1", f"s1.{ctx.sfx}@example.com")  # A, 5
    u2 = await _make_user(db, ctx, "s2", f"s2.{ctx.sfx}@example.com")  # A, 310
    u3 = await _make_user(db, ctx, "s3", f"s3.{ctx.sfx}@example.com")  # B, 7 (미선택)
    await _give_empid(db, ctx, u1, store_a, 5)
    await _give_empid(db, ctx, u2, store_a, 310)
    await _give_empid(db, ctx, u3, store_b, 7)
    await db.commit()

    # 미선택(u3) 제외 확인 — 단일 시트
    content = await svc.build_selected_export_xlsx(
        db, ctx.org_id, [(u1, store_a), (u2, store_a)])
    wb = load_workbook(io.BytesIO(content))
    ws = wb.worksheets[0]
    body = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
    assert {r[2] for r in body} == {"Imp Test s1", "Imp Test s2"}

    # split_by=band — 1-99 / 300-399 시트 구분 (Instructions 마지막)
    content = await svc.build_selected_export_xlsx(
        db, ctx.org_id, [(u1, store_a), (u2, store_a)], split_by="band")
    wb = load_workbook(io.BytesIO(content))
    assert wb.sheetnames[:2] == ["1-99", "300-399"]
    assert wb.sheetnames[-1] == "Instructions"

    # split_by=store — 매장별 시트
    content = await svc.build_selected_export_xlsx(
        db, ctx.org_id, [(u1, store_a), (u3, store_b)], split_by="store")
    wb = load_workbook(io.BytesIO(content))
    assert len(wb.sheetnames) == 3  # 매장 2 + Instructions
    # roster 에 role 정보가 실리는지 (export 필터 축)
    rows = await svc.roster(db, ctx.org_id)
    mine = next(r for r in rows if r["store_id"] == str(store_a))
    assert all(m["role_name"] == "staff" for m in mine["members"])


async def test_roster_lists_store_members_sorted(db: AsyncSession, ctx: Ctx) -> None:
    store = await _make_store(db, ctx, "A")
    u1 = await _make_user(db, ctx, "r1", f"r1.{ctx.sfx}@example.com")
    u2 = await _make_user(db, ctx, "r2", f"r2.{ctx.sfx}@example.com")
    u3 = await _make_user(db, ctx, "r3", f"r3.{ctx.sfx}@example.com")
    await _give_empid(db, ctx, u1, store, 7)
    await _give_empid(db, ctx, u2, store, 3)
    # u3 는 배정만, 번호 없음
    db.add(OrgMemberStore(org_member_id=await _member_id(db, ctx, u3), store_id=store))
    await db.commit()

    rows = await svc.roster(db, ctx.org_id)
    mine = next(r for r in rows if r["store_id"] == str(store))
    empids = [m["empid"] for m in mine["members"]]
    assert empids == [3, 7, None]  # 오름차순 + 번호 없는 사람 마지막
    assert {m["user_id"] for m in mine["members"]} == {str(u1), str(u2), str(u3)}
