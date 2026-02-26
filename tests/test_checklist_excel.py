"""체크리스트 Excel Import/Export 유닛 테스트 (Mock DB).

Tests generate_sample_excel, _parse_recurrence, verification_type parsing,
and the full import_from_excel flow with a mocked async DB session.
"""

from io import BytesIO
from unittest.mock import AsyncMock, MagicMock, patch
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook

from app.services.checklist_service import ChecklistService, DAY_MAP


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_excel(rows: list[list[str]], headers: list[str] | None = None) -> bytes:
    """Build a minimal xlsx from a list of rows."""
    wb = Workbook()
    ws = wb.active
    if headers is None:
        headers = [
            "store", "shift", "position",
            "recurrence", "item_title", "item_description",
            "verification_type",
        ]
    ws.append(headers)
    ws.append([""] * len(headers))  # guide row at row 2 (skipped by importer)
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_mock_db():
    """Create a mocked async DB session with execute/flush/refresh."""
    db = AsyncMock()
    # For select queries returning None (no existing data)
    mock_result = MagicMock()
    mock_result.scalar_one_or_none.return_value = None
    mock_result.scalar.return_value = 0
    db.execute.return_value = mock_result
    return db


# ---------------------------------------------------------------------------
# 1. generate_sample_excel() 테스트
# ---------------------------------------------------------------------------

class TestGenerateSampleExcel:
    """Sample Excel 생성 검증."""

    def test_generates_valid_xlsx(self):
        """생성된 파일이 유효한 xlsx인지 확인."""
        data = ChecklistService.generate_sample_excel()
        assert isinstance(data, bytes)
        assert len(data) > 0
        # openpyxl로 열 수 있어야 함
        wb = load_workbook(filename=BytesIO(data))
        assert len(wb.sheetnames) == 2
        wb.close()

    def test_sheet_names(self):
        """시트 이름이 올바른지 확인."""
        wb = load_workbook(filename=BytesIO(ChecklistService.generate_sample_excel()))
        assert wb.sheetnames[0] == "Checklist Template"
        assert wb.sheetnames[1] == "Guide"
        wb.close()

    def test_has_required_headers(self):
        """필수 컬럼 헤더가 존재하는지 확인."""
        wb = load_workbook(filename=BytesIO(ChecklistService.generate_sample_excel()))
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for col in ["store", "shift", "position", "recurrence", "item_title"]:
            assert col in headers, f"Missing header: {col}"
        wb.close()

    def test_has_diverse_stores(self):
        """여러 매장이 포함되어 있는지 확인."""
        wb = load_workbook(filename=BytesIO(ChecklistService.generate_sample_excel()))
        ws = wb.active
        stores = set()
        for row in ws.iter_rows(min_row=3, max_col=1, values_only=True):
            if row[0]:
                stores.add(row[0])
        assert len(stores) >= 3, f"Expected ≥3 stores, got: {stores}"
        wb.close()

    def test_has_diverse_shifts(self):
        """여러 시간대가 포함되어 있는지 확인."""
        wb = load_workbook(filename=BytesIO(ChecklistService.generate_sample_excel()))
        ws = wb.active
        shifts = set()
        for row in ws.iter_rows(min_row=3, min_col=2, max_col=2, values_only=True):
            if row[0]:
                shifts.add(row[0])
        assert len(shifts) >= 3, f"Expected ≥3 shifts, got: {shifts}"
        wb.close()

    def test_has_diverse_recurrence(self):
        """다양한 recurrence 패턴이 포함되어 있는지 확인."""
        wb = load_workbook(filename=BytesIO(ChecklistService.generate_sample_excel()))
        ws = wb.active
        recurrences = set()
        for row in ws.iter_rows(min_row=3, min_col=4, max_col=4, values_only=True):
            if row[0]:
                recurrences.add(row[0])
        assert "daily" in recurrences
        assert len(recurrences) >= 3, f"Expected ≥3 recurrence patterns, got: {recurrences}"
        wb.close()

    def test_has_multi_type_verification(self):
        """photo,text 같은 multi-type verification이 포함되어 있는지 확인."""
        wb = load_workbook(filename=BytesIO(ChecklistService.generate_sample_excel()))
        ws = wb.active
        vtypes = set()
        for row in ws.iter_rows(min_row=3, min_col=7, max_col=7, values_only=True):
            if row[0]:
                vtypes.add(row[0])
        assert "photo,text" in vtypes, f"Expected 'photo,text' in vtypes, got: {vtypes}"
        wb.close()

    def test_shows_none_and_empty(self):
        """'none'과 빈 문자열 양쪽이 모두 포함되어 있는지 확인."""
        wb = load_workbook(filename=BytesIO(ChecklistService.generate_sample_excel()))
        ws = wb.active
        has_none = False
        has_empty = False
        for row in ws.iter_rows(min_row=3, min_col=7, max_col=7, values_only=True):
            val = row[0]
            if val == "none":
                has_none = True
            if val is None or val == "":
                has_empty = True
        assert has_none, "Expected at least one row with explicit 'none'"
        assert has_empty, "Expected at least one row with empty verification_type"
        wb.close()

    def test_guide_sheet_in_english(self):
        """가이드 시트가 영어로 작성되어 있는지 확인."""
        wb = load_workbook(filename=BytesIO(ChecklistService.generate_sample_excel()))
        guide = wb["Guide"]
        headers = [cell.value for cell in next(guide.iter_rows(min_row=1, max_row=1))]
        assert "Column" in headers
        assert "Required" in headers
        assert "Description" in headers
        wb.close()


# ---------------------------------------------------------------------------
# 2. _parse_recurrence() 테스트
# ---------------------------------------------------------------------------

class TestParseRecurrence:
    """recurrence 파싱 유닛 테스트."""

    def test_daily(self):
        assert ChecklistService._parse_recurrence("daily") == ("daily", None)

    def test_daily_case_insensitive(self):
        assert ChecklistService._parse_recurrence("Daily") == ("daily", None)
        assert ChecklistService._parse_recurrence(" DAILY ") == ("daily", None)

    def test_single_day(self):
        assert ChecklistService._parse_recurrence("mon") == ("weekly", [0])

    def test_multiple_days(self):
        assert ChecklistService._parse_recurrence("mon,wed,fri") == ("weekly", [0, 2, 4])

    def test_days_sorted_and_deduplicated(self):
        assert ChecklistService._parse_recurrence("fri,mon,fri") == ("weekly", [0, 4])

    def test_weekend(self):
        assert ChecklistService._parse_recurrence("sat,sun") == ("weekly", [5, 6])

    def test_all_days(self):
        """7일 전부 선택 → daily로 정규화."""
        result = ChecklistService._parse_recurrence("mon,tue,wed,thu,fri,sat,sun")
        assert result == ("daily", None)

    def test_with_spaces(self):
        assert ChecklistService._parse_recurrence(" mon , wed , fri ") == ("weekly", [0, 2, 4])

    def test_invalid_day_raises(self):
        with pytest.raises(ValueError, match="Invalid day"):
            ChecklistService._parse_recurrence("monday")

    def test_empty_days_raises(self):
        with pytest.raises(ValueError):
            ChecklistService._parse_recurrence("")


# ---------------------------------------------------------------------------
# 3. verification_type 파싱 테스트 (import_from_excel 내부 로직)
# ---------------------------------------------------------------------------

class TestVerificationTypeParsing:
    """Excel import 시 verification_type 파싱 테스트."""

    @staticmethod
    def _parse_vtype(raw: str) -> str:
        """import_from_excel 내부 v_type 파싱 로직을 재현."""
        v_type_raw = raw.lower().strip()
        if v_type_raw:
            valid_types = {"none", "photo", "text", "video"}
            parts = [p.strip() for p in v_type_raw.split(",") if p.strip()]
            parts = [p for p in parts if p in valid_types]
            if len(parts) > 1:
                parts = [p for p in parts if p != "none"]
            v_type = ",".join(parts) if parts else "none"
        else:
            v_type = "none"
        return v_type

    def test_empty_string(self):
        assert self._parse_vtype("") == "none"

    def test_none_explicit(self):
        assert self._parse_vtype("none") == "none"

    def test_photo_single(self):
        assert self._parse_vtype("photo") == "photo"

    def test_text_single(self):
        assert self._parse_vtype("text") == "text"

    def test_photo_text_multi(self):
        assert self._parse_vtype("photo,text") == "photo,text"

    def test_text_photo_preserves_order(self):
        """입력 순서대로 유지 (text,photo → text,photo)."""
        assert self._parse_vtype("text,photo") == "text,photo"

    def test_none_combined_with_photo(self):
        """none이 다른 타입과 같이 있으면 none 제거."""
        assert self._parse_vtype("none,photo") == "photo"

    def test_none_combined_with_text(self):
        assert self._parse_vtype("none,text") == "text"

    def test_none_combined_with_both(self):
        assert self._parse_vtype("none,photo,text") == "photo,text"

    def test_invalid_type_ignored(self):
        assert self._parse_vtype("picture") == "none"

    def test_mixed_valid_invalid(self):
        assert self._parse_vtype("photo,video") == "photo,video"

    def test_case_insensitive(self):
        assert self._parse_vtype("Photo") == "photo"
        assert self._parse_vtype("PHOTO,TEXT") == "photo,text"

    def test_with_spaces(self):
        assert self._parse_vtype(" photo , text ") == "photo,text"


# ---------------------------------------------------------------------------
# 4. import_from_excel() 통합 테스트 (Mock DB)
# ---------------------------------------------------------------------------

class TestImportFromExcel:
    """import_from_excel Mock 통합 테스트."""

    @pytest.fixture
    def service(self):
        return ChecklistService()

    @pytest.fixture
    def org_id(self):
        return uuid4()

    @pytest.mark.asyncio
    async def test_missing_columns_raises(self, service, org_id):
        """필수 컬럼이 없으면 ValueError."""
        bad_excel = make_excel(
            [["Store1", "Morning"]],
            headers=["store_name", "shift_name"],  # missing position_name, recurrence, item_title
        )
        db = make_mock_db()
        with pytest.raises(ValueError, match="Missing required columns"):
            await service.import_from_excel(db, org_id, bad_excel)

    @pytest.mark.asyncio
    async def test_empty_data_raises(self, service, org_id):
        """데이터 행이 없으면 ValueError."""
        empty_excel = make_excel([])
        db = make_mock_db()
        with pytest.raises(ValueError, match="No data rows"):
            await service.import_from_excel(db, org_id, empty_excel)

    @pytest.mark.asyncio
    async def test_skip_blank_rows(self, service, org_id):
        """빈 행(store_name 등이 비어 있음)은 건너뛰기."""
        excel = make_excel([
            ["", "Morning", "Grill", "daily", "Item1", "", "photo"],  # no store_name
            ["Store1", "Morning", "Grill", "daily", "", "", "photo"],  # no item_title
        ])
        db = make_mock_db()
        with pytest.raises(ValueError, match="No data rows"):
            await service.import_from_excel(db, org_id, excel)

    @pytest.mark.asyncio
    async def test_basic_import_flow(self, service, org_id):
        """기본 import 플로우: store/shift/position 생성 + 템플릿/항목 생성."""
        excel = make_excel([
            ["LA Downtown", "Morning", "Grill", "daily", "Preheat grill", "Heat to 400F", "photo"],
            ["LA Downtown", "Morning", "Grill", "daily", "Check oil", "", "text"],
        ])

        # Mock store/shift/position creation
        mock_store = MagicMock(id=uuid4())
        mock_shift = MagicMock(id=uuid4())
        mock_position = MagicMock(id=uuid4())
        mock_template = MagicMock(id=uuid4())

        with patch.object(service, "_get_or_create_store", return_value=(mock_store, True)) as m_store, \
             patch.object(service, "_get_or_create_shift", return_value=(mock_shift, True)) as m_shift, \
             patch.object(service, "_get_or_create_position", return_value=(mock_position, True)) as m_pos, \
             patch("app.services.checklist_service.checklist_repository") as mock_repo:

            mock_repo.check_duplicate = AsyncMock(return_value=False)
            mock_repo.create = AsyncMock(return_value=mock_template)
            mock_repo.create_items_bulk = AsyncMock(return_value=[])

            db = make_mock_db()
            result = await service.import_from_excel(db, org_id, excel)

            assert result["created_stores"] == 1
            assert result["created_shifts"] == 1
            assert result["created_positions"] == 1
            assert result["created_templates"] == 1
            assert result["created_items"] == 2
            assert result["errors"] == []

            # 템플릿 생성 확인 — 제목 형식: "{store} - {shift} - {position}"
            mock_repo.create.assert_called_once()
            create_args = mock_repo.create.call_args[0]
            assert create_args[1]["title"] == "LA Downtown - Morning - Grill"

            # 항목 생성 확인 — recurrence는 item 레벨에 저장
            mock_repo.create_items_bulk.assert_called_once()
            items = mock_repo.create_items_bulk.call_args[0][1]
            assert len(items) == 2
            assert items[0]["title"] == "Preheat grill"
            assert items[0]["verification_type"] == "photo"
            assert items[0]["recurrence_type"] == "daily"
            assert items[0]["recurrence_days"] is None
            assert items[1]["title"] == "Check oil"
            assert items[1]["verification_type"] == "text"
            assert items[1]["recurrence_type"] == "daily"
            assert items[1]["recurrence_days"] is None

    @pytest.mark.asyncio
    async def test_multi_type_verification_import(self, service, org_id):
        """photo,text 같은 multi-type verification이 올바르게 파싱되는지 확인."""
        excel = make_excel([
            ["Store1", "Morning", "Grill", "daily", "Item1", "", "photo,text"],
            ["Store1", "Morning", "Grill", "daily", "Item2", "", ""],
            ["Store1", "Morning", "Grill", "daily", "Item3", "", "none"],
        ])

        mock_store = MagicMock(id=uuid4())
        mock_shift = MagicMock(id=uuid4())
        mock_position = MagicMock(id=uuid4())
        mock_template = MagicMock(id=uuid4())

        with patch.object(service, "_get_or_create_store", return_value=(mock_store, False)), \
             patch.object(service, "_get_or_create_shift", return_value=(mock_shift, False)), \
             patch.object(service, "_get_or_create_position", return_value=(mock_position, False)), \
             patch("app.services.checklist_service.checklist_repository") as mock_repo:

            mock_repo.check_duplicate = AsyncMock(return_value=False)
            mock_repo.create = AsyncMock(return_value=mock_template)
            mock_repo.create_items_bulk = AsyncMock(return_value=[])

            db = make_mock_db()
            result = await service.import_from_excel(db, org_id, excel)

            assert result["created_items"] == 3

            items = mock_repo.create_items_bulk.call_args[0][1]
            assert items[0]["verification_type"] == "photo,text"
            assert items[1]["verification_type"] == "none"  # empty → none
            assert items[2]["verification_type"] == "none"  # explicit none

    @pytest.mark.asyncio
    async def test_weekly_recurrence_import(self, service, org_id):
        """weekly recurrence 파싱 확인."""
        excel = make_excel([
            ["Store1", "Morning", "Counter", "mon,wed,fri", "POS check", "", "text"],
        ])

        mock_store = MagicMock(id=uuid4())
        mock_shift = MagicMock(id=uuid4())
        mock_position = MagicMock(id=uuid4())
        mock_template = MagicMock(id=uuid4())

        with patch.object(service, "_get_or_create_store", return_value=(mock_store, False)), \
             patch.object(service, "_get_or_create_shift", return_value=(mock_shift, False)), \
             patch.object(service, "_get_or_create_position", return_value=(mock_position, False)), \
             patch("app.services.checklist_service.checklist_repository") as mock_repo:

            mock_repo.check_duplicate = AsyncMock(return_value=False)
            mock_repo.create = AsyncMock(return_value=mock_template)
            mock_repo.create_items_bulk = AsyncMock(return_value=[])

            db = make_mock_db()
            result = await service.import_from_excel(db, org_id, excel)

            # recurrence는 item 레벨에 저장
            items = mock_repo.create_items_bulk.call_args[0][1]
            assert items[0]["recurrence_type"] == "weekly"
            assert items[0]["recurrence_days"] == [0, 2, 4]

    @pytest.mark.asyncio
    async def test_duplicate_skip(self, service, org_id):
        """중복 시 skip 동작 확인."""
        excel = make_excel([
            ["Store1", "Morning", "Grill", "daily", "Item1", "", "photo"],
        ])

        mock_store = MagicMock(id=uuid4())
        mock_shift = MagicMock(id=uuid4())
        mock_position = MagicMock(id=uuid4())

        with patch.object(service, "_get_or_create_store", return_value=(mock_store, False)), \
             patch.object(service, "_get_or_create_shift", return_value=(mock_shift, False)), \
             patch.object(service, "_get_or_create_position", return_value=(mock_position, False)), \
             patch("app.services.checklist_service.checklist_repository") as mock_repo:

            mock_repo.check_duplicate = AsyncMock(return_value=True)

            db = make_mock_db()
            result = await service.import_from_excel(db, org_id, excel, duplicate_action="skip")

            assert result["skipped_templates"] == 1
            assert result["created_templates"] == 0
            mock_repo.create.assert_not_called()

    @pytest.mark.asyncio
    async def test_invalid_recurrence_recorded_as_error(self, service, org_id):
        """잘못된 recurrence가 에러로 기록되는지 확인."""
        excel = make_excel([
            ["Store1", "Morning", "Grill", "badday", "Item1", "", "photo"],
            ["Store1", "Morning", "Grill", "daily", "Item2", "", "text"],
        ])

        mock_store = MagicMock(id=uuid4())
        mock_shift = MagicMock(id=uuid4())
        mock_position = MagicMock(id=uuid4())
        mock_template = MagicMock(id=uuid4())

        with patch.object(service, "_get_or_create_store", return_value=(mock_store, False)), \
             patch.object(service, "_get_or_create_shift", return_value=(mock_shift, False)), \
             patch.object(service, "_get_or_create_position", return_value=(mock_position, False)), \
             patch("app.services.checklist_service.checklist_repository") as mock_repo:

            mock_repo.check_duplicate = AsyncMock(return_value=False)
            mock_repo.create = AsyncMock(return_value=mock_template)
            mock_repo.create_items_bulk = AsyncMock(return_value=[])

            db = make_mock_db()
            result = await service.import_from_excel(db, org_id, excel)

            assert len(result["errors"]) == 1
            assert "Invalid day" in result["errors"][0]
            assert result["created_items"] == 1  # only Item2

    @pytest.mark.asyncio
    async def test_sample_excel_roundtrip(self, service, org_id):
        """generate_sample_excel → import_from_excel 라운드트립 테스트."""
        sample_bytes = ChecklistService.generate_sample_excel()

        mock_store = MagicMock(id=uuid4())
        mock_shift = MagicMock(id=uuid4())
        mock_position = MagicMock(id=uuid4())
        mock_template = MagicMock(id=uuid4())

        with patch.object(service, "_get_or_create_store", return_value=(mock_store, True)), \
             patch.object(service, "_get_or_create_shift", return_value=(mock_shift, True)), \
             patch.object(service, "_get_or_create_position", return_value=(mock_position, True)), \
             patch("app.services.checklist_service.checklist_repository") as mock_repo:

            mock_repo.check_duplicate = AsyncMock(return_value=False)
            mock_repo.create = AsyncMock(return_value=mock_template)
            mock_repo.create_items_bulk = AsyncMock(return_value=[])

            db = make_mock_db()
            result = await service.import_from_excel(db, org_id, sample_bytes)

            # 에러 없이 임포트되어야 함
            assert result["errors"] == [], f"Import errors: {result['errors']}"
            assert result["created_templates"] > 0
            assert result["created_items"] > 0
            print(f"\n✅ Roundtrip result: {result}")

    @pytest.mark.asyncio
    async def test_multiple_stores_grouped(self, service, org_id):
        """여러 매장의 데이터가 올바르게 그룹화되는지 확인."""
        excel = make_excel([
            ["Store1", "Morning", "Grill", "daily", "Item1", "", "photo"],
            ["Store1", "Morning", "Grill", "daily", "Item2", "", "text"],
            ["Store2", "Evening", "Counter", "mon,fri", "Item3", "", "photo,text"],
        ])

        store1 = MagicMock(id=uuid4())
        store2 = MagicMock(id=uuid4())
        mock_shift = MagicMock(id=uuid4())
        mock_position = MagicMock(id=uuid4())
        mock_template = MagicMock(id=uuid4())

        call_count = {"store": 0}

        async def fake_get_or_create_store(db, org_id, name):
            call_count["store"] += 1
            if name == "Store1":
                return (store1, True)
            return (store2, True)

        with patch.object(service, "_get_or_create_store", side_effect=fake_get_or_create_store), \
             patch.object(service, "_get_or_create_shift", return_value=(mock_shift, True)), \
             patch.object(service, "_get_or_create_position", return_value=(mock_position, True)), \
             patch("app.services.checklist_service.checklist_repository") as mock_repo:

            mock_repo.check_duplicate = AsyncMock(return_value=False)
            mock_repo.create = AsyncMock(return_value=mock_template)
            mock_repo.create_items_bulk = AsyncMock(return_value=[])

            db = make_mock_db()
            result = await service.import_from_excel(db, org_id, excel)

            assert result["created_stores"] == 2
            assert result["created_templates"] == 2
            assert result["created_items"] == 3
