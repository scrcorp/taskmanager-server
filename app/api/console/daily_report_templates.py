from typing import Annotated
from uuid import UUID

from fastapi import APIRouter, Depends, Query, UploadFile, File, Form
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import require_permission
from app.database import get_db
from app.models.user import User
from app.schemas.daily_report import (
    DailyReportTemplateCreate,
    DailyReportTemplateResponse,
    DailyReportTemplateSectionInput,
    DailyReportTemplateSectionResponse,
    DailyReportTemplateUpdate,
)
from app.services.daily_report_service import daily_report_service
from app.utils.exceptions import BadRequestError

router: APIRouter = APIRouter()


def _build_template_response(template) -> dict:
    return {
        "id": str(template.id),
        "organization_id": str(template.organization_id) if template.organization_id else None,
        "store_id": str(template.store_id) if template.store_id else None,
        "name": template.name,
        "is_default": template.is_default,
        "is_active": template.is_active,
        "created_at": template.created_at,
        "sections": [
            {
                "id": str(s.id),
                "title": s.title,
                "description": s.description,
                "sort_order": s.sort_order,
                "is_required": s.is_required,
            }
            for s in template.sections
        ],
    }


@router.get("", response_model=list[DailyReportTemplateResponse])
async def list_templates(
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permission("daily_reports:read"))],
    store_id: Annotated[str | None, Query()] = None,
    is_active: Annotated[bool | None, Query()] = None,
) -> list[dict]:
    templates = await daily_report_service.list_templates(
        db,
        organization_id=current_user.organization_id,
        store_id=UUID(store_id) if store_id else None,
        is_active=is_active,
    )
    return [_build_template_response(t) for t in templates]


@router.post("", response_model=DailyReportTemplateResponse, status_code=201)
async def create_template(
    data: DailyReportTemplateCreate,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permission("daily_reports:create"))],
) -> dict:
    template = await daily_report_service.create_template(
        db, current_user.organization_id, data
    )
    return _build_template_response(template)


# --- Excel routes (MUST be before /{template_id} to avoid path conflicts) ---


@router.get("/excel/sample")
async def download_sample_excel() -> FileResponse:
    """Download sample Excel template for bulk section upload."""
    from pathlib import Path
    sample_path = Path(__file__).resolve().parent.parent.parent.parent / "static" / "daily_report_template_sample.xlsx"
    return FileResponse(
        path=str(sample_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="daily_report_template_sample.xlsx",
    )


@router.post("/upload-excel", response_model=DailyReportTemplateResponse, status_code=201)
async def create_template_from_excel(
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permission("daily_reports:create"))],
    file: UploadFile = File(...),
    name: str = Form(...),
    store_id: str | None = Form(None),
) -> dict:
    """Create a template by uploading an Excel file.

    Excel format: Title | Description | Required (Y/N)
    First row is header, data starts from row 2.
    sort_order is auto-assigned by row order.
    """
    from openpyxl import load_workbook
    import io

    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise BadRequestError("Only .xlsx files are supported")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True)
    except Exception:
        raise BadRequestError("Invalid Excel file")

    ws = wb.active
    sections: list[DailyReportTemplateSectionInput] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        title = str(row[0]).strip()
        description = str(row[1]).strip() if len(row) > 1 and row[1] else None
        is_required = str(row[2]).strip().upper() in ("Y", "YES", "TRUE", "1") if len(row) > 2 and row[2] else False
        sections.append(DailyReportTemplateSectionInput(
            title=title, description=description, is_required=is_required,
        ))

    if not sections:
        raise BadRequestError("No valid sections found in the Excel file")

    data = DailyReportTemplateCreate(
        name=name,
        store_id=store_id,
        sections=sections,
    )
    template = await daily_report_service.create_template(
        db, current_user.organization_id, data
    )
    return _build_template_response(template)


# --- Template detail routes ---


@router.get("/{template_id}", response_model=DailyReportTemplateResponse)
async def get_template(
    template_id: UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permission("daily_reports:read"))],
) -> dict:
    template = await daily_report_service.get_template_detail(
        db, template_id, current_user.organization_id
    )
    return _build_template_response(template)


@router.put("/{template_id}", response_model=DailyReportTemplateResponse)
async def update_template(
    template_id: UUID,
    data: DailyReportTemplateUpdate,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permission("daily_reports:update"))],
) -> dict:
    template = await daily_report_service.update_template(
        db, template_id, current_user.organization_id, data
    )
    return _build_template_response(template)


@router.delete("/{template_id}", status_code=204)
async def delete_template(
    template_id: UUID,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(require_permission("daily_reports:delete"))],
) -> None:
    await daily_report_service.delete_template(
        db, template_id, current_user.organization_id
    )
