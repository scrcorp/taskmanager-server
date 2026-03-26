"""체크리스트 서비스 — 체크리스트 템플릿/항목 비즈니스 로직.

Checklist Service — Business logic for checklist template and item management.
Handles template CRUD, item CRUD, reordering, store ownership validation,
and Excel import/export for bulk template creation.
"""

from io import BytesIO
from typing import Sequence
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.checklist import ChecklistTemplate, ChecklistTemplateItem
from app.models.organization import Store
from app.models.schedule import StoreWorkRole
from app.models.work import Position, Shift
from app.repositories.checklist_repository import checklist_repository
from app.schemas.common import (
    ChecklistBulkItemCreate,
    ChecklistItemCreate,
    ChecklistItemUpdate,
    ChecklistTemplateCreate,
    ChecklistTemplateUpdate,
)
from app.utils.exceptions import DuplicateError, ForbiddenError, NotFoundError

# 요일 약어 → 숫자 매핑 (Monday=0 ~ Sunday=6, Python weekday() 기준)
DAY_MAP: dict[str, int] = {
    "mon": 0, "tue": 1, "wed": 2, "thu": 3,
    "fri": 4, "sat": 5, "sun": 6,
}

# Excel 필수 컬럼 목록
REQUIRED_COLUMNS: list[str] = [
    "store", "shift", "position", "recurrence", "item_title",
]


class ChecklistService:
    """체크리스트 서비스.

    Checklist service providing template and item business logic.

    Methods handle validation, authorization, and delegate DB operations
    to the checklist repository.
    """

    @staticmethod
    def _normalize_recurrence(
        recurrence_type: str | None,
        recurrence_days: list[int] | None,
    ) -> tuple[str, list[int] | None]:
        """recurrence_days 기반으로 recurrence_type을 자동 결정합니다.

        Derive recurrence_type from recurrence_days as a backend safety net.
        - None / empty / all 7 days → ("daily", None)
        - Otherwise → ("weekly", sorted days)
        """
        if not recurrence_days or len(recurrence_days) == 0 or set(recurrence_days) == {0, 1, 2, 3, 4, 5, 6}:
            return "daily", None
        return "weekly", sorted(recurrence_days)

    async def _validate_store_ownership(
        self,
        db: AsyncSession,
        store_id: UUID,
        organization_id: UUID,
    ) -> Store:
        """매장이 해당 조직에 속하는지 검증합니다.

        Verify that a store belongs to the specified organization.

        Args:
            db: 비동기 데이터베이스 세션 (Async database session)
            store_id: 매장 UUID (Store UUID)
            organization_id: 조직 UUID (Organization UUID)

        Returns:
            Store: 검증된 매장 (Verified store)

        Raises:
            NotFoundError: 매장이 없을 때 (When store not found)
            ForbiddenError: 다른 조직의 매장일 때 (When store belongs to another org)
        """
        result = await db.execute(select(Store).where(Store.id == store_id))
        store: Store | None = result.scalar_one_or_none()

        if store is None:
            raise NotFoundError("Store not found")
        if store.organization_id != organization_id:
            raise ForbiddenError("No permission for this store")

        return store

    # --- 템플릿 CRUD (Template CRUD) ---

    async def list_all_templates(
        self,
        db: AsyncSession,
        organization_id: UUID,
        store_id: UUID | None = None,
        shift_id: UUID | None = None,
        position_id: UUID | None = None,
    ) -> Sequence[ChecklistTemplate]:
        """조직 전체의 체크리스트 템플릿 목록을 조회합니다.

        List all checklist templates for an organization with optional filters.

        Args:
            db: 비동기 데이터베이스 세션 (Async database session)
            organization_id: 조직 UUID (Organization UUID)
            store_id: 매장 UUID 필터, 선택 (Optional store UUID filter)
            shift_id: 근무조 UUID 필터, 선택 (Optional shift UUID filter)
            position_id: 포지션 UUID 필터, 선택 (Optional position UUID filter)

        Returns:
            Sequence[ChecklistTemplate]: 템플릿 목록 (List of templates)
        """
        return await checklist_repository.get_all_by_org(
            db, organization_id, store_id, shift_id, position_id
        )

    async def list_templates(
        self,
        db: AsyncSession,
        store_id: UUID,
        organization_id: UUID,
        shift_id: UUID | None = None,
        position_id: UUID | None = None,
    ) -> Sequence[ChecklistTemplate]:
        """매장의 체크리스트 템플릿 목록을 조회합니다.

        List checklist templates for a store with optional filters.

        Args:
            db: 비동기 데이터베이스 세션 (Async database session)
            store_id: 매장 UUID (Store UUID)
            organization_id: 조직 UUID (Organization UUID)
            shift_id: 근무조 UUID 필터, 선택 (Optional shift UUID filter)
            position_id: 포지션 UUID 필터, 선택 (Optional position UUID filter)

        Returns:
            Sequence[ChecklistTemplate]: 템플릿 목록 (List of templates)

        Raises:
            NotFoundError: 매장이 없을 때 (When store not found)
            ForbiddenError: 다른 조직 매장일 때 (When store belongs to another org)
        """
        await self._validate_store_ownership(db, store_id, organization_id)
        return await checklist_repository.get_by_store(db, store_id, shift_id, position_id)

    async def get_template_detail(
        self,
        db: AsyncSession,
        template_id: UUID,
        organization_id: UUID,
    ) -> ChecklistTemplate:
        """체크리스트 템플릿 상세를 항목과 함께 조회합니다.

        Get checklist template detail with items.

        Args:
            db: 비동기 데이터베이스 세션 (Async database session)
            template_id: 템플릿 UUID (Template UUID)
            organization_id: 조직 UUID (Organization UUID)

        Returns:
            ChecklistTemplate: 항목 포함 템플릿 상세 (Template detail with items)

        Raises:
            NotFoundError: 템플릿이 없을 때 (When template not found)
            ForbiddenError: 다른 조직 매장일 때 (When store belongs to another org)
        """
        template: ChecklistTemplate | None = await checklist_repository.get_with_items(
            db, template_id
        )
        if template is None:
            raise NotFoundError("Checklist template not found")

        # 조직 소유권 검증 — Verify org ownership via store
        await self._validate_store_ownership(db, template.store_id, organization_id)
        return template

    async def create_template(
        self,
        db: AsyncSession,
        store_id: UUID,
        organization_id: UUID,
        data: ChecklistTemplateCreate,
    ) -> ChecklistTemplate:
        """새 체크리스트 템플릿을 생성합니다.

        Create a new checklist template (unique store+shift+position).

        Args:
            db: 비동기 데이터베이스 세션 (Async database session)
            store_id: 매장 UUID (Store UUID)
            organization_id: 조직 UUID (Organization UUID)
            data: 템플릿 생성 데이터 (Template creation data)

        Returns:
            ChecklistTemplate: 생성된 템플릿 (Created template)

        Raises:
            NotFoundError: 매장이 없을 때 (When store not found)
            ForbiddenError: 다른 조직 매장일 때 (When store belongs to another org)
            DuplicateError: 동일 조합 존재 시 (When same combination already exists)
        """
        store: Store = await self._validate_store_ownership(db, store_id, organization_id)

        shift_id: UUID = UUID(data.shift_id)
        position_id: UUID = UUID(data.position_id)

        # 중복 검사 — Check for duplicate combination
        is_duplicate: bool = await checklist_repository.check_duplicate(
            db, store_id, shift_id, position_id
        )
        if is_duplicate:
            raise DuplicateError(
                "해당 매장+근무조+포지션 조합의 템플릿이 이미 존재합니다 "
                "(Template for this store+shift+position combination already exists)"
            )

        # Shift/Position 이름 조회 — Lookup names for title generation
        shift_result = await db.execute(select(Shift).where(Shift.id == shift_id))
        shift: Shift = shift_result.scalar_one()
        position_result = await db.execute(select(Position).where(Position.id == position_id))
        position: Position = position_result.scalar_one()

        # 제목 자동 생성 — Auto-generate title: '{store} - {shift} - {position} [(title)]'
        base_title: str = f"{store.name} - {shift.name} - {position.name}"
        extra: str = data.title.strip() if data.title else ""
        template_title: str = f"{base_title} ({extra})" if extra else base_title

        try:
            template: ChecklistTemplate = await checklist_repository.create(
                db,
                {
                    "store_id": store_id,
                    "shift_id": shift_id,
                    "position_id": position_id,
                    "title": template_title,
                },
            )
            # 매칭되는 work_role에 자동 연결 — Auto-link to matching work_role
            await self._auto_link_work_role(db, template)
            await db.commit()
            return template
        except Exception:
            await db.rollback()
            raise

    async def update_template(
        self,
        db: AsyncSession,
        template_id: UUID,
        organization_id: UUID,
        data: ChecklistTemplateUpdate,
    ) -> ChecklistTemplate:
        """체크리스트 템플릿을 업데이트합니다.

        Update a checklist template (title, shift_id, position_id).

        Args:
            db: 비동기 데이터베이스 세션 (Async database session)
            template_id: 템플릿 UUID (Template UUID)
            organization_id: 조직 UUID (Organization UUID)
            data: 업데이트 데이터 (Update data with optional title/shift_id/position_id)

        Returns:
            ChecklistTemplate: 업데이트된 템플릿 (Updated template)

        Raises:
            NotFoundError: 템플릿이 없을 때 (When template not found)
            ForbiddenError: 다른 조직 매장일 때 (When store belongs to another org)
            DuplicateError: 동일 조합 존재 시 (When same combination already exists)
        """
        template: ChecklistTemplate | None = await checklist_repository.get_with_items(
            db, template_id
        )
        if template is None:
            raise NotFoundError("Checklist template not found")

        await self._validate_store_ownership(db, template.store_id, organization_id)

        update_fields: dict = {}
        if data.title is not None:
            update_fields["title"] = data.title

        new_shift_id: UUID = UUID(data.shift_id) if data.shift_id else template.shift_id
        new_position_id: UUID = UUID(data.position_id) if data.position_id else template.position_id

        # shift_id 또는 position_id가 변경되었으면 중복 검사 — Check duplicate if shift/position changed
        if new_shift_id != template.shift_id or new_position_id != template.position_id:
            is_duplicate: bool = await checklist_repository.check_duplicate(
                db, template.store_id, new_shift_id, new_position_id
            )
            if is_duplicate:
                raise DuplicateError(
                    "해당 매장+근무조+포지션 조합의 템플릿이 이미 존재합니다 "
                    "(Template for this store+shift+position combination already exists)"
                )
            if data.shift_id is not None:
                update_fields["shift_id"] = new_shift_id
            if data.position_id is not None:
                update_fields["position_id"] = new_position_id

        if not update_fields:
            return template

        try:
            updated: ChecklistTemplate | None = await checklist_repository.update(
                db, template_id, update_fields
            )
            if updated is None:
                raise NotFoundError("Checklist template not found")
            # shift/position 변경 시 새 조합에 매칭되는 work_role에 자동 연결
            # Auto-link to matching work_role when shift/position changed
            await self._auto_link_work_role(db, updated)
            await db.commit()
            return updated
        except Exception:
            await db.rollback()
            raise

    async def delete_template(
        self,
        db: AsyncSession,
        template_id: UUID,
        organization_id: UUID,
    ) -> bool:
        """체크리스트 템플릿을 삭제합니다 (cascade로 항목도 삭제).

        Delete a checklist template (items are cascade-deleted).

        Args:
            db: 비동기 데이터베이스 세션 (Async database session)
            template_id: 템플릿 UUID (Template UUID)
            organization_id: 조직 UUID (Organization UUID)

        Returns:
            bool: 삭제 성공 여부 (Whether the deletion was successful)

        Raises:
            NotFoundError: 템플릿이 없을 때 (When template not found)
            ForbiddenError: 다른 조직 매장일 때 (When store belongs to another org)
        """
        template: ChecklistTemplate | None = await checklist_repository.get_with_items(
            db, template_id
        )
        if template is None:
            raise NotFoundError("Checklist template not found")

        await self._validate_store_ownership(db, template.store_id, organization_id)
        try:
            result = await checklist_repository.delete(db, template_id)
            await db.commit()
            return result
        except Exception:
            await db.rollback()
            raise

    # --- 내부 헬퍼 (Internal helpers) ---

    @staticmethod
    async def _auto_link_work_role(
        db: AsyncSession,
        template: ChecklistTemplate,
    ) -> None:
        """체크리스트 템플릿 생성/수정 시 매칭되는 work_role에 자동 연결합니다.

        store_id+shift_id+position_id가 일치하는 work_role 중 default_checklist_id가
        null인 것을 찾아 이 템플릿 ID로 설정합니다.
        (이미 다른 템플릿이 연결된 경우는 변경하지 않음)

        Auto-link this template to a matching work_role that has no default_checklist_id set.
        Matching is by store_id + shift_id + position_id.
        Does not overwrite existing links.
        """
        result = await db.execute(
            select(StoreWorkRole).where(
                StoreWorkRole.store_id == template.store_id,
                StoreWorkRole.shift_id == template.shift_id,
                StoreWorkRole.position_id == template.position_id,
                StoreWorkRole.default_checklist_id.is_(None),
            )
        )
        work_role: StoreWorkRole | None = result.scalar_one_or_none()
        if work_role is not None:
            work_role.default_checklist_id = template.id
            await db.flush()

    # --- 항목 CRUD (Item CRUD) ---

    async def list_items(
        self,
        db: AsyncSession,
        template_id: UUID,
        organization_id: UUID,
    ) -> Sequence[ChecklistTemplateItem]:
        """템플릿의 항목 목록을 조회합니다.

        List items for a checklist template.

        Args:
            db: 비동기 데이터베이스 세션 (Async database session)
            template_id: 템플릿 UUID (Template UUID)
            organization_id: 조직 UUID (Organization UUID)

        Returns:
            Sequence[ChecklistTemplateItem]: 정렬된 항목 목록 (Sorted item list)

        Raises:
            NotFoundError: 템플릿이 없을 때 (When template not found)
        """
        # 템플릿 존재 및 소유권 검증 — Verify template exists and ownership
        await self.get_template_detail(db, template_id, organization_id)
        return await checklist_repository.get_items(db, template_id)

    async def add_item(
        self,
        db: AsyncSession,
        template_id: UUID,
        organization_id: UUID,
        data: ChecklistItemCreate,
    ) -> ChecklistTemplateItem:
        """템플릿에 새 항목을 추가합니다.

        Add a new item to a checklist template.

        Args:
            db: 비동기 데이터베이스 세션 (Async database session)
            template_id: 템플릿 UUID (Template UUID)
            organization_id: 조직 UUID (Organization UUID)
            data: 항목 생성 데이터 (Item creation data)

        Returns:
            ChecklistTemplateItem: 생성된 항목 (Created item)

        Raises:
            NotFoundError: 템플릿이 없을 때 (When template not found)
        """
        await self.get_template_detail(db, template_id, organization_id)

        rec_type, rec_days = self._normalize_recurrence(data.recurrence_type, data.recurrence_days)
        try:
            item: ChecklistTemplateItem = await checklist_repository.create_item(
                db,
                {
                    "template_id": template_id,
                    "title": data.title,
                    "description": data.description,
                    "verification_type": data.verification_type,
                    "min_photos": data.min_photos if "photo" in (data.verification_type or "") else 0,
                    "recurrence_type": rec_type,
                    "recurrence_days": rec_days,
                    "sort_order": data.sort_order,
                },
            )
            await db.commit()
            return item
        except Exception:
            await db.rollback()
            raise

    async def add_items_bulk(
        self,
        db: AsyncSession,
        template_id: UUID,
        organization_id: UUID,
        data: ChecklistBulkItemCreate,
    ) -> list[ChecklistTemplateItem]:
        """템플릿에 여러 항목을 일괄 추가합니다.

        Bulk-add multiple items to a checklist template in a single transaction.

        Args:
            db: 비동기 데이터베이스 세션 (Async database session)
            template_id: 템플릿 UUID (Template UUID)
            organization_id: 조직 UUID (Organization UUID)
            data: 일괄 생성 데이터 (Bulk creation data)

        Returns:
            list[ChecklistTemplateItem]: 생성된 항목 목록 (List of created items)

        Raises:
            NotFoundError: 템플릿이 없을 때 (When template not found)
        """
        await self.get_template_detail(db, template_id, organization_id)

        items_data: list[dict] = []
        for item in data.items:
            rec_type, rec_days = self._normalize_recurrence(item.recurrence_type, item.recurrence_days)
            items_data.append(
                {
                    "template_id": template_id,
                    "title": item.title,
                    "description": item.description,
                    "verification_type": item.verification_type,
                    "min_photos": item.min_photos if "photo" in (item.verification_type or "") else 0,
                    "recurrence_type": rec_type,
                    "recurrence_days": rec_days,
                    "sort_order": item.sort_order,
                }
            )

        try:
            result = await checklist_repository.create_items_bulk(db, items_data)
            await db.commit()
            return result
        except Exception:
            await db.rollback()
            raise

    async def update_item(
        self,
        db: AsyncSession,
        item_id: UUID,
        organization_id: UUID,
        data: ChecklistItemUpdate,
    ) -> ChecklistTemplateItem:
        """체크리스트 항목을 업데이트합니다.

        Update a checklist template item.

        Args:
            db: 비동기 데이터베이스 세션 (Async database session)
            item_id: 항목 UUID (Item UUID)
            organization_id: 조직 UUID (Organization UUID)
            data: 항목 업데이트 데이터 (Item update data)

        Returns:
            ChecklistTemplateItem: 업데이트된 항목 (Updated item)

        Raises:
            NotFoundError: 항목이 없을 때 (When item not found)
            ForbiddenError: 다른 조직의 항목일 때 (When item belongs to another org)
        """
        item: ChecklistTemplateItem | None = await checklist_repository.get_item_by_id(
            db, item_id
        )
        if item is None:
            raise NotFoundError("Checklist item not found")

        # 소유권 검증 — Verify ownership via template's store
        template: ChecklistTemplate | None = await checklist_repository.get_with_items(
            db, item.template_id
        )
        if template is None:
            raise NotFoundError("Checklist template not found")
        await self._validate_store_ownership(db, template.store_id, organization_id)

        # None이 아닌 필드만 업데이트 — Only update non-None fields
        update_data: dict = data.model_dump(exclude_unset=True)

        # recurrence_days가 넘어왔으면 recurrence_type도 자동 결정
        if "recurrence_days" in update_data or "recurrence_type" in update_data:
            rec_type_in = update_data.get("recurrence_type", item.recurrence_type)
            rec_days_in = update_data.get("recurrence_days", item.recurrence_days)
            rec_type, rec_days = self._normalize_recurrence(rec_type_in, rec_days_in)
            update_data["recurrence_type"] = rec_type
            update_data["recurrence_days"] = rec_days

        try:
            updated: ChecklistTemplateItem | None = await checklist_repository.update_item(
                db, item_id, update_data
            )
            if updated is None:
                raise NotFoundError("Checklist item not found")
            await db.commit()
            return updated
        except Exception:
            await db.rollback()
            raise

    async def reorder_items(
        self,
        db: AsyncSession,
        template_id: UUID,
        organization_id: UUID,
        item_ids: list[str],
    ) -> None:
        """항목의 정렬 순서를 재배치합니다.

        Reorder items within a template.

        Args:
            db: 비동기 데이터베이스 세션 (Async database session)
            template_id: 템플릿 UUID (Template UUID)
            organization_id: 조직 UUID (Organization UUID)
            item_ids: 새 순서대로 정렬된 항목 ID 문자열 목록
                      (List of item ID strings in the desired order)

        Raises:
            NotFoundError: 템플릿이 없을 때 (When template not found)
        """
        await self.get_template_detail(db, template_id, organization_id)
        uuid_ids: list[UUID] = [UUID(id_str) for id_str in item_ids]
        await checklist_repository.reorder_items(db, template_id, uuid_ids)

    async def reorder_items_by_item_id(
        self,
        db: AsyncSession,
        item_id: UUID,
        organization_id: UUID,
        item_ids: list[str],
    ) -> None:
        """항목 ID로부터 template_id를 추출하여 재배치합니다.

        Resolve template_id from an item reference and reorder items.
        Encapsulates repository access that was previously done in the router.

        Args:
            db: 비동기 데이터베이스 세션 (Async database session)
            item_id: 기준 항목 UUID (Reference item UUID to resolve template)
            organization_id: 조직 UUID (Organization UUID)
            item_ids: 새 순서대로 정렬된 항목 ID 문자열 목록
                      (List of item ID strings in the desired order)

        Raises:
            NotFoundError: 항목이 없을 때 (When item not found)
        """
        item: ChecklistTemplateItem | None = await checklist_repository.get_item_by_id(
            db, item_id
        )
        if item is None:
            raise NotFoundError("Checklist item not found")

        try:
            await self.reorder_items(db, item.template_id, organization_id, item_ids)
            await db.commit()
        except Exception:
            await db.rollback()
            raise

    async def delete_item(
        self,
        db: AsyncSession,
        item_id: UUID,
        organization_id: UUID,
    ) -> bool:
        """체크리스트 항목을 삭제합니다.

        Delete a checklist template item.

        Args:
            db: 비동기 데이터베이스 세션 (Async database session)
            item_id: 항목 UUID (Item UUID)
            organization_id: 조직 UUID (Organization UUID)

        Returns:
            bool: 삭제 성공 여부 (Whether the deletion was successful)

        Raises:
            NotFoundError: 항목이 없을 때 (When item not found)
            ForbiddenError: 다른 조직의 항목일 때 (When item belongs to another org)
        """
        item: ChecklistTemplateItem | None = await checklist_repository.get_item_by_id(
            db, item_id
        )
        if item is None:
            raise NotFoundError("Checklist item not found")

        template: ChecklistTemplate | None = await checklist_repository.get_with_items(
            db, item.template_id
        )
        if template is None:
            raise NotFoundError("Checklist template not found")
        await self._validate_store_ownership(db, template.store_id, organization_id)

        try:
            result = await checklist_repository.delete_item(db, item_id)
            await db.commit()
            return result
        except Exception:
            await db.rollback()
            raise

    # --- Excel Import/Export ---

    @staticmethod
    def _parse_recurrence(value: str) -> tuple[str, list[int] | None]:
        """Excel recurrence 값을 (recurrence_type, recurrence_days)로 변환.

        Args:
            value: "daily" 또는 "mon,wed,fri" 형식의 문자열

        Returns:
            ("daily", None) 또는 ("weekly", [0, 2, 4])

        Raises:
            ValueError: 잘못된 요일 약어가 있을 때
        """
        value = value.strip().lower()
        if value == "daily":
            return ("daily", None)

        days: list[int] = []
        for day_str in value.split(","):
            day_str = day_str.strip()
            if day_str not in DAY_MAP:
                raise ValueError(
                    f"Invalid day: '{day_str}'. Use: mon,tue,wed,thu,fri,sat,sun"
                )
            days.append(DAY_MAP[day_str])

        if not days:
            raise ValueError("At least one day must be specified for weekly recurrence")

        unique_days = sorted(set(days))
        # 7일 전부 선택 = daily
        if len(unique_days) == 7:
            return ("daily", None)
        return ("weekly", unique_days)

    async def _get_or_create_store(
        self,
        db: AsyncSession,
        organization_id: UUID,
        store_name: str,
    ) -> tuple[Store, bool]:
        """Store를 이름으로 조회하고, 없으면 생성."""
        result = await db.execute(
            select(Store).where(
                Store.organization_id == organization_id,
                Store.name == store_name,
            )
        )
        store: Store | None = result.scalar_one_or_none()
        if store is not None:
            return (store, False)

        store = Store(organization_id=organization_id, name=store_name)
        db.add(store)
        await db.flush()
        await db.refresh(store)
        return (store, True)

    async def _get_or_create_shift(
        self,
        db: AsyncSession,
        store_id: UUID,
        shift_name: str,
    ) -> tuple[Shift, bool]:
        """Shift를 이름으로 조회하고, 없으면 생성."""
        result = await db.execute(
            select(Shift).where(
                Shift.store_id == store_id,
                Shift.name == shift_name,
            )
        )
        shift: Shift | None = result.scalar_one_or_none()
        if shift is not None:
            return (shift, False)

        max_order_result = await db.execute(
            select(func.coalesce(func.max(Shift.sort_order), -1)).where(
                Shift.store_id == store_id
            )
        )
        next_order: int = (max_order_result.scalar() or 0) + 1

        shift = Shift(store_id=store_id, name=shift_name, sort_order=next_order)
        db.add(shift)
        await db.flush()
        await db.refresh(shift)
        return (shift, True)

    async def _get_or_create_position(
        self,
        db: AsyncSession,
        store_id: UUID,
        position_name: str,
    ) -> tuple[Position, bool]:
        """Position을 이름으로 조회하고, 없으면 생성."""
        result = await db.execute(
            select(Position).where(
                Position.store_id == store_id,
                Position.name == position_name,
            )
        )
        position: Position | None = result.scalar_one_or_none()
        if position is not None:
            return (position, False)

        max_order_result = await db.execute(
            select(func.coalesce(func.max(Position.sort_order), -1)).where(
                Position.store_id == store_id
            )
        )
        next_order: int = (max_order_result.scalar() or 0) + 1

        position = Position(
            store_id=store_id, name=position_name, sort_order=next_order
        )
        db.add(position)
        await db.flush()
        await db.refresh(position)
        return (position, True)

    async def import_from_excel(
        self,
        db: AsyncSession,
        organization_id: UUID,
        file_content: bytes,
        duplicate_action: str = "skip",
    ) -> dict:
        """Excel 파일에서 체크리스트 템플릿을 일괄 생성합니다.

        Args:
            db: 비동기 데이터베이스 세션
            organization_id: 조직 UUID
            file_content: Excel 파일 바이트
            duplicate_action: "skip" | "overwrite" | "append"

        Returns:
            dict: 생성 결과 통계
        """
        wb = load_workbook(filename=BytesIO(file_content), read_only=True)
        # Use "Checklist Template" sheet if present, otherwise fall back to
        # the first non-Guide sheet, then finally wb.active.
        ws = None
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() == "checklist template":
                ws = wb[sheet_name]
                break
        if ws is None:
            for sheet_name in wb.sheetnames:
                if sheet_name.lower() != "guide":
                    ws = wb[sheet_name]
                    break
        if ws is None:
            ws = wb.active

        # 헤더 검증 — Validate required columns
        headers: list[str] = [
            str(cell.value).strip().lower() if cell.value else ""
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        missing: list[str] = [c for c in REQUIRED_COLUMNS if c not in headers]
        if missing:
            wb.close()
            raise ValueError(f"Missing required columns: {', '.join(missing)}")

        col_idx: dict[str, int] = {name: i for i, name in enumerate(headers)}

        # 행 파싱 — Parse rows
        result: dict = {
            "created_templates": 0,
            "created_items": 0,
            "created_stores": 0,
            "created_shifts": 0,
            "created_positions": 0,
            "skipped_templates": 0,
            "updated_templates": 0,
            "errors": [],
        }

        # 그룹핑: (store_name, shift_name, position_name) → items (각 item에 recurrence 포함)
        groups: dict[tuple[str, str, str], dict] = {}

        for row_num, row in enumerate(ws.iter_rows(min_row=3), start=3):
            cells: list = list(row)

            def get_val(col_name: str) -> str:
                idx = col_idx.get(col_name)
                if idx is None or idx >= len(cells):
                    return ""
                val = cells[idx].value
                return str(val).strip() if val is not None else ""

            store_name: str = get_val("store")
            shift_name: str = get_val("shift")
            position_name: str = get_val("position")
            recurrence_raw: str = get_val("recurrence")
            item_title: str = get_val("item_title")

            # 빈 행 무시
            if not store_name or not shift_name or not position_name:
                continue
            if not item_title:
                continue

            # recurrence 파싱 — 각 item에 개별 적용
            try:
                rec_type, rec_days = self._parse_recurrence(
                    recurrence_raw if recurrence_raw else "daily"
                )
            except ValueError as e:
                result["errors"].append(f"Row {row_num}: {e}")
                continue

            key: tuple[str, str, str] = (store_name, shift_name, position_name)
            if key not in groups:
                groups[key] = {
                    "items": [],
                }

            item_desc: str = get_val("item_description")
            v_type_raw: str = get_val("verification_type").lower()
            # Support comma-separated multi-type (e.g. "photo,text")
            if v_type_raw:
                valid_types = {"none", "photo", "text", "video"}
                parts = [p.strip() for p in v_type_raw.split(",") if p.strip()]
                parts = [p for p in parts if p in valid_types]
                # Remove "none" if combined with others (e.g. "none,photo" → "photo")
                if len(parts) > 1:
                    parts = [p for p in parts if p != "none"]
                v_type = ",".join(parts) if parts else "none"
            else:
                v_type = "none"

            groups[key]["items"].append(
                {
                    "title": item_title,
                    "description": item_desc if item_desc else None,
                    "verification_type": v_type,
                    "min_photos": 1 if "photo" in v_type else 0,
                    "recurrence_type": rec_type,
                    "recurrence_days": rec_days,
                }
            )

        wb.close()

        if not groups:
            raise ValueError("No data rows found in the Excel file")

        # get_or_create 캐시 — 같은 이름 반복 조회 방지
        store_cache: dict[str, Store] = {}
        shift_cache: dict[tuple[UUID, str], Shift] = {}
        position_cache: dict[tuple[UUID, str], Position] = {}

        for (store_name, shift_name, position_name), group_data in groups.items():
            # Store
            if store_name in store_cache:
                store = store_cache[store_name]
            else:
                store, created = await self._get_or_create_store(
                    db, organization_id, store_name
                )
                store_cache[store_name] = store
                if created:
                    result["created_stores"] += 1

            # Shift
            shift_key = (store.id, shift_name)
            if shift_key in shift_cache:
                shift = shift_cache[shift_key]
            else:
                shift, created = await self._get_or_create_shift(
                    db, store.id, shift_name
                )
                shift_cache[shift_key] = shift
                if created:
                    result["created_shifts"] += 1

            # Position
            pos_key = (store.id, position_name)
            if pos_key in position_cache:
                position = position_cache[pos_key]
            else:
                position, created = await self._get_or_create_position(
                    db, store.id, position_name
                )
                position_cache[pos_key] = position
                if created:
                    result["created_positions"] += 1

            # Template 조회/처리
            is_dup: bool = await checklist_repository.check_duplicate(
                db, store.id, shift.id, position.id
            )

            if is_dup:
                if duplicate_action == "skip":
                    result["skipped_templates"] += 1
                    continue

                # 기존 template 조회
                existing_result = await db.execute(
                    select(ChecklistTemplate).where(
                        ChecklistTemplate.store_id == store.id,
                        ChecklistTemplate.shift_id == shift.id,
                        ChecklistTemplate.position_id == position.id,
                    )
                )
                existing_template: ChecklistTemplate = existing_result.scalar_one()

                if duplicate_action == "overwrite":
                    # 기존 items 전체 삭제
                    await db.execute(
                        delete(ChecklistTemplateItem).where(
                            ChecklistTemplateItem.template_id == existing_template.id
                        )
                    )
                    await db.flush()
                    start_order = 0
                else:
                    # append: 기존 max sort_order 이후부터
                    max_result = await db.execute(
                        select(
                            func.coalesce(
                                func.max(ChecklistTemplateItem.sort_order), -1
                            )
                        ).where(
                            ChecklistTemplateItem.template_id == existing_template.id
                        )
                    )
                    start_order = (max_result.scalar() or 0) + 1

                items_data: list[dict] = [
                    {
                        "template_id": existing_template.id,
                        "title": item["title"],
                        "description": item["description"],
                        "verification_type": item["verification_type"],
                        "min_photos": item.get("min_photos", 0),
                        "recurrence_type": item["recurrence_type"],
                        "recurrence_days": item["recurrence_days"],
                        "sort_order": start_order + i,
                    }
                    for i, item in enumerate(group_data["items"])
                ]
                await checklist_repository.create_items_bulk(db, items_data)
                result["updated_templates"] += 1
                result["created_items"] += len(items_data)
            else:
                # 새 Template 생성
                template_title: str = f"{store_name} - {shift_name} - {position_name}"
                template: ChecklistTemplate = await checklist_repository.create(
                    db,
                    {
                        "store_id": store.id,
                        "shift_id": shift.id,
                        "position_id": position.id,
                        "title": template_title,
                    },
                )

                items_data = [
                    {
                        "template_id": template.id,
                        "title": item["title"],
                        "description": item["description"],
                        "verification_type": item["verification_type"],
                        "min_photos": item.get("min_photos", 0),
                        "recurrence_type": item["recurrence_type"],
                        "recurrence_days": item["recurrence_days"],
                        "sort_order": i,
                    }
                    for i, item in enumerate(group_data["items"])
                ]
                await checklist_repository.create_items_bulk(db, items_data)
                result["created_templates"] += 1
                result["created_items"] += len(items_data)

        try:
            await db.commit()
        except Exception:
            await db.rollback()
            raise

        return result



# 싱글턴 인스턴스 — Singleton instance
checklist_service: ChecklistService = ChecklistService()
