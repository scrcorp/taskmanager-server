"""Payroll export 서비스 — 기간 → xlsx (Phase 4).

DUMMY 포맷 v1 (E3 — 회계사 양식 확인 전 선개발): 실제 양식이 확정되면
컬럼/행 매핑만 바꾸면 되도록 스왑 지점을 고립해 둔다.

    양식 스왑 지점 (여기만 바꾸면 됨):
        - EXPORT_COLUMNS: 헤더 목록
        - export_row(): 행 1건 → 셀 값 목록 (동결/미확정 공용)

두 가지 원천을 같은 시트 모양으로 낸다:
    - confirmed 기간 → FROZEN payroll_entries. 금액 스칼라는 confirm 시
      breakdown 합계와 일치 검증된 값이라 재계산 없이 그대로 (rate 목록만
      breakdown 참조).
    - open 기간 → LIVE preview 계산 결과를 **DRAFT** 로. 확정 전 숫자라
      배너 행 + 파일명(payroll_draft_…)으로 오해 여지를 없앤다.

행 매핑은 한 벌뿐이다 (export_row) — 동결 entry 와 preview 행이 같은 필드
이름을 쓰므로 포맷을 두 군데 유지하지 않는다.

empid 매칭 (스펙 §계산규칙 외 v1 제외 항목):
    EMPID 셀 = empid → 없으면 crewid fallback → 둘 다 없으면 빈칸.
    둘 다 없는 직원은 Warnings 시트에 나열한다 (빈칸+경고).
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO
from typing import Optional, Protocol, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased
from uuid import UUID

from app.models.org_member import OrgMember, OrgMemberStore
from app.models.payroll import PayPeriod, PayrollEntry
from app.models.rate import HourlyRateHistory
from app.models.user import User
from app.schemas.payroll import EntryBreakdown, PayrollPreviewRow
from app.services.payroll_calc_service import (
    parse_frozen_breakdown,
    payroll_calc_service,
)
from app.core.payroll_rules import payroll_period_label
from app.utils.download import download_stamp, payroll_range_tag, safe_filename
from app.utils.names import display_name

EXPORT_SHEET_TITLE = "Payroll"
WARNINGS_SHEET_TITLE = "Warnings"
RATE_CHANGES_SHEET_TITLE = "Rate Changes"

# 미확정 기간 export 상단 배너 — 시트를 열자마자 draft 임을 알린다.
DRAFT_BANNER = "DRAFT — period not confirmed; numbers may change"

# ── DUMMY 포맷 v1 — 양식 스왑 지점 ①: 헤더 ──────────────────────────
EXPORT_COLUMNS: list[str] = [
    "EMPID",
    "CREWID",
    "Name",
    "Regular Hours",
    "OT Hours",
    "DT Hours",
    "Rate(s)",
    "Regular Pay",
    "OT Pay",
    "DT Pay",
    "Penalty Pay",
    "Card Tips",
    "Gross Pay",
]
_COLUMN_WIDTHS: list[int] = [10, 10, 22, 14, 12, 12, 16, 13, 12, 12, 13, 12, 12]

RATE_CHANGES_COLUMNS: list[str] = [
    "Name",
    "EMPID",
    "Old Rate",
    "New Rate",
    "Effective Date",
    "Memo",
    "Changed By",
    "Changed At (UTC)",
]
_RATE_CHANGES_WIDTHS: list[int] = [22, 10, 11, 11, 14, 40, 18, 20]

WARNINGS_COLUMNS: list[str] = ["Name", "Issue"]
_UNMATCHED_ISSUE = (
    "No EMPID or CREWID on file — match this employee manually and assign an "
    "EMPID so future exports line up automatically"
)

_TWO_DP = Decimal("0.01")


def minutes_to_hours(minutes: int) -> Decimal:
    """분 → 시간 (소수 2자리, HALF_UP) — export 표기 전용 (계산은 항상 분)."""
    return (Decimal(minutes) / Decimal(60)).quantize(_TWO_DP, rounding=ROUND_HALF_UP)


def format_rates(breakdown: EntryBreakdown) -> str:
    """breakdown segments → rate 셀 문자열 — 멀티 rate 는 콤마 목록.

    segments 는 동결 시 rate 오름차순 — 순서를 보존하고 중복만 제거한다.
    """
    seen: list[str] = []
    for segment in breakdown.segments:
        formatted = str(segment.rate.quantize(_TWO_DP))
        if formatted not in seen:
            seen.append(formatted)
    return ", ".join(seen)


class ExportRowSource(Protocol):
    """행 매핑이 요구하는 최소 형태 — 동결 entry 와 preview 행의 공통 필드.

    두 원천(PayrollEntry / PayrollPreviewRow)이 이름·의미가 같은 값을 들고
    있으므로, 포맷을 복제하는 대신 이 형태 하나로 받는다.
    """

    empid: Optional[int]
    crewid: Optional[int]
    member_name: str
    regular_minutes: int
    ot_minutes: int
    dt_minutes: int
    regular_pay: Decimal
    ot_pay: Decimal
    dt_pay: Decimal
    penalty_pay: Decimal
    card_tips: Decimal
    gross_pay: Decimal


# ── DUMMY 포맷 v1 — 양식 스왑 지점 ②: 행 매핑 ────────────────────────
def export_row(source: ExportRowSource, breakdown: EntryBreakdown) -> list:
    """행 1건 → 셀 값 목록 (EXPORT_COLUMNS 순서와 1:1) — 동결/미확정 공용.

    EMPID 셀은 empid → crewid fallback → 빈칸(None). CREWID 셀은 항상
    crewid 그대로 (fallback 이 일어났는지 대조 가능하도록).
    """
    empid_cell: Optional[int] = (
        source.empid if source.empid is not None else source.crewid
    )
    return [
        empid_cell,
        source.crewid,
        source.member_name,
        minutes_to_hours(source.regular_minutes),
        minutes_to_hours(source.ot_minutes),
        minutes_to_hours(source.dt_minutes),
        format_rates(breakdown),
        source.regular_pay,
        source.ot_pay,
        source.dt_pay,
        source.penalty_pay,
        source.card_tips,
        source.gross_pay,
    ]


@dataclass(frozen=True)
class IdleMemberRow:
    """기간에 급여 활동이 없는 재직 직원 — 선택 포함 시 0 행으로 내보낸다.

    preview/entries 로스터에는 없는 사람이다 (has_payroll_activity 가 거른
    빈 행이 아니라 애초에 후보가 아니었던 사람). ExportRowSource 형태를 그대로
    갖춰 export_row 한 벌로 매핑된다.
    """

    user_id: UUID
    member_name: str
    empid: Optional[int]
    crewid: Optional[int]
    regular_minutes: int = 0
    ot_minutes: int = 0
    dt_minutes: int = 0
    regular_pay: Decimal = Decimal("0.00")
    ot_pay: Decimal = Decimal("0.00")
    dt_pay: Decimal = Decimal("0.00")
    penalty_pay: Decimal = Decimal("0.00")
    card_tips: Decimal = Decimal("0.00")
    gross_pay: Decimal = Decimal("0.00")


_EMPTY_BREAKDOWN = EntryBreakdown()


def idle_member_export_row(row: IdleMemberRow) -> list:
    """무활동 재직 직원 1건 → 0 셀 행 (Rate(s) 는 빈칸)."""
    return export_row(row, _EMPTY_BREAKDOWN)


def entry_export_row(entry: PayrollEntry, breakdown: EntryBreakdown) -> list:
    """동결 entry 1건 → 셀 값 목록 (breakdown 은 JSONB 파싱본을 받는다)."""
    return export_row(entry, breakdown)


def preview_export_row(row: PayrollPreviewRow) -> list:
    """preview 행 1건 → 셀 값 목록 (breakdown 은 행이 이미 들고 있다)."""
    return export_row(row, row.breakdown)


def _style_headers(
    ws: Worksheet, headers: list[str], widths: list[int], *, row: int = 1
) -> None:
    """헤더 행 스타일 — dashboard export 와 동일한 톤."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="2D3436", end_color="2D3436", fill_type="solid"
    )
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=row, column=col_idx).column_letter].width = (
            width
        )


def _add_draft_banner(ws: Worksheet, column_count: int) -> None:
    """1행에 DRAFT 배너 — 헤더 폭만큼 병합해 눈에 띄는 경고색으로."""
    cell = ws.cell(row=1, column=1, value=DRAFT_BANNER)
    cell.font = Font(bold=True, color="9C2A2A", size=12)
    cell.fill = PatternFill(
        start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
    )
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=column_count
    )
    ws.row_dimensions[1].height = 22


@dataclass
class RateChangeExportRow:
    """Rate Changes 시트 1행 — 기간 내 effective_date 를 갖는 시급 변경."""

    name: str
    empid: Optional[int]
    old_rate: Optional[Decimal]
    new_rate: Decimal
    effective_date: date
    memo: Optional[str]
    changed_by: Optional[str]
    changed_at: Optional[str]  # "YYYY-MM-DD HH:MM" (UTC)


def _assemble_workbook(
    rows: Sequence[list],
    unmatched_names: Sequence[str],
    *,
    draft: bool,
    rate_changes: Sequence[RateChangeExportRow] = (),
) -> Workbook:
    """셀 값 행 목록 → 워크북 (draft 면 배너 1행 뒤에 헤더).

    Payroll 시트 1개 + (empid/crewid 둘 다 없는 직원이 있을 때만) Warnings
    시트 + (기간 내 시급 변경이 있을 때만) Rate Changes 시트.
    부속 시트들은 draft 여부와 무관하게 같은 모양이다.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = EXPORT_SHEET_TITLE
    header_row = 1
    if draft:
        _add_draft_banner(ws, len(EXPORT_COLUMNS))
        header_row = 2
    _style_headers(ws, EXPORT_COLUMNS, _COLUMN_WIDTHS, row=header_row)

    for row in rows:
        ws.append(row)

    if unmatched_names:
        warn_ws = wb.create_sheet(WARNINGS_SHEET_TITLE)
        _style_headers(warn_ws, WARNINGS_COLUMNS, [22, 90])
        for name in unmatched_names:
            warn_ws.append([name, _UNMATCHED_ISSUE])

    if rate_changes:
        rc_ws = wb.create_sheet(RATE_CHANGES_SHEET_TITLE)
        _style_headers(rc_ws, RATE_CHANGES_COLUMNS, _RATE_CHANGES_WIDTHS)
        for rc in rate_changes:
            rc_ws.append([
                rc.name,
                rc.empid,
                rc.old_rate,
                rc.new_rate,
                rc.effective_date.isoformat(),
                rc.memo,
                rc.changed_by,
                rc.changed_at,
            ])

    return wb


def _append_idle_rows(
    rows: list[list], unmatched: list[str], idle_rows: Sequence[IdleMemberRow]
) -> None:
    """무활동 직원 0 행을 데이터 행 **뒤에** 이름순으로 덧붙인다.

    지급 행과 섞어 정렬하면 회계사가 0 행을 골라내야 하므로 블록을 분리한다.
    empid/crewid 둘 다 없는 직원은 지급 행과 같은 규칙으로 Warnings 에 올린다.
    """
    for row in sorted(idle_rows, key=lambda r: (r.member_name, str(r.user_id))):
        rows.append(idle_member_export_row(row))
        if row.empid is None and row.crewid is None:
            unmatched.append(row.member_name)


def build_export_workbook(
    entries: Sequence[PayrollEntry],
    idle_rows: Sequence[IdleMemberRow] = (),
    rate_changes: Sequence[RateChangeExportRow] = (),
) -> Workbook:
    """동결 entries (+ 선택: 무활동 재직 직원 0 행 / Rate Changes 시트) → 워크북.

    순수 함수 — DB 없음. 선택 인자 둘은 서로 독립이다(0 행은 Payroll 시트 뒤,
    시급 변경은 별도 시트). **키워드로 넘길 것** — 위치 인자로 넘기면 두
    기능이 서로 다른 자리에 꽂힌다.


    breakdown 은 계약 파서 경유 — calc_version 이 다르면 조용히 오독하지 않고
    400 으로 실패한다.
    """
    rows: list[list] = []
    unmatched: list[str] = []
    for entry in entries:
        breakdown = parse_frozen_breakdown(entry.breakdown)
        rows.append(entry_export_row(entry, breakdown))
        if entry.empid is None and entry.crewid is None:
            unmatched.append(entry.member_name)
    _append_idle_rows(rows, unmatched, idle_rows)
    return _assemble_workbook(
        rows, unmatched, draft=False, rate_changes=rate_changes
    )


def build_draft_workbook(
    preview_rows: Sequence[PayrollPreviewRow],
    idle_rows: Sequence[IdleMemberRow] = (),
    rate_changes: Sequence[RateChangeExportRow] = (),
) -> Workbook:
    """live preview 행 (+ 선택: 무활동 0 행 / Rate Changes) → DRAFT 워크북.


    동결본과 같은 컬럼/행 매핑이고, 상단 배너로만 구분된다. preview 행의
    breakdown 은 이미 계약 모델이라 파싱 단계가 없다.
    """
    rows = [preview_export_row(row) for row in preview_rows]
    unmatched = [
        row.member_name
        for row in preview_rows
        if row.empid is None and row.crewid is None
    ]
    _append_idle_rows(rows, unmatched, idle_rows)
    return _assemble_workbook(
        rows, unmatched, draft=True, rate_changes=rate_changes
    )


def workbook_bytes(wb: Workbook) -> bytes:
    """워크북 → xlsx bytes (StreamingResponse 용)."""
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _payroll_filename(
    kind: str,
    scope_name: str,
    start_date: date,
    end_date: date,
    *,
    draft: bool,
    generated_at: datetime | None,
    ext: str,
    extra: str | None = None,
) -> str:
    """급여 산출물 파일명 공통 규칙 — CFS/기간 export 가 같은 모양을 쓰도록.

    `{kind}_{스코프}[_{extra}]_{기간ID}_{날짜범위}_{상태}_{생성시각}.{ext}`
    """
    parts = [kind, safe_filename(scope_name)]
    if extra:
        parts.append(safe_filename(extra))
    parts += [
        payroll_period_label(start_date),
        payroll_range_tag(start_date, end_date),
        "DRAFT" if draft else "FINAL",
        download_stamp(generated_at),
    ]
    return "_".join(parts) + f".{ext}"


def export_filename(
    scope_name: str,
    start_date: date,
    end_date: date,
    *,
    draft: bool = False,
    generated_at: datetime | None = None,
) -> str:
    """`Payroll_{스코프}_{기간ID}_{날짜범위}_{상태}_{생성시각}.xlsx`.

    예: `Payroll_ODG_2026.08.FH_20260801-0815_DRAFT_20260820-1352Z.xlsx`

    파일명 하나로 **어느 법인의 / 언제 기간 / 확정본인지 / 언제 받은 것인지**가
    구분돼야 한다 — 회계사에게 여러 번 보내는 파일이라 받는 쪽 폴더에서 섞인다.
        - 기간ID(2026.08.FH): 회계사 파일의 payroll_id 칸과 같은 값 — 대조 키
        - 상태: DRAFT / FINAL 를 **양쪽 다 명시**한다. 예전엔 확정본에 표시가 없어
          "표시 없음 = 확정본" 이라는 암묵 규칙이었고, 그건 파일만 봐선 모른다
        - 생성시각(UTC, Z): DRAFT 는 받을 때마다 숫자가 달라지므로 최신본 구분에 필수

    스코프명은 유니코드 그대로 둔다 (한글 법인/매장명이 남아야 구분된다).
    전송은 Content-Disposition 의 filename*(UTF-8)이 책임진다 — app/utils/download.
    """
    return _payroll_filename(
        "Payroll", scope_name, start_date, end_date,
        draft=draft, generated_at=generated_at, ext="xlsx",
    )


@dataclass
class PeriodExport:
    """build_period_export 결과 — is_draft 는 파일명 표기에 필요."""

    workbook: Workbook
    is_draft: bool


class PayrollExportService:
    """기간 export — 라우터에서 호출하는 파사드."""

    async def _load_rate_changes(
        self,
        db: AsyncSession,
        period: PayPeriod,
        *,
        user_ids: Sequence,
        name_by_user: dict,
        empid_by_user: dict,
    ) -> list[RateChangeExportRow]:
        """Rate Changes 시트 데이터 — 기간 내 effective_date 인 개인 시급 변경.

        스코프 = 이번 export 행에 등장하는 직원(user_ids). 이름/EMPID 는
        export 행의 스냅샷을 재사용해 본 시트와 표기가 어긋나지 않게 한다.
        """
        if not user_ids:
            return []
        changer = aliased(User)
        rows = (
            await db.execute(
                select(HourlyRateHistory, OrgMember.user_id, changer)
                .join(OrgMember, OrgMember.id == HourlyRateHistory.org_member_id)
                .outerjoin(changer, changer.id == HourlyRateHistory.changed_by)
                .where(
                    OrgMember.organization_id == period.organization_id,
                    OrgMember.user_id.in_(list(user_ids)),
                    HourlyRateHistory.effective_date >= period.start_date,
                    HourlyRateHistory.effective_date <= period.end_date,
                )
                .order_by(
                    HourlyRateHistory.effective_date.asc(),
                    HourlyRateHistory.created_at.asc(),
                )
            )
        ).all()
        return [
            RateChangeExportRow(
                name=name_by_user.get(user_id, ""),
                empid=empid_by_user.get(user_id),
                old_rate=history.old_rate,
                new_rate=history.new_rate,
                effective_date=history.effective_date,
                memo=history.reason,
                changed_by=display_name(changed_by_user) if changed_by_user else None,
                changed_at=(
                    history.created_at.strftime("%Y-%m-%d %H:%M")
                    if history.created_at
                    else None
                ),
            )
            for history, user_id, changed_by_user in rows
        ]

    async def build_period_export(
        self,
        db: AsyncSession,
        period: PayPeriod,
        *,
        include_idle_members: bool = False,
    ) -> PeriodExport:
        """기간 → 워크북. confirmed 는 동결 entries, open 은 live preview DRAFT.

        open 기간은 preview 와 같은 계산 경로를 타므로 payroll_events 가
        upsert 되는 부수효과가 있다 (flush 만 — commit 은 호출자 소유,
        preview 엔드포인트와 동일).

        Args:
            include_idle_members: True 면 기간에 급여 활동이 없는 **재직 중**
                직원도 0 행으로 덧붙인다 (화면 로스터에는 영향 없음 — 파일
                전용). 비활성 계정·퇴직 소속은 활동이 없으면 어느 경우에도
                나오지 않는다.
        """
        if period.status != "confirmed":
            rows = await payroll_calc_service.preview_period(db, period)
            idle = (
                await self._idle_member_rows(db, period, {r.user_id for r in rows})
                if include_idle_members
                else []
            )
            rate_changes = await self._load_rate_changes(
                db,
                period,
                user_ids=[r.user_id for r in rows],
                name_by_user={r.user_id: r.member_name for r in rows},
                empid_by_user={
                    r.user_id: (r.empid if r.empid is not None else r.crewid)
                    for r in rows
                },
            )
            return PeriodExport(
                workbook=build_draft_workbook(
                    rows, idle_rows=idle, rate_changes=rate_changes
                ),
                is_draft=True,
            )

        entries = (
            (
                await db.execute(
                    select(PayrollEntry)
                    .where(PayrollEntry.pay_period_id == period.id)
                    .order_by(
                        PayrollEntry.member_name.asc(), PayrollEntry.revision.asc()
                    )
                )
            )
            .scalars()
            .all()
        )
        idle = (
            await self._idle_member_rows(db, period, {e.user_id for e in entries})
            if include_idle_members
            else []
        )
        rate_changes = await self._load_rate_changes(
            db,
            period,
            user_ids=[e.user_id for e in entries if e.user_id is not None],
            name_by_user={
                e.user_id: e.member_name for e in entries if e.user_id is not None
            },
            empid_by_user={
                e.user_id: (e.empid if e.empid is not None else e.crewid)
                for e in entries
                if e.user_id is not None
            },
        )
        return PeriodExport(
            workbook=build_export_workbook(
                entries, idle_rows=idle, rate_changes=rate_changes
            ),
            is_draft=False,
        )

    async def _idle_member_rows(
        self, db: AsyncSession, period: PayPeriod, present_user_ids: set[UUID]
    ) -> list[IdleMemberRow]:
        """스코프 매장에 배정된 재직 직원 중 로스터에 없는 사람 → 0 행.

        재직 = 계정 활성(users.is_active) ∧ 소속 활성(org_members.status).
        이름/empid/crewid 는 preview 행과 같은 로더(_load_members)로 뽑아
        스냅샷 규칙(그룹 내 첫 non-null empid)을 공유한다.
        """
        stores = await payroll_calc_service._period_stores(db, period)
        store_ids = [s.id for s in stores]
        org_id = period.organization_id
        user_ids = set(
            (
                await db.execute(
                    select(OrgMember.user_id)
                    .join(User, User.id == OrgMember.user_id)
                    .join(OrgMemberStore, OrgMemberStore.org_member_id == OrgMember.id)
                    .where(
                        OrgMember.organization_id == org_id,
                        OrgMember.status == "active",
                        User.is_active.is_(True),
                        OrgMemberStore.store_id.in_(store_ids),
                    )
                    .distinct()
                )
            )
            .scalars()
            .all()
        )
        idle_ids = user_ids - present_user_ids
        if not idle_ids:
            return []
        members = await payroll_calc_service._load_members(
            db, idle_ids, org_id, store_ids
        )
        rows: list[IdleMemberRow] = []
        for user_id in idle_ids:
            user = members["users"].get(user_id)
            member = members["by_user"].get(user_id)
            if user is None:
                continue
            rows.append(
                IdleMemberRow(
                    user_id=user_id,
                    member_name=display_name(user),
                    empid=members["empid"].get(user_id),
                    crewid=member.crewid if member is not None else None,
                )
            )
        return rows


payroll_export_service = PayrollExportService()
