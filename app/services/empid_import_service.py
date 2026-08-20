"""EMPID 임포트 — 레거시 직원 마스터(xlsx/csv)를 매장별 empid 로 등록.

콘솔 /users/bulk/empid 탭의 백엔드. 레거시 도구(empid_reconcile_service — users.employee_no
기입)와 달리 이 서비스는 **org_member_stores.empid** (매장별 정수 순번)에 기록하며
users.employee_no 는 건드리지 않는다 (employee_no 는 폐기 방향 — 2026-08-03 결정).

설계:
- blind import 금지 → preview(사람×매장 버킷) → 운영자 검토(체크박스) → commit.
- 사람 매칭: 이메일 (레거시와 동일 — placeholder/deferred 로직 재사용).
- 매장 매칭: COMPANY/CORP_ABR_3 → org 의 Store name/code 정규화 비교. 미매칭 매장은 생략(리포트만).
- 기존 empid 는 전 매장 자동백필(1..N) 상태 — "이미 있으면 skip" 이 아니라 **rebind**:
  업로드값 == 현재값 → same(스킵), 다르면 rebind(기본 선택 = 업로드값).
- commit 은 매장 단위 3-phase (비우기 → 기입 → 잔여 재채번) 단일 트랜잭션:
  순열 교환(A:5→3, B:3→5)·소번호 재기입이 (store_id, empid) partial unique 에
  걸리지 않게 하고, 파일에 없는 기존 인원이 번호를 뺏기면 next_empid 로 재채번한다.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from uuid import UUID

from sqlalchemy import select, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.client_surface import current_channel
from app.models.empid_change import (
    EMPID_SOURCE_COMMIT,
    EMPID_SOURCE_RENUMBER,
    EmpidChange,
)
from app.models.org_member import (
    EMPID_KIND_SEQUENCE,
    EMPID_KINDS,
    OrgMember,
    OrgMemberStore,
)
from app.models.organization import Store
from app.repositories.store_repository import store_repository
from app.repositories.user_repository import user_repository
from app.services.empid_reconcile_service import (
    EmpRow,
    _is_placeholder_email,
    _name_tokens,
    _name_similar,
    _norm_email,
    _first_name_token,
    parse_emplist,
)
from app.services.org_numbering import (
    empid_cursor_state,
    empid_scope_store_ids,
    lock_empid_scope,
    next_empid,
)
from app.utils.exceptions import DuplicateError

# 액션 값 — 사람×매장 1건의 처리 분류.
ACTION_SAME = "same"                      # 업로드값 == 현재값 → 할 일 없음
ACTION_REBIND = "rebind"                  # 값이 다름 → 재기입 후보 (기본 선택 = 업로드값)
ACTION_NEW_ASSIGNMENT = "new_assignment"  # 매장 배정 행 없음 → 배정 생성 + 번호 등록
ACTION_UNMATCHED_STORE = "unmatched_store"  # COMPANY 를 org 매장으로 못 찾음 → 생략(리포트)
ACTION_INVALID = "invalid"                # emp_id 정수화 실패 / org_member 없음 등
ACTION_NEEDS_USER = "needs_user"          # 매장·번호 유효하나 유저 미확정 → 운영자가 직접 선택해 등록
ACTION_NEEDS_STORE = "needs_store"        # 그룹 스코프 매칭 — 사람은 확정, 그룹 내 배정 매장이 없어 운영자가 매장 선택


def _row_name_tokens(r: EmpRow) -> set[str]:
    """파일 행의 이름 토큰 — name 컬럼 + FIRST/LAST 분리 컬럼 합집합."""
    toks = _name_tokens(r.name)
    toks |= _name_tokens(getattr(r, "first_name", None))
    toks |= _name_tokens(getattr(r, "last_name", None))
    return toks


def _fuzzy_cover(a: set[str], b: set[str], thresh: float = 0.78) -> float:
    """a 의 토큰이 b 의 어떤 토큰과든 유사(비율≥thresh)한 비율 — 철자 흔들림 흡수.

    0.78 은 "deigo↔diego"(0.80) 같은 실제 급여 마스터 오탈자를 잡는 하한.
    퍼지 결과는 **제안 전용**(자동 등록 없음)이라 임계를 보수보다 회수율 쪽에 둔다.
    """
    if not a or not b:
        return 0.0
    hit = 0
    for t in a:
        best = max((SequenceMatcher(None, t, o).ratio() for o in b), default=0.0)
        if best >= thresh:
            hit += 1
    return hit / len(a)


def _fuzzy_score(a: set[str], b: set[str]) -> float:
    """양방향 커버리지의 최솟값 — 한쪽만 부분집합이어도 과신하지 않는다."""
    return min(_fuzzy_cover(a, b), _fuzzy_cover(b, a))


def _norm_key(value: str | None) -> str:
    """매장 매칭 키 정규화 — 대문자, 영숫자만."""
    if not value:
        return ""
    return re.sub(r"[^A-Z0-9]", "", value.upper())


def _emp_id_int(raw: str) -> int | None:
    """레거시 emp_id 문자열 → 양의 정수 (선행 0 허용: "07"→7). 실패 시 None."""
    s = raw.strip()
    if not s or not s.isdigit():
        return None
    v = int(s)
    return v if v >= 1 else None


def empid_band(empid: int) -> tuple[int, int, str]:
    """백 단위 번호대 → (lo, hi, label). 1-99, 100-199, 1000-1099 …

    export split_by="band" 시트 구분과 preview 분포가 **같은 규칙**을 쓰도록 한 곳에 둔다
    (계약 §3-5). 규칙 = (empid // 100) * 100, 0 대는 1 부터 표기.
    """
    base = (empid // 100) * 100
    lo = base if base > 0 else 1
    hi = base + 99
    return lo, hi, f"{lo}-{hi}"


def build_store_index(stores: list[Store]) -> dict[str, list[Store]]:
    """정규화 키 → 매장 목록. name/code 둘 다 키로 등록 (중복 키 = 모호 매칭)."""
    index: dict[str, list[Store]] = {}
    for s in stores:
        for key in {_norm_key(s.name), _norm_key(s.code)}:
            if key:
                index.setdefault(key, []).append(s)
    return index


def match_store(index: dict[str, list[Store]], company: str, corp_abr: str | None) -> Store | None:
    """CORP_ABR_3/COMPANY → Store. 유일 매칭만 인정 (모호하면 None).

    corp_abr(매장 코드)를 company(법인명)보다 먼저 본다 — 법인 하나가 여러 매장을
    거느리는 파일(M KOREAN BBQ = MKB+MSK)에서 company 키 오버라이드가 이미 코드로
    매칭된 형제 매장 행까지 삼키는 사고를 막는다.
    """
    for key in (_norm_key(corp_abr), _norm_key(company)):
        if not key:
            continue
        candidates = index.get(key, [])
        # name/code 가 같은 매장을 이중 등록했을 수 있으므로 id 로 유니크화
        unique = {s.id: s for s in candidates}
        if len(unique) == 1:
            return next(iter(unique.values()))
    return None


@dataclass
class GroupScope:
    """운영자가 라벨을 **그룹**에 매핑했을 때의 매칭 스코프.

    행의 매장은 파일이 아니라 **그 사람의 그룹 내 기존 배정**이 결정한다 —
    "M KOREAN BBQ" 아래 Gloria 가 MSK 소속이면 번호는 MSK 배정 행에 적힌다.
    겸업자(그룹 내 2개 매장 배정)는 전 배정에 같은 번호가 가도록 매장별 엔트리를
    만든다 (numbering_mode=group 의 1인 1번호 의미).

    라벨이 **매장**으로 풀렸어도 그 매장이 shared-numbering(mode=group) 그룹
    소속이면 이 스코프로 승격된다 — 번호가 그룹 단위인데 매장 하나만 갱신하면
    겸업자의 다른 배정에 옛 번호가 남는 모순이 생기기 때문 (2026-08-16 Jiho 건).
    그때 원래 매장은 hint_store_id 로 보존해 needs_store 의 기본 후보로 쓴다.
    """

    id: str
    name: str
    stores: list  # list[Store]
    hint_store_id: str | None = None  # 라벨이 지목했던 매장 (needs_store 프리필)


@dataclass
class ImportEntry:
    """사람×매장 1건 — preview 행."""

    store_id: str | None       # 매칭된 매장 UUID (None = unmatched)
    store_name: str | None     # 매칭된 매장 이름
    company: str               # 파일 COMPANY 원문
    emp_id_raw: str            # 파일 emp_id 원문 (선행 0 보존 표시용)
    emp_id: int | None         # 정수 정규화 값 (None = invalid)
    current_empid: int | None  # 현재 org_member_stores.empid
    has_assignment: bool       # 매장 배정 행 존재 여부
    action: str                # ACTION_* 값
    warning: str | None = None  # 그룹 스코프 충돌 등 경고 (블록 아님)
    dormant: bool = False      # 휴면 배정(is_work_assignment=False) 여부 — 번호만 쓰고 재활성화 안 함
    person_name: str | None = None  # 파일 행의 인물 이름 — placeholder(공유 이메일)에서 행별 picker 라벨
    corp_abr: str | None = None  # 파일 매장 코드 원문 (unmatched 매핑 키·표시용)
    # 그룹 스코프 매칭 정보 — needs_store/needs_user 에서 매장 picker 옵션으로 쓴다
    group_id: str | None = None
    group_name: str | None = None
    group_stores: list[dict] | None = None  # [{store_id, store_name}]
    hint_store_id: str | None = None  # 파일 corp 가 지목했던 매장 — picker 프리필


@dataclass
class PersonRow:
    """사람 1명 — preview 그룹."""

    email: str | None
    name: str
    user_id: str | None
    user_full_name: str | None
    entries: list[ImportEntry] = field(default_factory=list)
    note: str = ""
    similar: list[str] = field(default_factory=list)  # deferred 이름 유사 힌트 (표시용)
    members: list[str] = field(default_factory=list)  # placeholder — 파일 내 인물 나열 (표시용)
    # 유저 picker 기본값 후보 — 이름 유사 DB 유저 (구조화, 콘솔 select 프리필용)
    similar_users: list[dict] = field(default_factory=list)  # {user_id, full_name, email}
    matched_by: str | None = None  # "crewid" = 파일의 CREWID 로 정확 매칭 (email/이름 달라도 확정)


@dataclass
class ImportPreview:
    """preview 결과 — 버킷별 사람 목록 + 카운트."""

    people: list[PersonRow] = field(default_factory=list)      # user 매칭 성공 (액션 가능)
    placeholder: list[PersonRow] = field(default_factory=list)  # 더미/공유 이메일 (리포트)
    deferred: list[PersonRow] = field(default_factory=list)     # DB 미매칭 (리포트)
    excluded_rows: int = 0
    total_rows: int = 0
    # 매장 미매칭 원문 집계 — 콘솔이 "이 코드는 어느 매장?" 매핑 UI 를 그리는 재료.
    # [{key(정규화, store_overrides 의 키로 재사용), company, corp_abr, rows}]
    unmatched_stores: list[dict] = field(default_factory=list)
    # 자동 적용된 저장 별칭 — [{key, target_id, store_name(매장 확정 시) | group_name}]
    saved_aliases: list[dict] = field(default_factory=list)
    # 양측 대조 — 스코프(그룹/매장)별 empid diff.
    # [{scope, id, name, matched, htm_only:[{empid,user,store}], file_only:[{empid,name}]}]
    reconciliation: list[dict] = field(default_factory=list)
    # 업로드 번호의 백 단위 분포 — [{band, lo, hi, count}]. 대역 밖 번호가 눈에 띄게 하는 용도
    # (계약 §3-5). 자동 예외 추천은 하지 않는다 — 판단은 운영자 몫.
    distribution: list[dict] = field(default_factory=list)

    def counts(self) -> dict[str, int]:
        entry_actions = [e.action for p in self.people for e in p.entries]
        # unmatched/invalid 는 세 버킷 전체 엔트리에서 집계 — 미지 이메일 행(deferred/placeholder
        # 안의 매장 미매칭·비정수)이 타일에서 누락되면 카운터가 운영자를 오도한다.
        all_actions = entry_actions + [
            e.action
            for p in list(self.placeholder) + list(self.deferred)
            for e in p.entries
        ]
        return {
            "people": len(self.people),
            "same": entry_actions.count(ACTION_SAME),
            "rebind": entry_actions.count(ACTION_REBIND),
            "new_assignment": entry_actions.count(ACTION_NEW_ASSIGNMENT),
            "unmatched_store": all_actions.count(ACTION_UNMATCHED_STORE),
            "invalid": all_actions.count(ACTION_INVALID),
            "needs_user": all_actions.count(ACTION_NEEDS_USER),
            "placeholder": len(self.placeholder),
            # deferred = 등록 가능(needs_user) 엔트리가 있는 사람 수 — 리포트 온리 행만 있는
            # 사람은 unmatched/invalid 타일이 이미 설명한다.
            "deferred": sum(
                1 for p in self.deferred
                if any(e.action == ACTION_NEEDS_USER for e in p.entries)
            ),
            "excluded_rows": self.excluded_rows,
            "total_rows": self.total_rows,
            # 이름으로 자동 매칭된 사람 수 — 운영자가 우선 검토할 대상
            "matched_by_name": sum(1 for p in self.people if p.matched_by == "name"),
            "needs_store": all_actions.count(ACTION_NEEDS_STORE),
            # 양측 대조 합계 (사람 단위) — HTM 미매칭 배정 / 파일 미매칭 행
            "htm_unmatched": sum(len(r["htm_unmatched"]) for r in self.reconciliation),
            "file_unmatched": sum(len(r["file_unmatched"]) for r in self.reconciliation),
        }


async def _current_empid_map(
    db: AsyncSession, organization_id: UUID
) -> tuple[
    dict[UUID, UUID],
    dict[tuple[UUID, UUID], int | None],
    dict[tuple[UUID, UUID], bool],
    dict[tuple[UUID, UUID], str],
]:
    """(user_id→org_member_id, (member, store)→empid, →is_work_assignment, →empid_kind)."""
    member_rows = (
        await db.execute(
            select(OrgMember.id, OrgMember.user_id).where(
                OrgMember.organization_id == organization_id
            )
        )
    ).all()
    member_by_user = {r.user_id: r.id for r in member_rows}
    member_ids = [r.id for r in member_rows]
    empid_map: dict[tuple[UUID, UUID], int | None] = {}
    work_map: dict[tuple[UUID, UUID], bool] = {}
    kind_map: dict[tuple[UUID, UUID], str] = {}
    if member_ids:
        store_rows = (
            await db.execute(
                select(
                    OrgMemberStore.org_member_id,
                    OrgMemberStore.store_id,
                    OrgMemberStore.empid,
                    OrgMemberStore.is_work_assignment,
                    OrgMemberStore.empid_kind,
                ).where(OrgMemberStore.org_member_id.in_(member_ids))
            )
        ).all()
        empid_map = {(r.org_member_id, r.store_id): r.empid for r in store_rows}
        work_map = {(r.org_member_id, r.store_id): r.is_work_assignment for r in store_rows}
        kind_map = {(r.org_member_id, r.store_id): r.empid_kind for r in store_rows}
    return member_by_user, empid_map, work_map, kind_map


async def preview(
    db: AsyncSession, organization_id: UUID, content: bytes, filename: str = "",
    store_overrides: dict[str, str] | None = None,
    actor_id: UUID | None = None,
) -> ImportPreview:
    """업로드 파일 → 사람×매장 버킷 분류 + 스코프별 양측 대조.

    스코프 해석 — 운영자의 그루핑이 기준이다 (코드가 임의로 정하지 않는다):
    1. 당회 store_overrides (매장 또는 그룹 — **다매장 그룹 허용**)
    2. 저장 별칭(import_label_aliases — 이전 업로드에서 학습)
    3. 매장 name/code 자연 매칭 (유일할 때만)
    4. 그룹 name 자연 매칭 — 운영자가 그룹 이름을 파일 회사명과 같게 지었다면
       별칭 없이도 그 그룹으로 (매장 키와 겹치면 매장이 우선 — 기존 동작 보존)

    그룹 스코프의 매장 결정은 파일이 아니라 **그 사람의 그룹 내 기존 배정**이다.
    겸업자(그룹 내 복수 배정)는 전 배정 매장에 같은 번호가 가도록 엔트리를 만든다.
    배정이 없으면 needs_store — 운영자가 그룹 매장 중에서 고른다.

    reconciliation: 스코프별 양측 diff — HTM 에만 있는 번호(그룹 매장들의 현재
    empid 중 파일에 없는 것) / 파일에만 있는 번호. 그룹 스코프는 멤버 매장으로
    직접 매칭된 행(코드 MSK 등)도 접어 넣어 그룹 단위 그림을 만든다.

    store_overrides 는 org 별칭으로 upsert 된다 (한 번 가르치면 다음부터 자동).
    당회 명시 매핑 > 저장 별칭. preview 의 유일한 DB 기록은 이 upsert 뿐이다.
    """
    emp_rows, excluded = parse_emplist(content, filename)
    result = ImportPreview(excluded_rows=excluded, total_rows=len(emp_rows))

    users = await user_repository.get_by_org(db, organization_id)
    users_by_email: dict[str, list] = {}
    for u in users:
        e = _norm_email(u.email)
        if e:
            users_by_email.setdefault(e, []).append(u)
    user_name_by_id = {u.id: getattr(u, "full_name", "") for u in users}

    stores = await store_repository.get_by_org(db, organization_id, include_closed=True)
    store_index = build_store_index(stores)
    member_by_user, empid_map, work_map, _kind_map = await _current_empid_map(
        db, organization_id
    )
    member_user: dict[UUID, UUID] = {m: u for u, m in member_by_user.items()}

    # ── 스코프 해석 준비 ──────────────────────────────────────────────
    from app.models.import_label_alias import ImportLabelAlias
    from app.models.organization import StoreGroup

    store_by_id = {str(st.id): st for st in stores}
    groups = (
        await db.execute(
            select(StoreGroup).where(StoreGroup.organization_id == organization_id)
        )
    ).scalars().all()
    group_by_id = {str(g.id): g for g in groups}
    stores_by_group: dict[str, list[Store]] = {}
    for st in stores:
        if st.group_id is not None:
            stores_by_group.setdefault(str(st.group_id), []).append(st)

    def _promote_store(st: Store):
        """매장 → 그 매장이 shared-numbering(mode=group) 그룹 소속이면 그룹 스코프로 승격.

        번호가 그룹 단위인데 매장 하나만 갱신하면 겸업자의 다른 배정에 옛 번호가
        남는다 (파일 corp 가 "MSK" 처럼 매장 코드를 쓰는 행에서 실제로 발생).
        자연 매칭·저장 별칭·당회 오버라이드 어느 경로로 풀렸든 동일하게 적용해야
        스코프 해석이 채번 정책과 어긋나지 않는다. mode=store 그룹은 매장별 독립
        번호가 정책이므로 승격하지 않는다. 원래 매장은 hint 로 보존한다.
        """
        gid = str(st.group_id) if st.group_id is not None else None
        g = group_by_id.get(gid) if gid else None
        if g is None or getattr(g, "numbering_mode", None) != "group":
            return st
        members = stores_by_group.get(gid, [])
        if len(members) < 2:
            return st
        return GroupScope(
            id=gid, name=g.name, stores=members, hint_store_id=str(st.id)
        )

    def _make_scope(sid: str):
        """매장/그룹 id → Store | GroupScope | None.

        단일 매장 그룹은 Store 로 축약(지름길) — 매장 스코프와 동작 동일.
        """
        st = store_by_id.get(sid)
        if st is not None:
            return _promote_store(st)
        g = group_by_id.get(sid)
        if g is not None:
            members = stores_by_group.get(str(g.id), [])
            if len(members) == 1:
                return members[0]
            if members:
                return GroupScope(id=str(g.id), name=g.name, stores=members)
        return None

    # scope_map: 정규화 키 → Store | GroupScope. 낮은 우선순위부터 쌓아 덮어쓴다.
    scope_map: dict[str, object] = {}
    # (4) 그룹 이름/코드 자연 매칭 — 매장 키가 뒤에서 덮으므로 여기가 최하위.
    # 그룹 code(예: "ODG")는 급여 생태계가 이 법인을 부르는 표기 — 운영자가 그룹에
    # 코드를 넣어두면 파일의 그 표기가 별칭 없이 그룹으로 붙는다.
    for g in groups:
        members = stores_by_group.get(str(g.id), [])
        if not members:
            continue
        scope = (
            members[0] if len(members) == 1
            else GroupScope(id=str(g.id), name=g.name, stores=members)
        )
        for k in {_norm_key(g.name), _norm_key(getattr(g, "code", None))}:
            if k:
                scope_map[k] = scope
    # (3) 매장 자연 매칭 (유일 키만) — shared 그룹 소속 매장은 그룹으로 승격
    for key, lst in store_index.items():
        uniq = {st.id: st for st in lst}
        if len(uniq) == 1:
            scope_map[key] = _promote_store(next(iter(uniq.values())))
    # (2) 저장 별칭
    saved_rows = (
        await db.execute(
            select(ImportLabelAlias).where(
                ImportLabelAlias.organization_id == organization_id
            )
        )
    ).scalars().all()
    for alias in saved_rows:
        target_id = str(alias.store_id or alias.group_id or "")
        scope = _make_scope(target_id) if target_id else None
        if scope is not None and alias.key:
            scope_map[alias.key] = scope
            result.saved_aliases.append({
                "key": alias.key,
                "target_id": target_id,
                "store_id": str(scope.id) if isinstance(scope, Store) else None,
                "store_name": scope.name,
            })
    # (1) 당회 명시 매핑 — 최우선 + 별칭 upsert (학습)
    if store_overrides:
        upserts: list[tuple[str, str]] = []
        for raw_key, sid in store_overrides.items():
            k = _norm_key(raw_key)
            if not k:
                continue
            scope = _make_scope(str(sid))
            if scope is None:
                continue
            scope_map[k] = scope
            upserts.append((k, str(sid)))
        if upserts:
            existing = {a.key: a for a in saved_rows}
            for k, sid in upserts:
                sid_uuid = UUID(sid)
                is_store = sid in store_by_id
                row = existing.get(k)
                if row is None:
                    db.add(ImportLabelAlias(
                        organization_id=organization_id, key=k,
                        store_id=sid_uuid if is_store else None,
                        group_id=None if is_store else sid_uuid,
                        created_by=actor_id,
                    ))
                else:
                    row.store_id = sid_uuid if is_store else None
                    row.group_id = None if is_store else sid_uuid
            await db.commit()

    def _resolve_scope(r: EmpRow):
        """행 → Store | GroupScope | None. corp_abr(매장 코드) 키 우선."""
        for key in (_norm_key(r.corp_abr), _norm_key(r.company)):
            if key and key in scope_map:
                return scope_map[key]
        return None

    # ── 양측 대조 수집기 ──────────────────────────────────────────────
    # 파일 쪽: 스코프 정체성별 {emp_int: set(이름)}. 그룹은 멤버 매장 행도 접는다.
    recon_file: dict[tuple[str, str], dict[int, set[str]]] = {}
    engaged_groups: dict[str, GroupScope] = {}
    store_group_of: dict[str, str] = {}
    for gid, members in stores_by_group.items():
        for st in members:
            store_group_of[str(st.id)] = gid

    def _record_file_row(scope, r: EmpRow) -> None:
        emp_int = _emp_id_int(r.emp_id)
        if emp_int is None or scope is None:
            return
        if isinstance(scope, GroupScope):
            ident = ("group", scope.id)
            engaged_groups[scope.id] = scope
        else:
            ident = ("store", str(scope.id))
        recon_file.setdefault(ident, {}).setdefault(emp_int, set()).add(r.name)

    # 그룹 스코프 경고용 — store_id → 스코프 내 타 매장 사용 empid
    scope_cache: dict[UUID, set[int]] = {}

    async def _scope_other_empids(store_id: UUID) -> set[int]:
        if store_id not in scope_cache:
            scope_ids = await empid_scope_store_ids(db, store_id)
            used: set[int] = set()
            if len(scope_ids) > 1:
                for (m_id, s_id), emp in empid_map.items():
                    if emp is not None and s_id != store_id and s_id in scope_ids:
                        used.add(emp)
            scope_cache[store_id] = used
        return scope_cache[store_id]

    # ── 엔트리 빌더 (스코프 인식) ─────────────────────────────────────
    def _group_meta(scope: GroupScope) -> dict:
        return {
            "group_id": scope.id,
            "group_name": scope.name,
            "group_stores": [
                {"store_id": str(st.id), "store_name": st.name}
                for st in scope.stores
            ],
            # 매장→그룹 승격 행이면 파일 corp 가 지목했던 매장 — picker 프리필
            "hint_store_id": scope.hint_store_id,
        }

    async def build_entries(
        r: EmpRow, member_id: UUID | None, user: object | None,
    ) -> list[ImportEntry]:
        """EmpRow 1건 → 엔트리 목록 (그룹 스코프는 사람의 배정 매장 수만큼)."""
        scope = _resolve_scope(r)
        _record_file_row(scope, r)
        emp_int = _emp_id_int(r.emp_id)

        if scope is None:
            return [ImportEntry(
                store_id=None, store_name=None, company=r.company,
                emp_id_raw=r.emp_id, emp_id=emp_int, current_empid=None,
                has_assignment=False, action=ACTION_UNMATCHED_STORE,
                warning="no matching store in this org", person_name=r.name,
                corp_abr=r.corp_abr,
            )]

        if emp_int is None:
            name = scope.name
            return [ImportEntry(
                store_id=(None if isinstance(scope, GroupScope) else str(scope.id)),
                store_name=name, company=r.company,
                emp_id_raw=r.emp_id, emp_id=None, current_empid=None,
                has_assignment=False, action=ACTION_INVALID,
                warning="emp_id is not a positive integer", person_name=r.name,
                corp_abr=r.corp_abr,
                **(_group_meta(scope) if isinstance(scope, GroupScope) else {}),
            )]

        if isinstance(scope, GroupScope):
            if member_id is None:
                # 사람 미확정 — 유저와 매장 둘 다 운영자 선택 (그룹 매장 옵션 제공)
                return [ImportEntry(
                    store_id=None, store_name=scope.name, company=r.company,
                    emp_id_raw=r.emp_id, emp_id=emp_int, current_empid=None,
                    has_assignment=False, action=ACTION_NEEDS_USER,
                    person_name=r.name, corp_abr=r.corp_abr, **_group_meta(scope),
                )]
            assigned = [
                st for st in scope.stores if (member_id, st.id) in empid_map
            ]
            if not assigned:
                # 그룹 매핑은 맞지만 이 사람의 배정 매장이 없다 — 일부 일치, 경고 매핑
                return [ImportEntry(
                    store_id=None, store_name=scope.name, company=r.company,
                    emp_id_raw=r.emp_id, emp_id=emp_int, current_empid=None,
                    has_assignment=False, action=ACTION_NEEDS_STORE,
                    warning="matched in this group but not assigned to any of its stores — pick one",
                    person_name=r.name, corp_abr=r.corp_abr, **_group_meta(scope),
                )]
            entries: list[ImportEntry] = []
            for st in assigned:
                key = (member_id, st.id)
                current = empid_map.get(key)
                action = ACTION_SAME if current == emp_int else ACTION_REBIND
                dormant = not (work_map or {}).get(key, True)
                warning = (
                    "assignment is dormant (not in work assignment) — number is written but the person stays inactive"
                    if dormant else None
                )
                if action == ACTION_REBIND and len(assigned) > 1:
                    note = "unifying to one number across this group"
                    warning = f"{warning}; {note}" if warning else note
                entries.append(ImportEntry(
                    store_id=str(st.id), store_name=st.name, company=r.company,
                    emp_id_raw=r.emp_id, emp_id=emp_int, current_empid=current,
                    has_assignment=True, action=action, dormant=dormant,
                    warning=warning, corp_abr=r.corp_abr, **_group_meta(scope),
                ))
            return entries

        # 매장 스코프 — 기존 단일 매장 로직
        store = scope
        if member_id is None:
            return [ImportEntry(
                store_id=str(store.id), store_name=store.name, company=r.company,
                emp_id_raw=r.emp_id, emp_id=emp_int, current_empid=None,
                has_assignment=False, action=ACTION_NEEDS_USER,
                person_name=r.name, corp_abr=r.corp_abr,
            )]
        key = (member_id, store.id)
        has_row = key in empid_map
        current = empid_map.get(key)
        if not has_row:
            action = ACTION_NEW_ASSIGNMENT
        elif current == emp_int:
            action = ACTION_SAME
        else:
            action = ACTION_REBIND
        dormant = has_row and not (work_map or {}).get(key, True)
        entry = ImportEntry(
            store_id=str(store.id), store_name=store.name, company=r.company,
            emp_id_raw=r.emp_id, emp_id=emp_int, current_empid=current,
            has_assignment=has_row, action=action, dormant=dormant,
            warning=(
                "assignment is dormant (not in work assignment) — number is written but the person stays inactive"
                if dormant else None
            ),
            corp_abr=r.corp_abr,
        )
        others = await _scope_other_empids(store.id)
        if entry.emp_id in others:
            note = "same number already used by another store in this group"
            entry.warning = f"{entry.warning}; {note}" if entry.warning else note
        return [entry]

    # 이름 유사도 인덱스 (deferred 힌트)
    name_index: list[tuple[set[str], object]] = []
    for u in users:
        toks = _name_tokens(getattr(u, "full_name", None))
        if toks:
            name_index.append((toks, u))

    def _similar_users(name: str) -> list:
        toks = _name_tokens(name)
        if not toks:
            return []
        return [u for u_toks, u in name_index if _name_similar(toks, u_toks)][:5]

    def _similar(name: str) -> list[str]:
        return [
            f"{getattr(u, 'full_name', '')} <{getattr(u, 'email', None) or '-'}>"
            for u in _similar_users(name)
        ]

    def _similar_structured(name: str) -> list[dict]:
        """유저 picker 프리필용 — {user_id, full_name, email}."""
        return [
            {
                "user_id": str(u.id),
                "full_name": getattr(u, "full_name", ""),
                "email": getattr(u, "email", None),
            }
            for u in _similar_users(name)
        ]

    # ── CREWID 정확 매칭 (optional 컬럼) ────────────────────────────────
    users_by_id = {u.id: u for u in users}
    crew_rows = (
        await db.execute(
            select(OrgMember.user_id, OrgMember.id, OrgMember.crewid).where(
                OrgMember.organization_id == organization_id,
                OrgMember.crewid.isnot(None),
            )
        )
    ).all()
    member_by_crewid: dict[int, tuple[UUID, UUID]] = {
        r.crewid: (r.user_id, r.id) for r in crew_rows
    }

    def _crewid_int(r: EmpRow) -> int | None:
        if not r.crewid or not r.crewid.strip().isdigit():
            return None
        return int(r.crewid)

    def _crewid_warn(r: EmpRow) -> str | None:
        if r.crewid is None:
            return None
        v = _crewid_int(r)
        if v is None:
            return f"CREWID '{r.crewid}' is not a number — matched by email instead"
        return f"CREWID {v} not found in this org — matched by email instead"

    crewid_persons: dict[str, PersonRow] = {}   # user_id(str) → PersonRow (사전 매칭 병합용)
    person_seen: dict[str, set[tuple[str, str]]] = {}
    leftover: list[EmpRow] = []
    for r in emp_rows:
        cid = _crewid_int(r)
        resolved = member_by_crewid.get(cid) if cid is not None else None
        user = users_by_id.get(resolved[0]) if resolved else None
        if resolved is None or user is None:
            leftover.append(r)
            continue
        user_id, member_id = resolved
        key = str(user_id)
        person = crewid_persons.get(key)
        if person is None:
            person = PersonRow(
                email=_norm_email(getattr(user, "email", None)), name=r.name,
                user_id=key, user_full_name=user.full_name,
                matched_by="crewid", note=f"matched by CREWID {cid}",
            )
            crewid_persons[key] = person
            person_seen[key] = set()
            result.people.append(person)
        pair = (r.company, r.emp_id)
        if pair in person_seen[key]:
            continue
        person_seen[key].add(pair)
        new_entries = await build_entries(r, member_id, user)
        if r.email and r.email != _norm_email(getattr(user, "email", None)):
            note = "file email differs from account — matched by CREWID"
            for e in new_entries:
                e.warning = f"{e.warning}; {note}" if e.warning else note
        person.entries.extend(new_entries)

    # ── 이름 매칭 티어 — 이메일 없는 행 ──────────────────────────────
    name_index_full: list[tuple[set[str], object]] = []
    for u in users:
        toks = _name_tokens(getattr(u, "full_name", None))
        toks |= _name_tokens(getattr(u, "first_name", None))
        toks |= _name_tokens(getattr(u, "last_name", None))
        if toks:
            name_index_full.append((toks, u))

    def _rank_structured(cands: list) -> list[dict]:
        return [
            {
                "user_id": str(u.id),
                "full_name": getattr(u, "full_name", ""),
                "email": getattr(u, "email", None),
            }
            for u in cands[:5]
        ]

    for r in (row for row in leftover if not row.email):
        row_toks = _row_name_tokens(r)
        exact = [
            u for u_toks, u in name_index_full
            if row_toks and (row_toks <= u_toks or u_toks <= row_toks)
        ]
        exact_unique = {u.id: u for u in exact}

        if len(exact_unique) == 1:
            user = next(iter(exact_unique.values()))
            member_id = member_by_user.get(user.id)
            key = str(user.id)
            person = crewid_persons.get(key)
            if person is None:
                person = PersonRow(
                    email=_norm_email(getattr(user, "email", None)), name=r.name,
                    user_id=key, user_full_name=user.full_name,
                    matched_by="name",
                    note="matched by name — verify before commit",
                )
                crewid_persons[key] = person
                person_seen[key] = set()
                result.people.append(person)
            pair = (r.company, r.emp_id)
            if pair in person_seen[key]:
                continue
            person_seen[key].add(pair)
            if member_id is None:
                for e in await build_entries(r, None, None):
                    e.action = ACTION_INVALID
                    e.warning = "no org_member row"
                    person.entries.append(e)
                continue
            person.entries.extend(await build_entries(r, member_id, user))
            continue

        if len(exact_unique) > 1:
            cands = list(exact_unique.values())
            result.deferred.append(PersonRow(
                email=None, name=r.name, user_id=None, user_full_name=None,
                note="multiple accounts match this name — pick one",
                similar=[
                    f"{getattr(u, 'full_name', '')} <{getattr(u, 'email', None) or '-'}>"
                    for u in cands[:5]
                ],
                similar_users=_rank_structured(cands),
                entries=await build_entries(r, None, None),
            ))
            continue

        scored = sorted(
            (
                (score, u)
                for u_toks, u in name_index_full
                if (score := _fuzzy_score(row_toks, u_toks)) >= 0.66
            ),
            key=lambda x: -x[0],
        )
        if scored and (len(scored) == 1 or scored[0][0] - scored[1][0] > 0.15):
            cands = [u for _, u in scored]
            result.deferred.append(PersonRow(
                email=None, name=r.name, user_id=None, user_full_name=None,
                note="close name match — confirm the suggested account",
                similar=[
                    f"{getattr(u, 'full_name', '')} <{getattr(u, 'email', None) or '-'}>"
                    for u in cands[:5]
                ],
                similar_users=_rank_structured(cands),
                entries=await build_entries(r, None, None),
            ))
            continue

        pr = PersonRow(
            email=None, name=r.name, user_id=None, user_full_name=None,
            note="no email", similar=_similar(r.name),
            similar_users=_similar_structured(r.name),
            entries=await build_entries(r, None, None),
        )
        warn = _crewid_warn(r)
        if warn:
            for e in pr.entries:
                e.warning = f"{e.warning}; {warn}" if e.warning else warn
        result.deferred.append(pr)

    groups_by_email: dict[str, list[EmpRow]] = {}
    for r in leftover:
        if r.email:
            groups_by_email.setdefault(r.email, []).append(r)

    for email, rows in groups_by_email.items():
        first_names = {_first_name_token(r.name) for r in rows}
        same_person = len(first_names) == 1
        rep_name = rows[0].name
        db_users = users_by_email.get(email, [])

        if _is_placeholder_email(email) or not same_person:
            reason = ("internal email — shared placeholder" if _is_placeholder_email(email)
                      else "shared email — multiple people")
            unique_rows = list({(r.name, r.emp_id, r.company): r for r in rows}.values())
            members = [f"{r.name} = {r.emp_id} ({r.company})" for r in unique_rows]
            ph_entries = []
            for r in unique_rows:
                for e in await build_entries(r, None, None):
                    warn = _crewid_warn(r)
                    if warn:
                        e.warning = f"{e.warning}; {warn}" if e.warning else warn
                    ph_entries.append(e)
            result.placeholder.append(PersonRow(
                email=email, name=rep_name, user_id=None, user_full_name=None,
                note=reason, members=members, entries=ph_entries,
            ))
            continue

        if not db_users:
            df_entries = []
            for r in rows:
                for e in await build_entries(r, None, None):
                    warn = _crewid_warn(r)
                    if warn:
                        e.warning = f"{e.warning}; {warn}" if e.warning else warn
                    df_entries.append(e)
            result.deferred.append(PersonRow(
                email=email, name=rep_name, user_id=None, user_full_name=None,
                note="email present, no DB user", similar=_similar(rep_name),
                similar_users=_similar_structured(rep_name),
                entries=df_entries,
            ))
            continue

        user = db_users[0]
        member_id = member_by_user.get(user.id)
        merged = crewid_persons.get(str(user.id))
        person = merged if merged is not None else PersonRow(
            email=email, name=rep_name,
            user_id=str(user.id), user_full_name=user.full_name,
        )
        if member_id is None:
            person.note = "no org membership (legacy account) — cannot assign"
            person.entries = []
            for r in rows:
                for e in await build_entries(r, None, None):
                    e.action = ACTION_INVALID
                    e.warning = "no org_member row"
                    person.entries.append(e)
            if merged is None:
                result.people.append(person)
            continue

        seen_pairs: set[tuple[str, str]] = person_seen.get(str(user.id), set())
        confirm_rows: list[EmpRow] = []
        for r in rows:
            pair = (r.company, r.emp_id)
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)
            if _crewid_warn(r):
                confirm_rows.append(r)
                continue
            person.entries.extend(await build_entries(r, member_id, user))

        if confirm_rows:
            cf_entries = []
            for r in confirm_rows:
                for e in await build_entries(r, None, None):
                    warn = _crewid_warn(r)
                    if warn:
                        e.warning = f"{e.warning}; {warn}" if e.warning else warn
                    cf_entries.append(e)
            result.deferred.append(PersonRow(
                email=email, name=rep_name, user_id=None, user_full_name=None,
                note="CREWID mismatch — email matches an account, confirm before registering",
                similar=[f"{user.full_name} <{email}>"],
                similar_users=[{
                    "user_id": str(user.id),
                    "full_name": user.full_name,
                    "email": email,
                }],
                entries=cf_entries,
            ))

        # 같은 매장에 서로 다른 두 값이 오는 경우 — 뒤 값에 경고
        by_store: dict[str, int] = {}
        for e in person.entries:
            if e.store_id and e.emp_id is not None:
                if e.store_id in by_store and by_store[e.store_id] != e.emp_id:
                    e.action = ACTION_INVALID
                    e.warning = "conflicting numbers for the same store in file"
                else:
                    by_store[e.store_id] = e.emp_id
        if merged is None and person.entries:
            result.people.append(person)

    # ── 매장 미매칭 원문 집계 ────────────────────────────────────────
    agg: dict[str, dict] = {}
    for p in list(result.people) + list(result.placeholder) + list(result.deferred):
        for e in p.entries:
            if e.action != ACTION_UNMATCHED_STORE:
                continue
            key = _norm_key(e.corp_abr) or _norm_key(e.company) or "(blank)"
            row = agg.setdefault(key, {"key": key, "company": e.company,
                                       "corp_abr": e.corp_abr, "rows": 0})
            row["rows"] += 1
    result.unmatched_stores = sorted(agg.values(), key=lambda x: -x["rows"])

    # ── 양측 대조 (reconciliation) — 사람 단위 ───────────────────────
    # 번호 단위로 비교하면 "8→1001 rebind 대기" 인 사람이 htm_only(#8)와
    # file_only(#1001) 양쪽에 다 찍혀 모순처럼 보인다. 사람 기준 3분류:
    #   matched        — 파일↔HTM 매칭된 사람 (매장별 current→new 전이 요약)
    #   htm_unmatched  — HTM 에 등록돼 있는데 파일이 못 덮은 사람 (번호 직접 지정 대상)
    #   file_unmatched — 파일에만 있는 행 (deferred/placeholder 에서 사람 해결)
    def _scope_ident_of_entry(e: ImportEntry):
        if e.group_id:
            return ("group", e.group_id)
        if e.store_id:
            gid = store_group_of.get(e.store_id)
            if gid in engaged_groups:
                return ("group", gid)
            return ("store", e.store_id)
        return None

    matched_by_scope: dict[tuple[str, str], dict[str, dict]] = {}
    for person in result.people:
        if not person.user_id:
            continue
        for e in person.entries:
            if e.emp_id is None or e.action == ACTION_UNMATCHED_STORE:
                continue
            ident = _scope_ident_of_entry(e)
            if ident is None:
                continue
            bucket = matched_by_scope.setdefault(ident, {})
            row = bucket.setdefault(person.user_id, {
                "user_id": person.user_id,
                "name": person.user_full_name or person.name,
                "changes": [],
            })
            row["changes"].append({
                "store_id": e.store_id,
                "store_name": e.store_name,
                "current": e.current_empid,
                "new": e.emp_id,
                "pending_store": e.action == ACTION_NEEDS_STORE,
            })

    file_unmatched_by_scope: dict[tuple[str, str], list[dict]] = {}
    for p in list(result.placeholder) + list(result.deferred):
        for e in p.entries:
            if e.emp_id is None:
                continue
            ident = _scope_ident_of_entry(e)
            if ident is None:
                continue
            file_unmatched_by_scope.setdefault(ident, []).append({
                "empid": e.emp_id,
                "name": e.person_name or p.name,
            })

    def _htm_unmatched(ident: tuple[str, str], scope_store_ids: set, store_names: dict) -> list[dict]:
        """스코프 매장에 배정돼 있는데 매칭 안 된 사람 — 번호 지정 대상.

        번호가 없는 배정자도 포함한다 (그들이야말로 번호 지정이 필요한 사람).
        """
        matched_users = set(matched_by_scope.get(ident, {}))
        out: list[dict] = []
        for (m_id, s_id), emp in empid_map.items():
            if s_id not in scope_store_ids:
                continue
            uid = member_user.get(m_id)
            if uid is None or str(uid) in matched_users:
                continue
            out.append({
                "user_id": str(uid),
                "name": user_name_by_id.get(uid, "?"),
                "store_id": str(s_id),
                "store_name": store_names.get(s_id, "?"),
                "current_empid": emp,
            })
        out.sort(key=lambda x: (x["store_name"], x["name"].lower()))
        return out

    def _emit(ident: tuple[str, str], name: str, scope_store_ids: set, store_names: dict) -> None:
        matched = sorted(
            matched_by_scope.get(ident, {}).values(),
            key=lambda x: x["name"].lower(),
        )
        file_rows = sorted(
            file_unmatched_by_scope.get(ident, []), key=lambda x: x["empid"],
        )
        result.reconciliation.append({
            "scope": ident[0], "id": ident[1], "name": name,
            "matched": matched,
            "htm_unmatched": _htm_unmatched(ident, scope_store_ids, store_names),
            "file_unmatched": file_rows,
        })

    emitted_store_ids: set[str] = set()
    for gid, gscope in engaged_groups.items():
        sids = {st.id for st in gscope.stores}
        emitted_store_ids.update(str(st.id) for st in gscope.stores)
        _emit(("group", gid), gscope.name, sids,
              {st.id: st.name for st in gscope.stores})
    store_idents = (
        {i for i in matched_by_scope if i[0] == "store"}
        | {i for i in file_unmatched_by_scope if i[0] == "store"}
    )
    for ident in store_idents:
        sid = ident[1]
        if sid in emitted_store_ids:
            continue
        st = store_by_id.get(sid)
        if st is None:
            continue
        _emit(ident, st.name, {st.id}, {st.id: st.name})
    result.reconciliation.sort(key=lambda x: (x["scope"], x["name"]))
    result.distribution = _build_distribution(result)

    return result


def _build_distribution(result: ImportPreview) -> list[dict]:
    """업로드 파일의 유효 번호를 백 단위로 묶어 [{band, lo, hi, count}] (오름차순).

    세 버킷(people/placeholder/deferred) 전체의 정수화된 emp_id 를 센다 — 아직 사람이
    확정되지 않은 행도 "파일이 어떤 대역을 쓰는가" 의 일부이기 때문이다.
    번호 없는/정수화 실패 행은 세지 않는다 (분포는 번호대 이야기다).
    """
    buckets: dict[int, dict] = {}
    for person in list(result.people) + list(result.placeholder) + list(result.deferred):
        for e in person.entries:
            if e.emp_id is None:
                continue
            lo, hi, label = empid_band(e.emp_id)
            row = buckets.setdefault(lo, {"band": label, "lo": lo, "hi": hi, "count": 0})
            row["count"] += 1
    return [buckets[k] for k in sorted(buckets)]


def _build_entry(
    r: EmpRow,
    store_index: dict[str, list[Store]],
    member_id: UUID | None,
    user: object | None,
    empid_map: dict[tuple[UUID, UUID], int | None],
    work_map: dict[tuple[UUID, UUID], bool] | None = None,
) -> ImportEntry:
    """EmpRow 1건 → ImportEntry (매장 매칭 + 현재값 비교 + 액션 결정)."""
    store = match_store(store_index, r.company, r.corp_abr)
    emp_int = _emp_id_int(r.emp_id)
    if store is None:
        return ImportEntry(
            store_id=None, store_name=None, company=r.company,
            emp_id_raw=r.emp_id, emp_id=emp_int, current_empid=None,
            has_assignment=False, action=ACTION_UNMATCHED_STORE,
            warning="no matching store in this org", person_name=r.name,
            corp_abr=r.corp_abr,
        )
    if emp_int is None:
        return ImportEntry(
            store_id=str(store.id), store_name=store.name, company=r.company,
            emp_id_raw=r.emp_id, emp_id=None, current_empid=None,
            has_assignment=False, action=ACTION_INVALID,
            warning="emp_id is not a positive integer", person_name=r.name,
            corp_abr=r.corp_abr,
        )
    if member_id is None:
        # 매장·번호는 유효, 유저만 미확정 — 운영자가 picker 로 유저를 골라 등록 가능.
        return ImportEntry(
            store_id=str(store.id), store_name=store.name, company=r.company,
            emp_id_raw=r.emp_id, emp_id=emp_int, current_empid=None,
            has_assignment=False, action=ACTION_NEEDS_USER, person_name=r.name,
            corp_abr=r.corp_abr,
        )
    key = (member_id, store.id)
    has_row = key in empid_map
    current = empid_map.get(key)
    if not has_row:
        action = ACTION_NEW_ASSIGNMENT
    elif current == emp_int:
        action = ACTION_SAME
    else:
        action = ACTION_REBIND
    # 휴면 배정(과거 배정 해제 or 근무배정 제외) — 번호는 기록되지만 재활성화하지 않음을 표시.
    dormant = has_row and not (work_map or {}).get(key, True)
    return ImportEntry(
        store_id=str(store.id), store_name=store.name, company=r.company,
        emp_id_raw=r.emp_id, emp_id=emp_int, current_empid=current,
        has_assignment=has_row, action=action,
        dormant=dormant, corp_abr=r.corp_abr,
        warning=(
            "assignment is dormant (not in work assignment) — number is written but the person stays inactive"
            if dormant else None
        ),
    )


async def roster(db: AsyncSession, organization_id: UUID) -> list[dict]:
    """매장별 배정·empid 현황 — bulk 에디터·export 필터의 데이터 소스.

    live 매장(sort_order 순)별로 배정 인원과 현재 empid + 역할/부서를 나열한다.
    (역할·부서는 export 필터 축 — SV/Manager 등 role, FOH/BOH department.)
    """
    from app.models.user import Role, User

    stores = await store_repository.get_by_org(db, organization_id, include_closed=False)
    if not stores:
        return []
    store_ids = [s.id for s in stores]
    rows = (
        await db.execute(
            select(
                OrgMemberStore.store_id,
                OrgMemberStore.empid,
                OrgMemberStore.empid_kind,
                OrgMemberStore.is_work_assignment,
                OrgMemberStore.is_manager,
                OrgMember.user_id,
                OrgMember.crewid,
                OrgMember.department,
                User.full_name,
                User.email,
                User.is_active,
                Role.name.label("role_name"),
                Role.priority.label("role_priority"),
            )
            .join(OrgMember, OrgMember.id == OrgMemberStore.org_member_id)
            .join(User, User.id == OrgMember.user_id)
            .join(Role, Role.id == OrgMember.role_id)
            .where(
                OrgMemberStore.store_id.in_(store_ids),
                OrgMember.organization_id == organization_id,
            )
        )
    ).all()
    by_store: dict[UUID, list[dict]] = {}
    for r in rows:
        by_store.setdefault(r.store_id, []).append({
            "user_id": str(r.user_id),
            "full_name": r.full_name,
            "email": r.email,
            "empid": r.empid,
            # 번호 구분 — 콘솔 로스터의 조용한 컬럼/필터용. empid 가 없으면 무의미하나
            # 컬럼 타입을 흔들지 않도록 저장값(기본 sequence)을 그대로 낸다 (계약 §3-6).
            "empid_kind": r.empid_kind,
            "is_work_assignment": r.is_work_assignment,
            "is_manager": r.is_manager,
            # 계정 활성 여부 — 비활성자도 번호는 점유하므로 로스터에는 남기고,
            # export 후보에서 거를지는 콘솔 필터 몫.
            "is_active": r.is_active,
            "crewid": r.crewid,
            "role_name": r.role_name,
            "role_priority": r.role_priority,
            "department": r.department,
        })
    out: list[dict] = []
    for s in stores:
        members = by_store.get(s.id, [])
        # empid 오름차순 (없는 사람은 뒤), 그 다음 이름순
        members.sort(key=lambda m: (m["empid"] is None, m["empid"] or 0, m["full_name"] or ""))
        out.append({
            "store_id": str(s.id),
            "store_name": s.name,
            "group_id": str(s.group_id) if s.group_id else None,
            "members": members,
        })
    return out


def _sheet_title(raw: str) -> str:
    """Excel 시트명 제약 대응 — 금지문자 제거 + 31자 제한."""
    cleaned = re.sub(r"[\[\]:*?/\\]", "-", raw).strip() or "Sheet"
    return cleaned[:31]


def _roster_export_sheets(sheets: list[tuple[str, list[list]]]) -> bytes:
    """임포트 형식 xlsx 생성 — 데이터 시트들(구분 export 시 복수) + Instructions.

    parse_emplist 는 첫 번째 시트만 읽으므로 데이터 시트가 반드시 먼저다.
    (시트를 나눈 파일을 재업로드하면 첫 시트만 임포트됨 — Instructions 에 명시.)
    """
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    first = True
    used_titles: set[str] = set()
    for title, rows in sheets or [("Roster", [])]:
        base = _sheet_title(title)
        name = base
        n = 2
        while name in used_titles:  # 시트명 중복 방지
            name = _sheet_title(f"{base[:28]}-{n}")
            n += 1
        used_titles.add(name)
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        ws.append(["COMPANY", "CORP_ABR_3", "Name", "emp_id", "Email", "crewid"])
        for c in ws[1]:
            c.font = Font(bold=True)
        for r in rows:
            ws.append(r)
        for col, width in zip("ABCDEF", (34, 12, 26, 10, 32, 10)):
            ws.column_dimensions[col].width = width
        ws.freeze_panes = "A2"

    info = wb.create_sheet("Instructions")
    for line in (
        ("EMPID import format",),
        (),
        ("COMPANY", "Store name (matched against this org's store names)"),
        ("CORP_ABR_3", "Store code (fallback match, e.g. IFO / SWC)"),
        ("Name", "Person's name (display only — matching uses Email)"),
        ("emp_id", "Number to register for that person at that store. Rows with an empty emp_id are skipped."),
        ("Email", "Person's email — used to match the user in the system"),
        ("crewid", "Optional exact-match key — the person's org number (CREWID). When present, the person is matched by CREWID even if email/name differ (useful when staff signed up with a different email)."),
        (),
        ("Tip", "Export current EMPIDs, edit the emp_id column, then upload the file back on the EMPID Import tab."),
        ("Note", "Import reads the FIRST sheet only — if this file was exported split into multiple sheets, merge rows into the first sheet (or re-export unsplit) before re-uploading."),
    ):
        info.append(list(line))
    info.column_dimensions["A"].width = 14
    info.column_dimensions["B"].width = 90
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def build_template_xlsx(
    db: AsyncSession,
    organization_id: UUID,
    prefill: bool,
    *,
    store_ids: set[UUID] | None = None,
    people: str = "all",  # all | numbered | unnumbered
    include_dormant: bool = True,
    include_email: bool = True,
    include_numbers: bool = True,
) -> bytes:
    """임포트용 템플릿/현황 export xlsx.

    prefill=False: 헤더 + 안내 시트만 (빈 템플릿, 필터 무관).
    prefill=True: 현재 배정·번호 export — 필터로 범위를 좁힐 수 있다:
      store_ids: 매장 부분집합 (None=전체) / people: 번호 유무 필터
      include_dormant: 휴면(근무배정 제외) 포함 여부
      include_email: Email 컬럼 값 포함 여부 — 빼면 재업로드 매칭 불가(공유용)
      include_numbers: emp_id 값 포함 여부 — 빼면 번호 비운 작성용 양식
    헤더 5컬럼은 항상 유지 (임포트 파서 형식 보존).
    """
    rows: list[list] = []
    if prefill:
        stores = await store_repository.get_by_org(db, organization_id, include_closed=False)
        code_by_id = {str(s.id): (s.code or "") for s in stores}
        for st in await roster(db, organization_id):
            if store_ids is not None and UUID(st["store_id"]) not in store_ids:
                continue
            for m in st["members"]:
                if not include_dormant and not m["is_work_assignment"]:
                    continue
                if people == "numbered" and m["empid"] is None:
                    continue
                if people == "unnumbered" and m["empid"] is not None:
                    continue
                rows.append([
                    st["store_name"],
                    code_by_id.get(st["store_id"], ""),
                    m["full_name"],
                    (m["empid"] if m["empid"] is not None else "") if include_numbers else "",
                    (m["email"] or "") if include_email else "",
                    m["crewid"] if m["crewid"] is not None else "",
                ])
    return _roster_export_sheets([("Roster", rows)])


async def build_selected_export_xlsx(
    db: AsyncSession,
    organization_id: UUID,
    selected: list[tuple[UUID, UUID]],  # (user_id, store_id) — 콘솔에서 사람 단위로 확정한 목록
    *,
    include_email: bool = True,
    include_numbers: bool = True,
    split_by: str = "none",  # none | store | role | band(백 단위 번호대)
) -> bytes:
    """사람 단위 선택 export — 콘솔이 필터·개별 넣고빼기로 확정한 (user, store) 목록을 그대로 굽는다.

    필터 축(역할/부서/번호대/매장/개인)은 전부 클라이언트 몫 — 서버는 선택 결과만 받아
    어떤 새 축이 생겨도 변경이 필요 없다. split_by 로 시트를 구분(1차/2차식 배포용):
    store=매장별, role=역할별, band=백 단위 번호대별. 재업로드는 첫 시트만 읽힌다.
    """
    wanted = set(selected)
    if not wanted:
        return _roster_export_sheets([("Roster", [])])
    stores = await store_repository.get_by_org(db, organization_id, include_closed=False)
    code_by_id = {str(s.id): (s.code or "") for s in stores}

    entries: list[dict] = []
    for st in await roster(db, organization_id):
        for m in st["members"]:
            if (UUID(m["user_id"]), UUID(st["store_id"])) not in wanted:
                continue
            entries.append({
                "row": [
                    st["store_name"],
                    code_by_id.get(st["store_id"], ""),
                    m["full_name"],
                    (m["empid"] if m["empid"] is not None else "") if include_numbers else "",
                    (m["email"] or "") if include_email else "",
                    m["crewid"] if m["crewid"] is not None else "",
                ],
                "store_name": st["store_name"],
                "role_name": m["role_name"],
                "empid": m["empid"],
            })

    if split_by == "store":
        keys: list[str] = []
        grouped: dict[str, list[list]] = {}
        for e in entries:
            k = e["store_name"]
            if k not in grouped:
                grouped[k] = []
                keys.append(k)
            grouped[k].append(e["row"])
        sheets = [(k, grouped[k]) for k in keys]
    elif split_by == "role":
        keys = []
        grouped = {}
        for e in entries:
            k = e["role_name"] or "No role"
            if k not in grouped:
                grouped[k] = []
                keys.append(k)
            grouped[k].append(e["row"])
        sheets = [(k, grouped[k]) for k in keys]
    elif split_by == "band":
        # 백 단위 번호대 — 1-99, 100-199, 300-399 … 번호 없는 사람은 "No number" 시트 마지막
        def band_key(empid: int | None) -> tuple[int, str]:
            if empid is None:
                return (10**9, "No number")
            lo, _hi, label = empid_band(empid)
            return (lo, label)

        tagged = sorted(
            ((band_key(e["empid"]), e["row"]) for e in entries), key=lambda t: t[0][0]
        )
        keys = []
        grouped = {}
        for (order, label), row in tagged:
            if label not in grouped:
                grouped[label] = []
                keys.append(label)
            grouped[label].append(row)
        sheets = [(k, grouped[k]) for k in keys]
    else:
        sheets = [("Roster", [e["row"] for e in entries])]
    return _roster_export_sheets(sheets)


@dataclass
class CommitAssignment:
    """commit 요청 1건 — (사람, 매장) 의 번호 + 번호 구분 + 사유.

    empid=None 은 번호 삭제. empid_kind 는 **요청이 말하는 값을 그대로** 쓴다 —
    경로(임포트 탭/bulk 에디터/스태프 상세)로 추론하지 않는다 (INV-6).
    """

    user_id: UUID
    store_id: UUID
    empid: int | None = None
    empid_kind: str = EMPID_KIND_SEQUENCE
    reason: str | None = None


@dataclass
class CommitResult:
    """commit 결과 — 반영/재채번/스킵/거절 내역 + 예외 건수/커밋 후 커서."""

    applied: list[dict] = field(default_factory=list)      # {user, store, empid, created}
    renumbered: list[dict] = field(default_factory=list)   # {user, store, old, new}
    skipped: list[dict] = field(default_factory=list)      # {user, store, empid, reason}
    rejected: list[dict] = field(default_factory=list)     # {user_id, store_id, reason}
    # 이번 커밋이 예외(exception)로 기입한 건수 — 콘솔 요약 "예외 N건 제외 → 커서 X"
    exception_count: int = 0
    # 커밋이 건드린 스코프의 **커밋 후 커서** — {scope_id: next_empid}
    cursor_after: dict[str, int] = field(default_factory=dict)


async def _cursor_after(db: AsyncSession, store_ids: list[UUID]) -> dict[str, int]:
    """커밋이 건드린 스코프들의 현재 커서 — {scope_id: next_empid}.

    commit(=수동 기입)은 커서를 전진시키지 않는다(INV-5). 이 값은 "그래서 지금 커서가
    얼마인가" 를 조용히 바꾸지 않고 운영자에게 그대로 알려주기 위한 것이고, 어긋났다면
    운영자가 커서 재계산으로 맞춘다.

    스코프 판정·커서 읽기는 전부 org_numbering(단일 게이트웨이)에 맡긴다.
    """
    out: dict[str, int] = {}
    seen: set[UUID] = set()
    for store_id in store_ids:
        state = await empid_cursor_state(db, store_id=store_id)
        if state.scope_id in seen:  # 그룹 공유 — 형제 매장들이 같은 커서를 가리킨다
            continue
        seen.add(state.scope_id)
        if state.next_empid is not None:  # 미초기화(NULL) 스코프는 알릴 값이 없다
            out[str(state.scope_id)] = state.next_empid
    return out


async def commit(
    db: AsyncSession,
    organization_id: UUID,
    assignments: list[CommitAssignment] | list[tuple[UUID, UUID, int | None]],
    actor_id: UUID | None = None,
) -> CommitResult:
    """확정 — 매장 단위 3-phase 로 empid 재기입. 단일 트랜잭션, 멱등.

    empid=None 은 번호 삭제(비우기) — 배정 행은 유지, 번호만 해제되어 재사용 가능.
    users.employee_no 는 기록하지 않는다 (폐기 방향).
    변경 1건마다 empid_changes 이력을 남긴다 (actor_id = 실행 운영자).

    assignments 는 CommitAssignment 목록. (user_id, store_id, empid) 3-튜플도 받는다
    (기존 호출부 하위호환 — 그때 empid_kind 는 기본값 sequence).
    """
    result = CommitResult()
    items: list[CommitAssignment] = [
        a if isinstance(a, CommitAssignment) else CommitAssignment(*a)
        for a in assignments
    ]

    member_by_user, empid_map, _work_map, kind_map = await _current_empid_map(
        db, organization_id
    )
    member_user: dict[UUID, UUID] = {m: u for u, m in member_by_user.items()}

    # org 매장 검증용 (폐점 포함 — 번호는 폐점 매장에도 존재)
    stores = await store_repository.get_by_org(db, organization_id, include_closed=True)
    store_by_id = {s.id: s for s in stores}
    user_names: dict[UUID, str] = {}
    for u in await user_repository.get_by_org(db, organization_id):
        user_names[u.id] = u.full_name

    def _label(member_id: UUID, store_id: UUID) -> tuple[str, str]:
        user_id = member_user.get(member_id)
        return (
            user_names.get(user_id, str(user_id)) if user_id else str(member_id),
            store_by_id[store_id].name,
        )

    channel = current_channel()

    def _log_change(
        member_id: UUID, store_id: UUID,
        old: int | None, new: int | None, source: str,
        reason: str | None = None,
    ) -> None:
        """empid 변경 이력 1건 — 같은 트랜잭션에 적재 (commit 실패 시 함께 롤백)."""
        if old == new:
            return
        user_id = member_user.get(member_id)
        db.add(EmpidChange(
            organization_id=organization_id,
            store_id=store_id,
            store_name=store_by_id[store_id].name,
            user_id=user_id,
            person_name=user_names.get(user_id) if user_id else None,
            old_empid=old, new_empid=new, reason=reason,
            source=source, channel=channel, changed_by=actor_id,
        ))

    # 매장별 mapping 구성 — {store_id: {org_member_id: empid}}.
    # claims 는 매장별 역맵(empid→member) — 같은 매장에 같은 번호를 두 사람이 요청하면
    # 현재 보유자를 우선하고 나머지는 거절한다 (전체 롤백/무단 재채번 방지).
    # kinds/reasons 는 mapping 과 같은 키(store_id → member_id)를 쓰는 병렬 맵 —
    # 3-phase 비교 로직(값 충돌·멱등)은 번호만 보므로 mapping 의 모양을 건드리지 않는다.
    per_store: dict[UUID, dict[UUID, int | None]] = {}
    kinds: dict[UUID, dict[UUID, str]] = {}
    reasons: dict[UUID, dict[UUID, str | None]] = {}
    claims: dict[UUID, dict[int, UUID]] = {}
    for item in items:
        user_id, store_id, empid = item.user_id, item.store_id, item.empid
        if store_id not in store_by_id:
            result.rejected.append({"user_id": str(user_id), "store_id": str(store_id),
                                    "reason": "store not in this org"})
            continue
        member_id = member_by_user.get(user_id)
        if member_id is None:
            result.rejected.append({"user_id": str(user_id), "store_id": str(store_id),
                                    "reason": "no org membership"})
            continue
        if empid is not None and empid < 1:
            result.rejected.append({"user_id": str(user_id), "store_id": str(store_id),
                                    "reason": "empid must be >= 1"})
            continue
        if item.empid_kind not in EMPID_KINDS:
            result.rejected.append({"user_id": str(user_id), "store_id": str(store_id),
                                    "reason": "empid_kind must be sequence|exception"})
            continue
        mapping = per_store.setdefault(store_id, {})
        store_kinds = kinds.setdefault(store_id, {})
        store_reasons = reasons.setdefault(store_id, {})
        store_claims = claims.setdefault(store_id, {})
        if member_id in mapping and mapping[member_id] != empid:
            result.rejected.append({"user_id": str(user_id), "store_id": str(store_id),
                                    "reason": "conflicting values for the same store in request"})
            continue
        if empid is None:
            # 번호 삭제 — 값 충돌 검사 대상 아님. 배정 행이 없으면 할 일 없음(스킵).
            if (member_id, store_id) not in empid_map or empid_map[(member_id, store_id)] is None:
                name, store_name = _label(member_id, store_id)
                result.skipped.append({"user": name, "store": store_name,
                                       "empid": None, "reason": "nothing to clear"})
                continue
            mapping[member_id] = None
            # 번호가 사라지면 구분은 의미를 잃는다 → 기본값으로 되돌린다.
            store_kinds[member_id] = EMPID_KIND_SEQUENCE
            store_reasons[member_id] = item.reason
            continue
        other = store_claims.get(empid)
        if other is not None and other != member_id:
            if empid_map.get((member_id, store_id)) == empid:
                # 이번 요청자가 현재 보유자 — 앞서 등록된 다른 사람을 밀어내고 거절.
                del mapping[other]
                store_kinds.pop(other, None)
                store_reasons.pop(other, None)
                other_user = member_user.get(other)
                result.rejected.append({
                    "user_id": str(other_user) if other_user else str(other),
                    "store_id": str(store_id),
                    "reason": "duplicate empid for this store in request",
                })
            else:
                result.rejected.append({"user_id": str(user_id), "store_id": str(store_id),
                                        "reason": "duplicate empid for this store in request"})
                continue
        store_claims[empid] = member_id
        mapping[member_id] = empid
        store_kinds[member_id] = item.empid_kind
        store_reasons[member_id] = item.reason

    try:
        # Pass 1 — 매장별: 멱등 스킵 → 스코프 락 → 비우기 → 기입. 재채번은 미룬다.
        # (그룹 공유 스코프에서 형제 매장의 기입값이 재채번 max 에 반영되도록 2-pass.)
        deferred: list[tuple[UUID, dict[UUID, int], list[OrgMemberStore]]] = []
        for store_id, mapping in per_store.items():
            store_kinds = kinds.get(store_id, {})
            store_reasons = reasons.get(store_id, {})
            # 이미 같은 값이면 skip (멱등) — mapping 에서 제거.
            # 단, 번호는 그대로인데 **구분만** 바뀌는 요청은 스킵이 아니다 (스태프 상세에서
            # "이 번호를 예외로 표시" 하는 경로가 조용히 무시되면 안 된다).
            for member_id in list(mapping.keys()):
                value = mapping[member_id]
                if empid_map.get((member_id, store_id)) != value:
                    continue
                name, store_name = _label(member_id, store_id)
                want_kind = store_kinds.get(member_id, EMPID_KIND_SEQUENCE)
                if value is not None and kind_map.get((member_id, store_id)) != want_kind:
                    await db.execute(
                        update(OrgMemberStore)
                        .where(
                            OrgMemberStore.org_member_id == member_id,
                            OrgMemberStore.store_id == store_id,
                        )
                        .values(empid_kind=want_kind)
                    )
                    result.applied.append({"user": name, "store": store_name,
                                           "empid": value, "created": False})
                    if want_kind != EMPID_KIND_SEQUENCE:
                        result.exception_count += 1
                else:
                    result.skipped.append({"user": name, "store": store_name,
                                           "empid": value,
                                           "reason": "already set"})
                del mapping[member_id]
            if not mapping:
                continue

            # 그룹 공유 스코프면 advisory lock 선취 — 커밋까지 유지되어 동시 채번과 직렬화.
            await lock_empid_scope(db, store_id)

            rows = (
                await db.execute(
                    select(OrgMemberStore).where(OrgMemberStore.store_id == store_id)
                )
            ).scalars().all()
            by_member = {row.org_member_id: row for row in rows}
            # None(번호 삭제)은 값 충돌 대상 아님
            wanted_values = {v for v in mapping.values() if v is not None}

            # phase 1 — 비우기: 대상 행 + 원하는 값을 점유 중인 행의 empid 를 NULL 로.
            cleared: list[OrgMemberStore] = []
            for row in rows:
                if row.empid is None:
                    continue
                if row.org_member_id in mapping or row.empid in wanted_values:
                    cleared.append(row)
                    row.empid = None
            await db.flush()

            # phase 2 — 기입: 행이 있으면 갱신, 없으면 배정 생성(신규 배정).
            for member_id, value in mapping.items():
                name, store_name = _label(member_id, store_id)
                row = by_member.get(member_id)
                # 번호 삭제면 구분도 기본값으로. 그 외엔 요청값 그대로 (경로 추론 없음).
                want_kind = (
                    EMPID_KIND_SEQUENCE if value is None
                    else store_kinds.get(member_id, EMPID_KIND_SEQUENCE)
                )
                reason = store_reasons.get(member_id)
                if value is not None and want_kind != EMPID_KIND_SEQUENCE:
                    result.exception_count += 1
                if row is None:
                    db.add(OrgMemberStore(
                        org_member_id=member_id, store_id=store_id,
                        is_manager=False, is_work_assignment=True, empid=value,
                        empid_kind=want_kind,
                    ))
                    result.applied.append({"user": name, "store": store_name,
                                           "empid": value, "created": True})
                    _log_change(member_id, store_id, None, value, EMPID_SOURCE_COMMIT,
                                reason=reason)
                else:
                    row.empid = value
                    row.empid_kind = want_kind
                    result.applied.append({"user": name, "store": store_name,
                                           "empid": value, "created": False})
                    _log_change(
                        member_id, store_id,
                        empid_map.get((member_id, store_id)), value,
                        EMPID_SOURCE_COMMIT, reason=reason,
                    )
            await db.flush()
            deferred.append((store_id, mapping, cleared))

        # Pass 2 — 전 매장 기입 완료 후 재채번: 번호를 뺏겼는데 mapping 에 없는 기존 인원.
        # 이 시점의 그룹 max 는 형제 매장의 phase 2 기입값을 포함하므로 임포트가
        # 스스로 그룹 스코프 중복을 만들 수 없다.
        for store_id, mapping, cleared in deferred:
            for row in cleared:
                if row.org_member_id in mapping:
                    continue
                old = empid_map.get((row.org_member_id, store_id))
                row.empid = await next_empid(db, store_id)
                row.empid_kind = EMPID_KIND_SEQUENCE  # 커서가 준 순번 — 예외가 아니다
                name, store_name = _label(row.org_member_id, store_id)
                result.renumbered.append({"user": name, "store": store_name,
                                          "old": old, "new": row.empid})
                _log_change(row.org_member_id, store_id, old, row.empid,
                            EMPID_SOURCE_RENUMBER)
            await db.flush()

        await db.commit()
        # 커밋이 끝난 뒤의 커서를 그대로 알려준다 — 조용히 바꾸지 않는다 (INV-5).
        result.cursor_after = await _cursor_after(db, list(per_store.keys()))
    except IntegrityError as e:
        await db.rollback()
        raise DuplicateError(
            "empid assignment conflict (concurrent change) — please retry"
        ) from e
    except Exception:
        await db.rollback()
        raise
    return result
