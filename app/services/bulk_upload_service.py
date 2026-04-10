"""Bulk upload 서비스 — 직원 CSV + 스케줄 Excel 대량 등록."""

import csv
import io
import re
from datetime import date, timedelta
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.organization import Store
from app.models.user import Role, User
from app.models.user_store import UserStore
from app.schemas.schedule import ScheduleCreate
from app.schemas.user import UserCreate
from app.services.schedule_service import schedule_service
from app.services.user_service import user_service
from app.utils.exceptions import BadRequestError


class BulkUploadService:

    # ── Dispatch (CSV or Excel) ───────────────────────────

    async def process_employees(
        self, db: AsyncSession, organization_id: UUID,
        file_bytes: bytes, caller: User, filename: str = "",
    ) -> dict:
        if filename.endswith(".xlsx"):
            rows = self._parse_excel_rows(file_bytes)
        else:
            rows = self._parse_csv(file_bytes)
        return await self._process_employee_rows(db, organization_id, rows, caller)

    async def process_schedules(
        self, db: AsyncSession, organization_id: UUID,
        file_bytes: bytes, created_by: UUID, filename: str = "",
    ) -> dict:
        if filename.endswith(".xlsx"):
            return await self.process_schedules_excel(db, organization_id, file_bytes, created_by)
        else:
            return await self.process_schedules_csv(db, organization_id, file_bytes, created_by)

    # ── Employee ──────────────────────────────────────────

    async def _process_employee_rows(
        self,
        db: AsyncSession,
        organization_id: UUID,
        rows: list[dict],
        caller: User,
    ) -> dict:
        """직원 rows를 처리하여 등록 + 매장 배정.

        Row keys: username, password, full_name, role, store_name, email, hourly_rate
        """
        # "id" 컬럼을 "username"으로 정규화
        for row in rows:
            if "id" in row and "username" not in row:
                row["username"] = row.pop("id")

        required = {"username", "password", "full_name", "role", "store_name"}
        if rows and not required.issubset(rows[0].keys()):
            missing = required - set(rows[0].keys())
            raise BadRequestError(f"Missing columns: {', '.join(missing)}")

        roles_map = await self._load_roles_map(db, organization_id)
        stores_map = await self._load_stores_map(db, organization_id)

        created = 0
        skipped = 0
        store_assigned = 0
        errors: list[str] = []
        users_list: list[dict] = []  # 결과에 유저 목록 포함

        for i, row in enumerate(rows, start=2):
            username = row.get("username", "").strip()
            password = row.get("password", "").strip()
            full_name = row.get("full_name", "").strip()
            role_name = row.get("role", "").strip().lower()
            store_names_raw = row.get("store_name", "").strip()
            email = row.get("email", "").strip() or None
            hourly_rate_str = row.get("hourly_rate", "").strip()

            if not username or not full_name:
                errors.append(f"Row {i}: username or full_name is empty")
                continue

            role = roles_map.get(role_name)
            if not role:
                errors.append(f"Row {i}: unknown role '{role_name}'")
                continue

            store_name_list = [s.strip() for s in store_names_raw.split(",") if s.strip()]
            resolved_stores: list[Store] = []
            for sn in store_name_list:
                store = stores_map.get(sn.lower())
                if not store:
                    errors.append(f"Row {i}: store not found '{sn}'")
                else:
                    resolved_stores.append(store)

            existing = await db.execute(
                select(User).where(
                    User.organization_id == organization_id,
                    User.username == username,
                    User.deleted_at.is_(None),
                )
            )
            existing_user = existing.scalar_one_or_none()

            if existing_user:
                skipped += 1
                user_id = existing_user.id
                users_list.append({"username": username, "full_name": existing_user.full_name, "status": "skipped"})
            else:
                if not password:
                    errors.append(f"Row {i}: password required for new user '{username}'")
                    continue
                try:
                    hourly_rate = float(hourly_rate_str) if hourly_rate_str else None
                    data = UserCreate(
                        username=username,
                        password=password,
                        full_name=full_name,
                        role_id=str(role.id),
                        email=email,
                    )
                    resp = await user_service.create_user(
                        db, organization_id, data, caller=caller
                    )
                    user_id = UUID(resp.id)

                    if hourly_rate is not None:
                        user_obj = await db.get(User, user_id)
                        if user_obj:
                            user_obj.hourly_rate = hourly_rate

                    users_list.append({"username": username, "full_name": full_name, "status": "created"})
                    created += 1
                except Exception as e:
                    errors.append(f"Row {i}: failed to create '{username}' — {e}")
                    continue

            for store in resolved_stores:
                exists = await db.execute(
                    select(UserStore).where(
                        UserStore.user_id == user_id,
                        UserStore.store_id == store.id,
                    )
                )
                if exists.scalar_one_or_none() is None:
                    db.add(UserStore(user_id=user_id, store_id=store.id, is_manager=False))
                    store_assigned += 1

        try:
            await db.commit()
        except Exception as e:
            await db.rollback()
            raise BadRequestError(f"DB commit failed: {e}")

        return {
            "created": created,
            "skipped": skipped,
            "store_assigned": store_assigned,
            "errors": errors,
            "total_rows": len(rows),
            "users": users_list,
        }

    # ── Schedule Excel ────────────────────────────────────

    async def process_schedules_excel(
        self,
        db: AsyncSession,
        organization_id: UUID,
        file_bytes: bytes,
        created_by: UUID,
    ) -> dict:
        """Excel 파일로 스케줄 대량 등록. 시트별 매장 구분.

        Each sheet format:
          Row 1: Week Start | {MM/DD/YYYY}
          Row 2: Employee | Sun | Mon | Tue | Wed | Thu | Fri | Sat
          Row 3+: {name} | {schedule per day}

        Sheet name = Store name
        Each cell: 9:00AM-6:00PM or 9:00AM-6:00PM(12:00PM-1:00PM) or empty/OFF
        """
        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as e:
            raise BadRequestError(f"Invalid Excel file: {e}")

        stores_map = await self._load_stores_map(db, organization_id)
        users_map = await self._load_users_map(db, organization_id)

        entries: list[ScheduleCreate] = []
        errors: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            store = stores_map.get(sheet_name.strip().lower())
            if not store:
                errors.append(f"Sheet '{sheet_name}': store not found")
                continue

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 3:
                errors.append(f"Sheet '{sheet_name}': need at least 3 rows (Week Start, Header, Data)")
                continue

            # Row 1: Week Start
            row1 = rows[0]
            week_start_str = str(row1[1]).strip() if len(row1) > 1 and row1[1] else ""

            # Handle date object from Excel
            if isinstance(row1[1], date):
                week_start = row1[1] if isinstance(row1[1], date) else row1[1].date()
            else:
                try:
                    parts = week_start_str.split("/")
                    week_start = date(int(parts[2]), int(parts[0]), int(parts[1]))
                except (ValueError, IndexError):
                    errors.append(f"Sheet '{sheet_name}': invalid date '{week_start_str}' (expected MM/DD/YYYY)")
                    continue

            # Row 2: Header (skip)
            # Row 3+: Data
            for row_idx in range(2, len(rows)):
                row = rows[row_idx]
                if not row or not row[0]:
                    continue

                employee_name = str(row[0]).strip()
                user = users_map.get(employee_name.lower())
                if not user:
                    errors.append(f"Sheet '{sheet_name}', row {row_idx + 1}: user ID not found '{employee_name}'")
                    continue

                for day_idx in range(7):
                    col_idx = day_idx + 1
                    if col_idx >= len(row) or not row[col_idx]:
                        continue
                    cell = str(row[col_idx]).strip()
                    if not cell or cell.upper() == "OFF":
                        continue

                    work_date = week_start + timedelta(days=day_idx)
                    parsed = self._parse_schedule_cell(cell)
                    if parsed is None:
                        day_name = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat'][day_idx]
                        errors.append(f"Sheet '{sheet_name}', row {row_idx + 1}, {day_name}: invalid '{cell}'")
                        continue

                    start_24, end_24, break_start_24, break_end_24 = parsed
                    entries.append(ScheduleCreate(
                        user_id=str(user.id),
                        store_id=str(store.id),
                        work_date=work_date,
                        start_time=start_24,
                        end_time=end_24,
                        break_start_time=break_start_24,
                        break_end_time=break_end_24,
                        status="confirmed",
                        force=True,
                    ))

        wb.close()

        if not entries and not errors:
            raise BadRequestError("No schedule entries found in Excel")

        result = await schedule_service.bulk_create(
            db, organization_id, entries, created_by, skip_on_conflict=True,
        )

        return {
            "created": result.created,
            "skipped": result.skipped,
            "failed": result.failed,
            "errors": errors + result.errors,
            "total_cells": len(entries),
        }

    # ── Helpers ───────────────────────────────────────────

    def _parse_csv(self, csv_bytes: bytes) -> list[dict]:
        text = csv_bytes.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return list(reader)

    def _parse_excel_rows(self, file_bytes: bytes) -> list[dict]:
        """Excel 첫 시트를 dict 리스트로 변환 (1행=헤더)."""
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows_raw = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows_raw) < 2:
            return []
        headers = [str(h).strip().lower() if h else "" for h in rows_raw[0]]
        result = []
        for row in rows_raw[1:]:
            d = {}
            for i, h in enumerate(headers):
                if h and i < len(row):
                    d[h] = str(row[i]).strip() if row[i] is not None else ""
                elif h:
                    d[h] = ""
            result.append(d)
        return result

    def _split_csv_line(self, line: str) -> list[str]:
        reader = csv.reader(io.StringIO(line))
        return next(reader, [])

    async def process_schedules_csv(
        self,
        db: AsyncSession,
        organization_id: UUID,
        csv_bytes: bytes,
        created_by: UUID,
    ) -> dict:
        """CSV 주간표 스케줄 등록. Store,{name} / Week Start,{date} / Employee,Sun,...,Sat"""
        raw_text = csv_bytes.decode("utf-8-sig")
        lines = raw_text.strip().splitlines()

        stores_map = await self._load_stores_map(db, organization_id)
        users_map = await self._load_users_map(db, organization_id)

        entries: list[ScheduleCreate] = []
        errors: list[str] = []

        # Split into blocks by "Store," rows
        blocks: list[tuple[int, list[str]]] = []
        current_block: list[str] = []
        current_start = 0

        for i, line in enumerate(lines):
            stripped = line.strip()
            if not stripped:
                continue
            cols = self._split_csv_line(stripped)
            if cols and cols[0].strip().lower() == "store" and current_block:
                blocks.append((current_start, current_block))
                current_block = []
                current_start = i
            current_block.append(stripped)
        if current_block:
            blocks.append((current_start, current_block))

        for block_start, block_lines in blocks:
            if len(block_lines) < 4:
                errors.append(f"Line {block_start + 1}: block too short")
                continue

            row1 = self._split_csv_line(block_lines[0])
            row2 = self._split_csv_line(block_lines[1])
            store_name = row1[1].strip() if len(row1) > 1 else ""
            week_start_str = row2[1].strip() if len(row2) > 1 else ""

            store = stores_map.get(store_name.lower())
            if not store:
                errors.append(f"Line {block_start + 1}: store not found '{store_name}'")
                continue

            try:
                parts = week_start_str.split("/")
                week_start = date(int(parts[2]), int(parts[0]), int(parts[1]))
            except (ValueError, IndexError):
                errors.append(f"Line {block_start + 2}: invalid date '{week_start_str}'")
                continue

            for rel_idx in range(3, len(block_lines)):
                abs_line = block_start + rel_idx + 1
                cols = self._split_csv_line(block_lines[rel_idx])
                if not cols or not cols[0].strip():
                    continue
                employee_name = cols[0].strip()
                user = users_map.get(employee_name.lower())
                if not user:
                    errors.append(f"Line {abs_line}: user ID not found '{employee_name}'")
                    continue
                for day_idx in range(7):
                    col_idx = day_idx + 1
                    if col_idx >= len(cols):
                        break
                    cell = cols[col_idx].strip()
                    if not cell or cell.upper() == "OFF":
                        continue
                    work_date = week_start + timedelta(days=day_idx)
                    parsed = self._parse_schedule_cell(cell)
                    if parsed is None:
                        day_name = ['Sun','Mon','Tue','Wed','Thu','Fri','Sat'][day_idx]
                        errors.append(f"Line {abs_line}, {day_name}: invalid '{cell}'")
                        continue
                    start_24, end_24, bs24, be24 = parsed
                    entries.append(ScheduleCreate(
                        user_id=str(user.id), store_id=str(store.id),
                        work_date=work_date, start_time=start_24, end_time=end_24,
                        break_start_time=bs24, break_end_time=be24,
                        status="confirmed", force=True,
                    ))

        if not entries and not errors:
            raise BadRequestError("No schedule entries found in CSV")

        result = await schedule_service.bulk_create(
            db, organization_id, entries, created_by, skip_on_conflict=True,
        )
        return {
            "created": result.created, "skipped": result.skipped,
            "failed": result.failed, "errors": errors + result.errors,
            "total_cells": len(entries),
        }

    async def _load_roles_map(self, db: AsyncSession, org_id: UUID) -> dict[str, Role]:
        result = await db.execute(
            select(Role).where(Role.organization_id == org_id)
        )
        return {r.name.lower(): r for r in result.scalars().all()}

    async def _load_stores_map(self, db: AsyncSession, org_id: UUID) -> dict[str, Store]:
        result = await db.execute(
            select(Store).where(
                Store.organization_id == org_id,
                Store.is_active.is_(True),
            )
        )
        return {s.name.lower(): s for s in result.scalars().all()}

    async def _load_users_map(self, db: AsyncSession, org_id: UUID) -> dict[str, User]:
        """username (lowercase) -> User."""
        result = await db.execute(
            select(User).where(
                User.organization_id == org_id,
                User.deleted_at.is_(None),
                User.is_active.is_(True),
            )
        )
        return {u.username.lower(): u for u in result.scalars().all()}

    def _parse_schedule_cell(self, cell: str) -> tuple[str, str, str | None, str | None] | None:
        """Parse '09:00AM-06:00PM' or '09:00AM-06:00PM(12:00PM-01:00PM)'.
        Returns (start_24h, end_24h, break_start_24h, break_end_24h) in HH:MM.
        """
        cell = cell.strip()

        break_start = None
        break_end = None
        break_match = re.search(r'\((.+?)\)', cell)
        if break_match:
            break_part = break_match.group(1)
            cell = cell[:break_match.start()].strip()
            break_times = break_part.split("-")
            if len(break_times) == 2:
                break_start = self._ampm_to_24(break_times[0].strip())
                break_end = self._ampm_to_24(break_times[1].strip())

        parts = cell.split("-")
        if len(parts) != 2:
            return None

        start = self._ampm_to_24(parts[0].strip())
        end = self._ampm_to_24(parts[1].strip())

        if start is None or end is None:
            return None

        return (start, end, break_start, break_end)

    def _ampm_to_24(self, time_str: str) -> str | None:
        """Convert '09:00AM' → '09:00', '01:30PM' → '13:30'. Also accepts 24h 'HH:MM'."""
        if not time_str:
            return None
        time_str = time_str.strip().upper()

        if re.match(r'^\d{1,2}:\d{2}$', time_str):
            return time_str

        match = re.match(r'^(\d{1,2}):(\d{2})\s*(AM|PM)$', time_str)
        if not match:
            return None

        hour = int(match.group(1))
        minute = int(match.group(2))
        period = match.group(3)

        if period == "AM":
            if hour == 12:
                hour = 0
        else:
            if hour != 12:
                hour += 12

        return f"{hour:02d}:{minute:02d}"


bulk_upload_service = BulkUploadService()
