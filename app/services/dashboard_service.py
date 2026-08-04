"""대시보드 서비스 — 대시보드 집계 비즈니스 로직.

Dashboard Service — Aggregation logic for admin dashboard.
Provides checklist completion rates, attendance summary, and overtime summary.
"""

from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Mapping, Sequence
from uuid import UUID
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, select, case, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.attendance import Attendance
from app.models.attendance_break import AttendanceBreak
from app.models.checklist import ChecklistInstance
from app.models.schedule import Schedule
from app.models.evaluation import Evaluation
from app.models.organization import Organization, Store
from app.models.user import User
from app.services.attendance_service import compute_net_work_minutes
from app.services.labor_law_service import (
    DEFAULT_MAX_WEEKLY_HOURS,
    resolve_weekly_max_hours,
)
from app.utils.timezone import DEFAULT_TIMEZONE


def week_start_of(d: date) -> date:
    """해당 날짜가 속한 주의 시작일(일요일) — 주는 항상 Sun→Sat."""
    return d - timedelta(days=(d.weekday() + 1) % 7)


def build_weekly_overtime_rows(
    attendances: Sequence[Attendance],
    breaks_map: Mapping[UUID, Sequence[AttendanceBreak]],
    max_weekly_by_store: Mapping[UUID, int],
) -> list[dict]:
    """user × 주(일요일 시작) 단위로 net 근무시간을 집계하고 주간 한도와 비교한다.

    P0-4: 기존 버그 — 조회 범위 전체 합계를 한 주 한도와 비교해 초과근무가
    부풀려졌음. 주 단위로 쪼개 각 주의 net 합을 그 주 한도와 비교한다.
    - net 은 attendance 별 C1 공식(compute_net_work_minutes) 합산 (gross 금지)
    - 멀티 매장 주간은 min() 기준 적용 — overtime alerts 와 동일한 보수적 결정
    - store 유실(SET NULL) 또는 map 에 없는 매장은 기본 40h

    Returns (week_start, user_id 순 정렬):
        {user_id, week_start, week_end, total_hours, max_weekly, overtime_hours}
    """
    agg: dict[tuple[UUID, date], dict] = {}
    for att in attendances:
        key = (att.user_id, week_start_of(att.work_date))
        entry = agg.setdefault(key, {"net": 0, "store_ids": set()})
        entry["net"] += compute_net_work_minutes(att, breaks_map.get(att.id, [])) or 0
        if att.store_id is not None:
            entry["store_ids"].add(att.store_id)

    rows: list[dict] = []
    for (uid, wk_start), entry in sorted(
        agg.items(), key=lambda kv: (kv[0][1], str(kv[0][0]))
    ):
        if entry["store_ids"]:
            max_weekly = min(
                max_weekly_by_store.get(sid, DEFAULT_MAX_WEEKLY_HOURS)
                for sid in entry["store_ids"]
            )
        else:
            max_weekly = DEFAULT_MAX_WEEKLY_HOURS
        total_hours = entry["net"] / 60
        rows.append({
            "user_id": uid,
            "week_start": wk_start,
            "week_end": wk_start + timedelta(days=6),
            "total_hours": round(total_hours, 1),
            "max_weekly": max_weekly,
            "overtime_hours": round(max(0.0, total_hours - max_weekly), 1),
        })
    return rows


async def _load_breaks_map(
    db: AsyncSession,
    attendance_ids: list[UUID],
) -> dict[UUID, list[AttendanceBreak]]:
    """attendance 별 break 세션 일괄 로드 — C1 net 계산용."""
    if not attendance_ids:
        return {}
    result = await db.execute(
        select(AttendanceBreak)
        .where(AttendanceBreak.attendance_id.in_(attendance_ids))
        .order_by(AttendanceBreak.started_at.asc())
    )
    out: dict[UUID, list[AttendanceBreak]] = {}
    for br in result.scalars().all():
        out.setdefault(br.attendance_id, []).append(br)
    return out


async def _resolve_max_weekly_map(
    db: AsyncSession,
    attendances: Sequence[Attendance],
) -> dict[UUID, int]:
    """attendance 에 등장하는 매장별 주간 한도 resolve — 매장당 1회 조회."""
    involved = {att.store_id for att in attendances if att.store_id is not None}
    return {sid: await resolve_weekly_max_hours(db, sid) for sid in involved}


class DashboardService:
    """대시보드 서비스.

    Dashboard aggregation service for admin dashboard views.
    """

    async def _resolve_today(
        self,
        db: AsyncSession,
        organization_id: UUID,
        store_id: UUID | None = None,
    ) -> date:
        """매장/조직 타임존 기준 오늘 날짜를 반환합니다."""
        if store_id:
            result = await db.execute(
                select(Store.timezone, Organization.timezone.label("org_timezone"))
                .join(Organization, Store.organization_id == Organization.id)
                .where(Store.id == store_id)
            )
            row = result.one_or_none()
            if row:
                tz_str = row.timezone or row.org_timezone or DEFAULT_TIMEZONE
                return datetime.now(ZoneInfo(tz_str)).date()
        # 조직 타임존 사용
        result = await db.execute(
            select(Organization.timezone).where(Organization.id == organization_id)
        )
        tz_str = result.scalar_one_or_none() or DEFAULT_TIMEZONE
        return datetime.now(ZoneInfo(tz_str)).date()

    async def get_checklist_completion(
        self,
        db: AsyncSession,
        organization_id: UUID,
        date_from: date | None = None,
        date_to: date | None = None,
        store_id: UUID | None = None,
    ) -> dict:
        """체크리스트 완료율 집계 (schedule + cl_instances 기반)."""
        today = await self._resolve_today(db, organization_id, store_id)
        if date_from is None:
            date_from = today - timedelta(days=7)
        if date_to is None:
            date_to = today

        base = (
            select(
                func.count(ChecklistInstance.id).label("total_assignments"),
                func.sum(
                    case(
                        (ChecklistInstance.status == "completed", 1),
                        else_=0,
                    )
                ).label("completed_assignments"),
            )
            .where(
                ChecklistInstance.organization_id == organization_id,
                ChecklistInstance.work_date >= date_from,
                ChecklistInstance.work_date <= date_to,
            )
        )
        if store_id:
            base = base.where(ChecklistInstance.store_id == store_id)

        result = await db.execute(base)
        row = result.one()
        total = row.total_assignments or 0
        completed = row.completed_assignments or 0
        rate = round((completed / total * 100), 1) if total > 0 else 0

        return {
            "date_from": str(date_from),
            "date_to": str(date_to),
            "total_assignments": total,
            "completed_assignments": completed,
            "completion_rate": rate,
        }

    async def get_attendance_summary(
        self,
        db: AsyncSession,
        organization_id: UUID,
        date_from: date | None = None,
        date_to: date | None = None,
        store_id: UUID | None = None,
    ) -> dict:
        """근태 요약 집계."""
        today = await self._resolve_today(db, organization_id, store_id)
        if date_from is None:
            date_from = today - timedelta(days=7)
        if date_to is None:
            date_to = today

        base = (
            select(
                func.count(Attendance.id).label("total"),
                func.sum(
                    case((Attendance.status == "completed", 1), else_=0)
                ).label("completed"),
                func.sum(
                    case((Attendance.status == "clocked_in", 1), else_=0)
                ).label("clocked_in"),
                func.avg(Attendance.total_work_minutes).label("avg_work_minutes"),
            )
            .where(
                Attendance.organization_id == organization_id,
                Attendance.work_date >= date_from,
                Attendance.work_date <= date_to,
            )
        )
        if store_id:
            base = base.where(Attendance.store_id == store_id)

        result = await db.execute(base)
        row = result.one()

        return {
            "date_from": str(date_from),
            "date_to": str(date_to),
            "total_records": row.total or 0,
            "completed": row.completed or 0,
            "clocked_in": row.clocked_in or 0,
            "avg_work_minutes": round(float(row.avg_work_minutes or 0), 1),
        }

    async def get_overtime_summary(
        self,
        db: AsyncSession,
        organization_id: UUID,
        week_date: date | None = None,
        store_id: UUID | None = None,
    ) -> dict:
        """초과근무 현황 요약.

        - 주간 시간은 attendance 별 C1 net(compute_net_work_minutes) 합산
        - 기준(max weekly)은 매장별 LaborLawSetting cascade — resolve_weekly_max_hours
          (기존 버그: org 내 임의 row .limit(1) + gross 분 합산)
        """
        today = await self._resolve_today(db, organization_id, store_id)
        target_date = week_date or today
        week_start = week_start_of(target_date)
        week_end = week_start + timedelta(days=6)

        query = (
            select(Attendance)
            .where(
                Attendance.organization_id == organization_id,
                Attendance.work_date >= week_start,
                Attendance.work_date <= week_end,
            )
        )
        if store_id:
            query = query.where(Attendance.store_id == store_id)

        result = await db.execute(query)
        attendances = list(result.scalars().all())

        breaks_map = await _load_breaks_map(db, [a.id for a in attendances])
        max_weekly_by_store = await _resolve_max_weekly_map(db, attendances)
        rows = build_weekly_overtime_rows(attendances, breaks_map, max_weekly_by_store)

        total_users = len(rows)
        overtime_users = sum(1 for r in rows if r["overtime_hours"] > 0)
        total_overtime_hours = round(sum(r["overtime_hours"] for r in rows), 1)

        # 표시용 한도: store 필터 시 그 매장 기준, 아니면 관련 매장 중 가장 엄격한 값.
        # (판정 자체는 user 별 근무 매장 기준 — build_weekly_overtime_rows)
        if store_id:
            max_weekly = await resolve_weekly_max_hours(db, store_id)
        elif max_weekly_by_store:
            max_weekly = min(max_weekly_by_store.values())
        else:
            max_weekly = DEFAULT_MAX_WEEKLY_HOURS

        return {
            "week_start": str(week_start),
            "week_end": str(week_end),
            "max_weekly_hours": max_weekly,
            "total_users_with_attendance": total_users,
            "overtime_users": overtime_users,
            "total_overtime_hours": total_overtime_hours,
        }

    async def get_evaluation_summary(
        self,
        db: AsyncSession,
        organization_id: UUID,
    ) -> dict:
        """평가 요약."""
        result = await db.execute(
            select(
                func.count(Evaluation.id).label("total"),
                func.sum(case((Evaluation.status == "draft", 1), else_=0)).label("draft"),
                func.sum(case((Evaluation.status == "submitted", 1), else_=0)).label("submitted"),
            )
            .where(
                Evaluation.organization_id == organization_id,
                # soft-deleted 평가는 집계에서 제외 (v1 redesign)
                Evaluation.deleted_at.is_(None),
            )
        )
        row = result.one()
        return {
            "total_evaluations": row.total or 0,
            "draft": row.draft or 0,
            "submitted": row.submitted or 0,
        }


    async def export_excel(
        self,
        db: AsyncSession,
        organization_id: UUID,
        date_from: date | None = None,
        date_to: date | None = None,
        store_id: UUID | None = None,
    ) -> bytes:
        """대시보드 데이터를 Excel 파일로 내보내기."""
        today = await self._resolve_today(db, organization_id, store_id)
        if date_from is None:
            date_from = today - timedelta(days=7)
        if date_to is None:
            date_to = today

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2D3436", end_color="2D3436", fill_type="solid")

        def style_headers(ws, headers: list[str]) -> None:
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        # --- Sheet 1: Checklist Completion ---
        ws1 = wb.active
        ws1.title = "Checklist Completion"
        headers1 = ["Store", "User", "Work Date", "Status", "Total Items", "Completed Items"]
        style_headers(ws1, headers1)

        checklist_query = (
            select(
                Store.name.label("store_name"),
                User.full_name.label("user_name"),
                ChecklistInstance.work_date,
                ChecklistInstance.status,
                ChecklistInstance.total_items,
                ChecklistInstance.completed_items,
            )
            .join(Store, ChecklistInstance.store_id == Store.id)
            .join(User, ChecklistInstance.user_id == User.id)
            .where(
                ChecklistInstance.organization_id == organization_id,
                ChecklistInstance.work_date >= date_from,
                ChecklistInstance.work_date <= date_to,
            )
            .order_by(ChecklistInstance.work_date.desc())
        )
        if store_id:
            checklist_query = checklist_query.where(ChecklistInstance.store_id == store_id)

        result = await db.execute(checklist_query)
        for row in result.all():
            ws1.append([row.store_name, row.user_name, str(row.work_date), row.status, row.total_items, row.completed_items])

        for i, w in enumerate([20, 20, 15, 15, 12, 15], 1):
            ws1.column_dimensions[ws1.cell(row=1, column=i).column_letter].width = w

        # --- Sheet 2: Attendance ---
        ws2 = wb.create_sheet("Attendance")
        headers2 = ["Store", "User", "Work Date", "Clock In", "Clock Out", "Break (min)", "Work (min)", "Status"]
        style_headers(ws2, headers2)

        att_query = (
            select(
                Store.name.label("store_name"),
                User.full_name.label("user_name"),
                Attendance.work_date,
                Attendance.clock_in,
                Attendance.clock_out,
                Attendance.total_break_minutes,
                Attendance.total_work_minutes,
                Attendance.status,
            )
            .join(Store, Attendance.store_id == Store.id)
            .join(User, Attendance.user_id == User.id)
            .where(
                Attendance.organization_id == organization_id,
                Attendance.work_date >= date_from,
                Attendance.work_date <= date_to,
            )
            .order_by(Attendance.work_date.desc())
        )
        if store_id:
            att_query = att_query.where(Attendance.store_id == store_id)

        result = await db.execute(att_query)
        for row in result.all():
            ws2.append([
                row.store_name,
                row.user_name,
                str(row.work_date),
                row.clock_in.isoformat() if row.clock_in else "",
                row.clock_out.isoformat() if row.clock_out else "",
                row.total_break_minutes or 0,
                row.total_work_minutes or 0,
                row.status,
            ])

        for i, w in enumerate([20, 20, 15, 22, 22, 12, 12, 15], 1):
            ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w

        # --- Sheet 3: Overtime ---
        # P0-4: 조회 범위를 일요일 시작 주(Sun–Sat)로 쪼개 user × 주 단위로
        # net 근무시간을 각 주의 매장별 한도와 비교한다.
        # (기존 버그: 범위 전체 gross 합 vs 한 주 한도 → 초과근무 부풀림)
        ws3 = wb.create_sheet("Overtime")
        headers3 = ["User", "Week Start", "Week End", "Total Hours", "Max Weekly", "Overtime Hours"]
        style_headers(ws3, headers3)

        # 경계 주는 통째로 포함 — date_from 이 속한 주 일요일부터
        # date_to 가 속한 주 토요일까지 (주간 합계가 잘리지 않도록)
        range_start = week_start_of(date_from)
        range_end = week_start_of(date_to) + timedelta(days=6)

        ot_query = (
            select(Attendance)
            .where(
                Attendance.organization_id == organization_id,
                Attendance.work_date >= range_start,
                Attendance.work_date <= range_end,
            )
        )
        if store_id:
            ot_query = ot_query.where(Attendance.store_id == store_id)

        result = await db.execute(ot_query)
        attendances = list(result.scalars().all())

        breaks_map = await _load_breaks_map(db, [a.id for a in attendances])
        max_weekly_by_store = await _resolve_max_weekly_map(db, attendances)
        ot_rows = build_weekly_overtime_rows(attendances, breaks_map, max_weekly_by_store)

        # 사용자 이름 일괄 조회 — row 당 개별 쿼리 제거
        user_ids = list({r["user_id"] for r in ot_rows})
        names_map: dict[UUID, str] = {}
        if user_ids:
            names_result = await db.execute(
                select(User.id, User.full_name).where(User.id.in_(user_ids))
            )
            names_map = {row.id: row.full_name for row in names_result}

        for r in ot_rows:
            ws3.append([
                names_map.get(r["user_id"]) or "Unknown",
                str(r["week_start"]),
                str(r["week_end"]),
                r["total_hours"],
                r["max_weekly"],
                r["overtime_hours"],
            ])

        for i, w in enumerate([20, 15, 15, 12, 12, 15], 1):
            ws3.column_dimensions[ws3.cell(row=1, column=i).column_letter].width = w

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


# 싱글턴 인스턴스
dashboard_service: DashboardService = DashboardService()
