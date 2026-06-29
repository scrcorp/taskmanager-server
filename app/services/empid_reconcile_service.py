"""EMPID 레거시 임포트 — 매칭/버킷 분류 + 확정(commit) 서비스.

Backoffice의 첫 도구. 레거시 직원 마스터(xlsx)의 사번(emp_id)을
우리 `users.employee_no`(org-scope)에 채운다.

SoT: docs/99_inbox/2026-06-24 HTM control-plane(=Backoffice) 운영자콘솔 + EMPID 임포트 설계.md
설계 핵심:
- blind import 금지 → 제안(버킷) → 사람 검토 → 확정(commit).
- 멱등: 이미 employee_no 있으면 skip(=저널). NULL만 채움.
- PURADAK(폐점) 행 제외. 신규 번호 생성은 이 도구 범위 밖.

버킷(설계 §5 + 확장):
1. auto       — 이메일 1:1 매칭 + 단일 emp_id + 대상 user employee_no NULL → 자동 제안
2. multiple   — 멀티스토어 동일인(이메일 같고 emp_id 여러 개) → 운영자가 canonical 선택
                (어떤 번호가 어느 COMPANY에서 왔는지 함께 표시)
3. placeholder— 더미/공용 이메일(내부 도메인 or 서로 다른 사람들이 공유) → 매칭 제외
4. assigned   — 매칭됐고 이미 employee_no 있음 + 파일 번호와 일치 → skip 표시
5. mismatch   — 이미 employee_no 있는데 파일의 번호가 다름 → 충돌, 운영자 확인 필요
6. deferred   — DB 미매칭(이메일 있으나 user 없음) / 무이메일 → 리포트만
                (이름이 비슷한 DB 유저 후보를 힌트로 함께 표시)
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from uuid import UUID

import openpyxl
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.utils.exceptions import DuplicateError
from app.repositories.employee_no_history_repository import (
    employee_no_history_repository,
)
from app.repositories.user_repository import user_repository
from app.schemas.user import _normalize_employee_no

# 더미/내부 이메일 도메인 — 매칭 키로 신뢰 불가 (placeholder 처리)
_PLACEHOLDER_DOMAINS = {"tigersplus.com"}
# 폐점 등 임포트 제외 회사명 (COMPANY 컬럼 부분일치, 대문자 비교)
_EXCLUDED_COMPANIES = ("PURADAK",)


def _norm_email(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip().lower()
    return s or None


def _name_tokens(name: object) -> set[str]:
    """이름 → 알파벳 토큰 집합(소문자, 별칭 괄호 제거). 이름 유사도 비교용."""
    if not name:
        return set()
    s = re.sub(r"\(.*?\)", "", str(name)).strip().lower()
    return {t for t in re.sub(r"[^a-z\s]", "", s).split() if len(t) > 1}


def _first_name_token(name: object) -> str:
    """이름에서 별칭 괄호 제거 후 첫 단어(소문자, 알파벳만) — 동일인 판별용."""
    if not name:
        return ""
    s = re.sub(r"\(.*?\)", "", str(name)).strip().lower()
    parts = re.sub(r"[^a-z\s]", "", s).split()
    return parts[0] if parts else ""


def _name_similar(a: set[str], b: set[str]) -> bool:
    """두 이름 토큰셋이 '비슷한가' — 운영자 수동확인 힌트용(느슨하게).

    기준: 공통 토큰 있고 (Jaccard ≥ 0.5  OR  한쪽이 다른 쪽의 부분집합).
    예: {john,doe} ~ {john,doe,jr}, {maria,santos} ~ {maria,l,santos}.
    """
    if not a or not b:
        return False
    inter = a & b
    if not inter:
        return False
    union = a | b
    return (len(inter) / len(union) >= 0.5) or a <= b or b <= a


@dataclass
class EmpRow:
    """xlsx 한 행 (정규화)."""

    company: str
    corp_abr: str | None
    name: str
    emp_id: str  # 문자열 보존(선행0)
    email: str | None


@dataclass
class Proposal:
    """매칭 제안 1건."""

    email: str | None
    name: str
    user_id: UUID | None  # 매칭된 우리 유저
    user_full_name: str | None
    emp_id: str | None  # 단일 후보(auto)
    emp_id_options: list[str] = field(default_factory=list)  # 멀티(operator 선택)
    note: str = ""
    # 멀티/충돌 — 각 emp_id가 어느 COMPANY에서 왔는지 (emp_id, "회사1, 회사2")
    emp_id_sources: list[tuple[str, str]] = field(default_factory=list)
    # mismatch — DB에 이미 들어있는 사번 (파일과 다를 때)
    db_emp_id: str | None = None
    # deferred — 이름이 비슷한 DB 유저 후보 [(full_name, email)]
    similar: list[tuple[str, str | None]] = field(default_factory=list)
    # placeholder — 같은 이메일을 쓰는 각 인물+번호 [(name, emp_id, company)]
    members: list[tuple[str, str, str]] = field(default_factory=list)
    # placeholder/공유 — 이 이메일을 실제로 쓰는 DB 계정 [(full_name, employee_no)]
    db_accounts: list[tuple[str, str | None]] = field(default_factory=list)


@dataclass
class ReconcileResult:
    """버킷별 분류 결과."""

    auto: list[Proposal] = field(default_factory=list)
    multiple: list[Proposal] = field(default_factory=list)
    placeholder: list[Proposal] = field(default_factory=list)
    assigned: list[Proposal] = field(default_factory=list)
    mismatch: list[Proposal] = field(default_factory=list)
    deferred: list[Proposal] = field(default_factory=list)
    excluded_rows: int = 0  # PURADAK 등 제외 행 수
    total_rows: int = 0

    def counts(self) -> dict[str, int]:
        return {
            "auto": len(self.auto),
            "multiple": len(self.multiple),
            "placeholder": len(self.placeholder),
            "assigned": len(self.assigned),
            "mismatch": len(self.mismatch),
            "deferred": len(self.deferred),
            "excluded_rows": self.excluded_rows,
            "total_rows": self.total_rows,
        }


# 헤더 별칭 — 컬럼명이 조금 달라도 매핑 (소문자/공백제거 기준)
def _hkey(s: object) -> str:
    return re.sub(r"\s+", "", str(s or "").strip().lower())


_HEADER_ALIASES = {
    "company": "company", "corp_abr_3": "corp_abr", "corpabr3": "corp_abr",
    "name": "name", "emp_id": "emp_id", "empid": "emp_id",
    "email": "email",
}


def _emp_id_str(raw: object) -> str | None:
    """emp_id 문자열 보존 (float면 .0 제거, 선행0 보존)."""
    if raw is None:
        return None
    if isinstance(raw, float) and raw.is_integer():
        return str(int(raw))
    s = str(raw).strip()
    return s or None


def _build_row(cells: dict[str, object]) -> EmpRow | None:
    """정규화 컬럼 dict(company/corp_abr/name/emp_id/email) → EmpRow (emp_id 없으면 None)."""
    emp_id = _emp_id_str(cells.get("emp_id"))
    if emp_id is None:
        return None
    return EmpRow(
        company=str(cells.get("company") or "").strip(),
        corp_abr=(str(cells.get("corp_abr")).strip() if cells.get("corp_abr") else None),
        name=str(cells.get("name") or "").strip(),
        emp_id=emp_id,
        email=_norm_email(cells.get("email")),
    )


def _is_excluded(company: str) -> bool:
    return any(x in company.upper() for x in _EXCLUDED_COMPANIES)


def _rows_from_records(records) -> tuple[list[EmpRow], int]:
    """헤더 매핑된 레코드 iterable(dict 정규화키) → (EmpRow 목록, 제외 수)."""
    out: list[EmpRow] = []
    excluded = 0
    for cells in records:
        company = str(cells.get("company") or "").strip()
        if _is_excluded(company):
            excluded += 1
            continue
        row = _build_row(cells)
        if row is not None:
            out.append(row)
    return out, excluded


def _parse_xlsx(content: bytes) -> tuple[list[EmpRow], int]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    header = [_HEADER_ALIASES.get(_hkey(c)) for c in next(rows_iter)]

    def records():
        for row in rows_iter:
            if row is None or all(c is None for c in row):
                continue
            yield {key: row[i] for i, key in enumerate(header) if key and i < len(row)}

    return _rows_from_records(records())


def _parse_csv(content: bytes) -> tuple[list[EmpRow], int]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        header = [_HEADER_ALIASES.get(_hkey(c)) for c in next(reader)]
    except StopIteration:
        return [], 0

    def records():
        for row in reader:
            if not row or all(not str(c).strip() for c in row):
                continue
            yield {key: row[i] for i, key in enumerate(header) if key and i < len(row)}

    return _rows_from_records(records())


def parse_emplist(content: bytes, filename: str = "") -> tuple[list[EmpRow], int]:
    """업로드 파일(xlsx/csv) → (제외 후 EmpRow 목록, 제외된 행 수).

    포맷은 확장자로 판별(.csv → CSV, 그 외 → Excel). PURADAK 등 폐점 회사 행 제외.
    """
    if filename.lower().endswith(".csv"):
        return _parse_csv(content)
    return _parse_xlsx(content)


def _is_placeholder_email(email: str) -> bool:
    domain = email.rsplit("@", 1)[-1] if "@" in email else ""
    return domain in _PLACEHOLDER_DOMAINS


def _sources_for(rows: list[EmpRow], emp_ids: list[str]) -> list[tuple[str, str]]:
    """각 emp_id가 어느 COMPANY 행에서 왔는지 (emp_id, "회사1, 회사2")."""
    by_id: dict[str, set[str]] = {}
    for r in rows:
        by_id.setdefault(r.emp_id, set()).add(r.company or "—")
    return [(eid, ", ".join(sorted(by_id.get(eid, set())))) for eid in emp_ids]


def classify(
    emp_rows: list[EmpRow],
    users_by_email: dict[str, list],
    users: list | None = None,
) -> ReconcileResult:
    """순수 분류 로직 — (xlsx 행, 이메일→유저목록) → 버킷.

    users_by_email: normalized email → [User, ...] (해당 org).
    users: 해당 org 전체 유저(이름 유사도 힌트용, 선택). 없으면 이름 매칭 생략.
    User는 .id, .full_name, .email, .employee_no 속성만 사용.
    """
    result = ReconcileResult()
    result.total_rows = len(emp_rows)

    # 이름 유사도 인덱스 (deferred 힌트용)
    name_index: list[tuple[set[str], object]] = []
    if users:
        for u in users:
            toks = _name_tokens(getattr(u, "full_name", None))
            if toks:
                name_index.append((toks, u))

    def _similar(name: str) -> list[tuple[str, str | None]]:
        toks = _name_tokens(name)
        if not toks:
            return []
        out: list[tuple[str, str | None]] = []
        for u_toks, u in name_index:
            if _name_similar(toks, u_toks):
                out.append((getattr(u, "full_name", ""), getattr(u, "email", None)))
        return out[:5]

    # 이메일 없는 행 → deferred(no-email) — 이름 유사 후보 힌트 부착
    for r in (row for row in emp_rows if not row.email):
        result.deferred.append(Proposal(
            email=None, name=r.name, user_id=None, user_full_name=None,
            emp_id=r.emp_id, note="no email", similar=_similar(r.name),
        ))

    # 이메일별 그룹
    groups: dict[str, list[EmpRow]] = {}
    for r in emp_rows:
        if r.email:
            groups.setdefault(r.email, []).append(r)

    for email, rows in groups.items():
        emp_ids = sorted({r.emp_id for r in rows})
        sources = _sources_for(rows, emp_ids)
        first_names = {_first_name_token(r.name) for r in rows}
        same_person = len(first_names) == 1
        rep_name = rows[0].name
        db_users = users_by_email.get(email, [])

        # 3. placeholder — 내부 도메인 or 서로 다른 사람이 공유
        if _is_placeholder_email(email) or not same_person:
            # 각 인물+번호를 그대로 보여준다(이름 하나로 뭉뚱그리지 않음). 중복 행 제거.
            members = list(dict.fromkeys((r.name, r.emp_id, r.company) for r in rows))
            # 이 이메일을 실제로 쓰는 DB 계정도 함께 표시
            db_accounts = [(u.full_name, getattr(u, "employee_no", None)) for u in db_users]
            reason = ("internal email — shared placeholder" if _is_placeholder_email(email)
                      else "shared email — multiple people")
            note = reason + (f"; {len(db_accounts)} DB account(s) use this email" if db_accounts else "")
            result.placeholder.append(
                Proposal(email=email, name=rep_name, user_id=None, user_full_name=None,
                         emp_id=None, emp_id_options=emp_ids, emp_id_sources=sources,
                         members=members, db_accounts=db_accounts, note=note)
            )
            continue

        # 동일인. DB 매칭?
        if not db_users:
            result.deferred.append(
                Proposal(email=email, name=rep_name, user_id=None, user_full_name=None,
                         emp_id=emp_ids[0] if len(emp_ids) == 1 else None, emp_id_options=emp_ids,
                         emp_id_sources=sources, note="email present, no DB user",
                         similar=_similar(rep_name))
            )
            continue

        user = db_users[0]  # 동일 이메일 1유저가 일반적
        db_emp = getattr(user, "employee_no", None)

        # 4/5. 이미 사번 있음 → 파일 번호와 일치하면 skip, 다르면 mismatch(충돌)
        if db_emp:
            if db_emp in emp_ids:
                result.assigned.append(
                    Proposal(email=email, name=rep_name, user_id=user.id, user_full_name=user.full_name,
                             emp_id=db_emp, note="already assigned (matches file)")
                )
            else:
                result.mismatch.append(
                    Proposal(email=email, name=rep_name, user_id=user.id, user_full_name=user.full_name,
                             emp_id=db_emp, db_emp_id=db_emp, emp_id_options=emp_ids,
                             emp_id_sources=sources,
                             note=f"DB has {db_emp}, file has {', '.join(emp_ids)}")
                )
            continue

        # 1/2. NULL → 자동 or 멀티
        if len(emp_ids) == 1:
            result.auto.append(
                Proposal(email=email, name=rep_name, user_id=user.id, user_full_name=user.full_name, emp_id=emp_ids[0])
            )
        else:
            result.multiple.append(
                Proposal(email=email, name=rep_name, user_id=user.id, user_full_name=user.full_name,
                         emp_id=None, emp_id_options=emp_ids, emp_id_sources=sources,
                         note="multi-store same person — pick canonical")
            )

    return result


async def reconcile(
    db: AsyncSession, organization_id: UUID, content: bytes, filename: str = ""
) -> ReconcileResult:
    """DB 연동 — 업로드 파일(xlsx/csv) 파싱 + 해당 org 유저 매칭 + 분류."""
    emp_rows, excluded = parse_emplist(content, filename)
    users = await user_repository.get_by_org(db, organization_id)
    by_email: dict[str, list] = {}
    for u in users:
        e = _norm_email(u.email)
        if e:
            by_email.setdefault(e, []).append(u)
    result = classify(emp_rows, by_email, users=users)
    result.excluded_rows = excluded
    return result


def build_report_csv(result: ReconcileResult) -> str:
    """버킷 결과 전체를 공유용 CSV 문자열로 직렬화.

    컬럼: bucket, name, email, db_emp_id, file_emp_ids, sources, similar_db_users, note
    """
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["bucket", "name", "email", "db_emp_id", "file_emp_ids",
                "sources", "similar_db_users", "note"])

    def _src(p: Proposal) -> str:
        if p.members:  # placeholder/공유 — 인물별 번호를 그대로
            return " | ".join(f"{name}={eid}({co})" for name, eid, co in p.members)
        if p.emp_id_sources:
            return " | ".join(f"{eid}={co}" for eid, co in p.emp_id_sources)
        return ""

    def _sim(p: Proposal) -> str:
        if p.db_accounts:  # placeholder/공유 — 실제 DB 계정
            return " | ".join(f"{fn}={emp or '-'}" for fn, emp in p.db_accounts)
        return " | ".join(f"{n} <{e or '-'}>" for n, e in p.similar)

    def _file_ids(p: Proposal) -> str:
        if p.emp_id_options:
            return ", ".join(p.emp_id_options)
        return p.emp_id or ""

    for bucket, props in (
        ("auto", result.auto),
        ("multiple", result.multiple),
        ("mismatch", result.mismatch),
        ("assigned", result.assigned),
        ("placeholder", result.placeholder),
        ("deferred", result.deferred),
    ):
        for p in props:
            w.writerow([
                bucket,
                p.user_full_name or p.name,
                p.email or "",
                p.db_emp_id or "",
                _file_ids(p),
                _src(p),
                _sim(p),
                p.note,
            ])
    return buf.getvalue()


def _p_sources(p: Proposal) -> str:
    return "\n".join(f"{eid} ← {co}" for eid, co in p.emp_id_sources)


def _p_file_ids(p: Proposal) -> str:
    return ", ".join(p.emp_id_options) if p.emp_id_options else (p.emp_id or "")


def _p_similar(p: Proposal) -> str:
    return "\n".join(f"{n} <{e or '-'}>" for n, e in p.similar)


def _p_members(p: Proposal) -> str:
    return "\n".join(f"{name} = {eid} ({co})" for name, eid, co in p.members)


def _p_db_accounts(p: Proposal) -> str:
    return "\n".join(f"{fn} = {emp or '(no emp_id)'}" for fn, emp in p.db_accounts)


def build_report_xlsx(result: ReconcileResult, org_name: str = "", filename: str = "") -> bytes:
    """버킷 결과를 보기 편한 .xlsx로 — 버킷별 시트 + Summary 시트, 헤더 스타일/틀고정/열폭."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    head_font = Font(bold=True, color="FFFFFF", size=11)
    wrap = Alignment(vertical="top", wrap_text=True)
    top = Alignment(vertical="top")

    def _sheet(title: str, color: str, headers: list[str], rows: list[list[str]]) -> None:
        ws = wb.create_sheet(title)
        fill = PatternFill("solid", fgColor=color)
        ws.append(headers)
        for c in ws[1]:
            c.font = head_font
            c.fill = fill
            c.alignment = Alignment(vertical="center")
        for r in rows:
            ws.append(r)
        # 열폭 자동(셀 줄바꿈 고려) + 줄바꿈/상단정렬
        for ci, _ in enumerate(headers, start=1):
            letter = get_column_letter(ci)
            longest = max(
                [len(headers[ci - 1])]
                + [max((len(ln) for ln in str(r[ci - 1]).split("\n")), default=0) for r in rows],
                default=10,
            )
            ws.column_dimensions[letter].width = min(max(longest + 2, 10), 48)
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.alignment = wrap if "\n" in str(c.value or "") else top
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # Summary 시트 (기본 시트 재사용)
    summ = wb.active
    summ.title = "Summary"
    summ.append(["HTM Backoffice — EMPID Reconciliation"])
    summ["A1"].font = Font(bold=True, size=14)
    summ.append([])
    if org_name:
        summ.append(["Organization", org_name])
    if filename:
        summ.append(["Source file", filename])
    summ.append([])
    counts = result.counts()
    summ.append(["Bucket", "Count"])
    for c in summ[summ.max_row]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0075DE")
    bucket_meta = [
        ("Auto-assign", "auto", "1AAE39"),
        ("Multiple numbers", "multiple", "DD5B00"),
        ("Mismatch (DB vs file)", "mismatch", "C0392B"),
        ("Already assigned", "assigned", "615D59"),
        ("Placeholder emails", "placeholder", "615D59"),
        ("Deferred", "deferred", "615D59"),
    ]
    for label, key, _ in bucket_meta:
        summ.append([label, counts.get(key, 0)])
    summ.append(["Excluded (PURADAK)", counts.get("excluded_rows", 0)])
    summ.append(["Total rows", counts.get("total_rows", 0)])
    summ.column_dimensions["A"].width = 26
    summ.column_dimensions["B"].width = 12

    _sheet("Auto-assign", "1AAE39",
           ["Name", "Email", "emp_id"],
           [[p.user_full_name or p.name, p.email or "", p.emp_id or ""] for p in result.auto])
    _sheet("Multiple numbers", "DD5B00",
           ["Name", "Email", "emp_id options", "Sources (← COMPANY)"],
           [[p.user_full_name or p.name, p.email or "", _p_file_ids(p), _p_sources(p)] for p in result.multiple])
    _sheet("Mismatch", "C0392B",
           ["Name", "Email", "DB emp_id", "File says", "Sources (← COMPANY)", "Note"],
           [[p.user_full_name or p.name, p.email or "", p.db_emp_id or "", _p_file_ids(p), _p_sources(p), p.note]
            for p in result.mismatch])
    _sheet("Already assigned", "615D59",
           ["Name", "Email", "emp_id"],
           [[p.user_full_name or p.name, p.email or "", p.emp_id or ""] for p in result.assigned])
    _sheet("Placeholder", "615D59",
           ["Email", "People in file (name = number)", "DB account(s) using email", "Note"],
           [[p.email or "", _p_members(p) or (p.name or ""), _p_db_accounts(p), p.note]
            for p in result.placeholder])
    _sheet("Deferred", "615D59",
           ["Name (file)", "Email", "emp_id", "Similar DB users", "Note"],
           [[p.name or p.user_full_name, p.email or "", _p_file_ids(p), _p_similar(p), p.note]
            for p in result.deferred])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


@dataclass
class CommitResult:
    assigned: list[tuple[str, str]] = field(default_factory=list)  # (full_name, emp_id)
    skipped: list[tuple[str, str]] = field(default_factory=list)  # 이미 있어 skip
    rejected: list[tuple[str, str]] = field(default_factory=list)  # (user_id, reason)


async def commit_assignments(
    db: AsyncSession,
    organization_id: UUID,
    assignments: list[tuple[UUID, str]],
) -> CommitResult:
    """확정 — NULL인 대상에만 employee_no 기록 (멱등).

    - org 불일치 user → reject (IDOR 방지)
    - 이미 employee_no 있음 → skip (저널, 덮어쓰지 않음)
    - 포맷 검증(_normalize_employee_no) + org 이력(ledger) burn 위반 → reject (blind overwrite 금지)
    - 성공 시 employee_no 기록 + ledger 적재 (옵션 A 영구 burn). 같은 트랜잭션.
    """
    result = CommitResult()
    try:
        for user_id, raw_emp in assignments:
            user = await user_repository.get_by_id(db, user_id, organization_id)
            if user is None:
                result.rejected.append((str(user_id), "not found in org"))
                continue
            if user.employee_no:
                result.skipped.append((user.full_name, user.employee_no))
                continue
            try:
                emp = _normalize_employee_no(raw_emp)
            except ValueError as e:
                result.rejected.append((str(user_id), str(e)))
                continue
            if emp is None:
                result.rejected.append((str(user_id), "empty employee_no"))
                continue
            # 이력 기반 영구 burn 체크 — 과거 사용/현재 활성 모두 차단 (덮어쓰지 않음).
            burned = await employee_no_history_repository.exists_for_org(
                db, organization_id, emp
            )
            if burned:
                result.rejected.append(
                    (str(user_id), f"employee_no {emp} already used in org (previously assigned, cannot be reused)")
                )
                continue
            user.employee_no = emp
            # ledger 적재 — 성공 부여 시 burn 기록. user 존재(FK 충족).
            await employee_no_history_repository.add(
                db, organization_id, emp, user.id
            )
            result.assigned.append((user.full_name, emp))

        if result.assigned:
            await db.commit()
    except IntegrityError as e:
        # 동시 임포트 등으로 ledger unique 충돌 → 부분 저장 방지하고 깔끔한 409.
        # (배치 전체 롤백; 운영자에게 재시도 안내.)
        await db.rollback()
        raise DuplicateError(
            "employee_no assignment conflict (concurrent change) — please retry"
        ) from e
    return result
