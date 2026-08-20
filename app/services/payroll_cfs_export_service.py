"""Payroll CFS Export — 회계사(CFS) 급여 입력 파일 생성.

DUMMY 포맷 v1(payroll_export_service)을 대체하는 **실제 양식**이다.
컬럼 정의·규칙의 근거는 docs/99_inbox/2026-08-11-payroll-cfs-export-결정사항.md.

시트 1장 = store_group 1개 (§3-3). 같은 그룹의 매장들은 한 시트에 합쳐진다 —
회계사 파일이 그렇게 생겼고, empid 채번도 그룹 단위라 자연스럽게 맞는다.

컬럼 19개 (원본 A~U 중 S·T 제외, df2_*/cfs_name/JobToDo 제외):
    corp payroll_id no name emp_id rate performance_bonus total_comp
    rgl ovr dbl total tip_apply earnedtip performanceb check note cash "premium pay"

원본과 다르게 내는 것 (의도된 정정):
    - tip_apply 는 항상 0/1 플래그. 원본에서 일부 매장이 금액을 넣은 건 오류다.
    - 팁 금액 칸은 earnedtip 하나. 원본의 S열(total 4070 earnedtip)·T열 중복은 안 만든다.
    - premium pay 에 meal/rest 미준수 수당이 들어간다 (원본은 전부 공란이었다).

이름 표기:
    `FIRST MID LAST (USERNAME)` 전부 대문자. 구조화 이름이 없으면 full_name 으로
    폴백하되 Warnings 시트에 남긴다 — 회계사 쪽 매칭 키라 조용히 넘기면 안 된다.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO
from typing import Sequence
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.org_member import OrgMember, OrgMemberStore
from app.models.organization import Store, StoreGroup
from app.models.payroll import PayPeriod, PayrollEntry
from app.models.rate import HourlyRateHistory
from app.models.user import User
from app.schemas.payroll import EntryBreakdown, PayrollPreviewRow
from app.services.payroll_calc_service import (
    parse_frozen_breakdown,
    payroll_calc_service,
)
from app.services.tip_prorate_service import tip_prorate_service
from app.core.payroll_rules import payroll_period_label
from app.utils.download import safe_filename

CFS_COLUMNS: list[str] = [
    "corp",
    "payroll_id",
    "no",
    "name",
    "emp_id",
    "rate",
    "performance_bonus",
    "total_comp",
    "rgl",
    "ovr",
    "dbl",
    "total",
    "tip_apply",
    "earnedtip",
    "performanceb",
    "check",
    "note",
    "cash",
    "premium pay",
]
_COLUMN_WIDTHS = [26, 12, 5, 30, 8, 8, 10, 10, 8, 8, 8, 8, 9, 10, 12, 10, 18, 9, 12]

WARNINGS_SHEET_TITLE = "Warnings"
WARNINGS_COLUMNS: list[str] = ["Sheet", "Name", "Issue"]

DRAFT_BANNER = "DRAFT — period not confirmed; numbers may change"

# note 자동 태그 (§6) — 사람이 매번 손으로 적던 값을 규칙으로 대체한다.
NOTE_NEW_EMP = "new emp"
NOTE_RATE_CHANGE = "rate change"

_TWO_DP = Decimal("0.01")
_ZERO = Decimal("0.00")


def minutes_to_hours(minutes: int) -> Decimal:
    """분 → 시간 (소수 2자리) — 표기 전용. 계산은 언제나 분 단위다."""
    return (Decimal(minutes) / Decimal(60)).quantize(_TWO_DP, rounding=ROUND_HALF_UP)


def daily_rounded_hours(days: list, field: str) -> Decimal:
    """일자별로 반올림한 시간의 합 — 급여 파일에 나가는 시간의 산출 방식.

    분 합계를 마지막에 한 번 반올림하는 쪽이 수학적으로는 더 정확하지만,
    그러면 파일의 일자별 값을 더한 것과 총합이 1센트 단위로 어긋난다.
    받는 쪽(회계사)이 검산하면 안 맞는 것처럼 보이고, 차이가 나는 방향도
    직원에게 불리할 수 있다. 그래서 **보수적으로** 일자별 반올림값을 더한다
    (검산이 항상 맞고, 반올림이 직원에게 불리하게만 작동하지 않는다).

    Args:
        days: EntryBreakdown.days
        field: "regular_minutes" | "ot_minutes" | "dt_minutes"
    """
    total = Decimal("0.00")
    for day in days:
        total += minutes_to_hours(int(getattr(day, field, 0) or 0))
    return total


def payroll_id_for(start_date: date) -> str:
    """기간 시작일 → `YYYY.MM.FH|LH`. 1–15 = FH, 16–말일 = LH.

    규칙 본체는 `app/core/payroll_rules.payroll_period_label` — 파일명과 시트 값이
    갈리지 않도록 한 곳에서만 만든다.
    """
    return payroll_period_label(start_date)


def cfs_name(user: User) -> tuple[str, bool]:
    """`FIRST MID LAST (USERNAME)` 대문자. (표기, 구조화이름_사용여부) 반환.

    구조화 이름이 없으면 full_name 으로 폴백한다. 폴백은 회계사 매칭이 어긋날 수
    있는 상태라 호출측이 Warnings 에 남긴다.
    """
    parts = [p for p in (user.first_name, user.middle_name, user.last_name) if p]
    structured = bool(parts)
    base = " ".join(parts) if structured else (user.full_name or "")
    username = user.username or ""
    label = f"{base} ({username})" if username else base
    return label.upper(), structured


@dataclass
class CfsRow:
    """시트 한 줄. 컬럼 순서와 1:1 대응한다."""

    corp: str
    payroll_id: str
    name: str
    emp_id: int | None
    rate: Decimal
    performance_bonus: Decimal
    rgl: Decimal
    ovr: Decimal
    dbl: Decimal
    tip_apply: int
    earnedtip: Decimal
    performanceb: Decimal
    check: str = ""
    note: str = ""
    cash: Decimal = _ZERO
    premium_pay: Decimal = _ZERO
    # 정렬·병합·경고용 (셀로 나가지 않는다)
    user_id: str = field(default="", compare=False)
    sort_key: tuple = field(default=(), compare=False)
    warnings: list[str] = field(default_factory=list, compare=False)

    @property
    def total_comp(self) -> Decimal:
        return (self.rate + self.performance_bonus).quantize(_TWO_DP)

    @property
    def total(self) -> Decimal:
        return self.rgl + self.ovr + self.dbl

    def cells(self, no: int) -> list:
        """`no` 는 시트 안에서 부여되는 순번이라 밖에서 받는다."""
        return [
            self.corp,
            self.payroll_id,
            no,
            self.name,
            self.emp_id,
            self.rate,
            self.performance_bonus,
            self.total_comp,
            self.rgl,
            self.ovr,
            self.dbl,
            self.total,
            self.tip_apply,
            self.earnedtip,
            self.performanceb,
            self.check,
            self.note,
            self.cash,
            self.premium_pay,
        ]


def merge_by_person(rows: list[CfsRow]) -> list[CfsRow]:
    """**레거시(전환 전 store 스코프 기간) 전용** — 매장별 행을 사람 기준 병합.

    회계사 파일은 그룹당 1인 1행이다. 매장별로 따로 내보내면 같은 사람이 두 줄이
    되고, 회계사 쪽에서 별개 직원으로 잡혀 급여가 두 번 나간다.

    ⚠️ 이 병합은 시간·금액을 합산할 뿐 OT 를 재분류하지 못한다 (분류는 이미
    매장별로 끝난 값). group 스코프 기간(2026-08-19 전환 이후)은 계산 엔진이
    법인 합산으로 1인 1행을 직접 만들므로 이 함수를 지나지 않는다 — 전환 전
    확정 원장을 당시 모습대로 재현하는 경로에만 남아 있다.

    empid 는 가장 작은 값을 쓴다 — 그룹 공유 채번이라 사람마다 번호가 하나여야
    맞지만, 매장별로 따로 채번된 과거 데이터가 있으면 둘 이상이 될 수 있다.
    이때 큰 쪽(나중에 붙은 번호)이 아니라 처음 받은 번호를 쓰는 게 안전하다.
    """
    merged: dict[str, CfsRow] = {}
    order: list[str] = []
    for row in rows:
        key = row.user_id or f"__{row.emp_id}"
        if key not in merged:
            merged[key] = row
            order.append(key)
            continue
        base = merged[key]
        base.rgl += row.rgl
        base.ovr += row.ovr
        base.dbl += row.dbl
        base.earnedtip += row.earnedtip
        base.performanceb += row.performanceb
        base.cash += row.cash
        base.premium_pay += row.premium_pay
        base.tip_apply = max(base.tip_apply, row.tip_apply)
        if base.emp_id is None or (row.emp_id is not None and row.emp_id < base.emp_id):
            base.emp_id = row.emp_id
        for note in (row.note or "").split(", "):
            if note and note not in base.note:
                base.note = f"{base.note}, {note}" if base.note else note
        base.warnings.extend(w for w in row.warnings if w not in base.warnings)
        base.sort_key = (base.emp_id is None, base.emp_id or 0, base.name)
    return [merged[key] for key in order]


@dataclass
class CfsSheet:
    """시트 1장 = store_group 1개."""

    title: str
    rows: list[CfsRow]
    is_draft: bool = False


def _style_headers(ws: Worksheet, headers: list[str], widths: list[int], row: int = 1) -> None:
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2D3436", end_color="2D3436", fill_type="solid")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=row, column=col_idx).column_letter].width = width


def _add_draft_banner(ws: Worksheet, column_count: int) -> None:
    cell = ws.cell(row=1, column=1, value=DRAFT_BANNER)
    cell.font = Font(bold=True, color="9C2A2A", size=12)
    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    ws.row_dimensions[1].height = 22


def build_workbook(sheets: Sequence[CfsSheet]) -> Workbook:
    """시트 목록 → 워크북 (순수 함수 — DB 없음).

    `no` 는 여기서 부여한다. 저장하지 않는 값이고, 매장 등록 순서(= empid 오름차순)를
    따른다 (§2 C열). 회계사 쪽 식별 키가 아니라 눈으로 세기 위한 번호다.
    """
    wb = Workbook()
    wb.remove(wb.active)
    warnings: list[list[str]] = []

    for sheet in sheets:
        ws = wb.create_sheet(sheet.title[:31])  # xlsx 시트명 상한
        header_row = 1
        if sheet.is_draft:
            _add_draft_banner(ws, len(CFS_COLUMNS))
            header_row = 2
        _style_headers(ws, CFS_COLUMNS, _COLUMN_WIDTHS, row=header_row)
        for no, row in enumerate(sorted(sheet.rows, key=lambda r: r.sort_key), start=1):
            ws.append(row.cells(no))
            for issue in row.warnings:
                warnings.append([sheet.title, row.name, issue])

    if warnings:
        warn_ws = wb.create_sheet(WARNINGS_SHEET_TITLE)
        _style_headers(warn_ws, WARNINGS_COLUMNS, [16, 30, 80])
        for warning in warnings:
            warn_ws.append(warning)
    return wb


def workbook_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_filename(
    label: str,
    start_date: date,
    end_date: date,
    *,
    draft: bool = False,
    generated_at: datetime | None = None,
) -> str:
    """`CFS_{조직}_{기간ID}_{날짜범위}_{상태}_{생성시각}.xlsx`.

    예: `CFS_ALL_2026.08.FH_20260801-0815_DRAFT_20260820-1352Z.xlsx`
    기간 export 와 같은 규칙 (payroll_export_service._payroll_filename 참조).
    """
    from app.services.payroll_export_service import _payroll_filename

    return _payroll_filename(
        "CFS", label, start_date, end_date,
        draft=draft, generated_at=generated_at, ext="xlsx",
    )


class PayrollCfsExportService:
    """기간 → CFS 양식 워크북. 그룹 단위로 시트를 만든다."""

    async def build_org_export(
        self,
        db: AsyncSession,
        *,
        organization_id: UUID,
        start_date: date,
        end_date: date,
    ) -> tuple[Workbook, bool]:
        """조직 전체 → 그룹별 시트 워크북. (워크북, draft 포함 여부) 반환."""
        groups = (
            await db.scalars(
                select(StoreGroup)
                .where(StoreGroup.organization_id == organization_id)
                .order_by(StoreGroup.sort_order.asc(), StoreGroup.name.asc())
            )
        ).all()

        sheets: list[CfsSheet] = []
        any_draft = False
        for group in groups:
            sheet = await self.build_group_sheet(
                db, group=group, start_date=start_date, end_date=end_date
            )
            if sheet is None:
                continue
            any_draft = any_draft or sheet.is_draft
            sheets.append(sheet)
        return build_workbook(sheets), any_draft

    async def build_group_sheet(
        self,
        db: AsyncSession,
        *,
        group: StoreGroup,
        start_date: date,
        end_date: date,
    ) -> CfsSheet | None:
        """그룹 1개 → 시트 1장. 대상 기간에 행이 하나도 없으면 None."""
        stores = (
            await db.scalars(
                select(Store)
                .where(Store.group_id == group.id)
                .order_by(Store.sort_order.asc(), Store.name.asc())
            )
        ).all()
        if not stores:
            return None

        corp = group.payroll_corp_name or group.name
        title = group.code or group.name

        # group 스코프 기간 (2026-08-19 전환 이후) — 계산이 이미 법인 합산이라
        # 사람당 1행이 그대로 나온다 (merge_by_person 불필요).
        group_period = await db.scalar(
            select(PayPeriod).where(
                PayPeriod.store_group_id == group.id,
                PayPeriod.start_date == start_date,
                PayPeriod.end_date == end_date,
            )
        )
        if group_period is not None:
            rows, is_draft = await self._rows_for_period(
                db, stores=list(stores), period=group_period, corp=corp
            )
            if not rows:
                return None
            return CfsSheet(title=title, rows=rows, is_draft=is_draft)

        # 레거시(전환 전) — 매장별 확정 원장을 당시 모습대로 모아 사람 기준 병합
        rows = []
        is_draft = False
        for store in stores:
            period = await db.scalar(
                select(PayPeriod).where(
                    PayPeriod.store_id == store.id,
                    PayPeriod.start_date == start_date,
                    PayPeriod.end_date == end_date,
                )
            )
            if period is None:
                continue
            store_rows, draft = await self._rows_for_period(
                db, stores=[store], period=period, corp=corp
            )
            is_draft = is_draft or draft
            rows.extend(store_rows)

        if not rows:
            return None
        return CfsSheet(title=title, rows=merge_by_person(rows), is_draft=is_draft)

    # ── 내부 ────────────────────────────────────────────────────────

    async def _rows_for_period(
        self,
        db: AsyncSession,
        *,
        stores: list[Store],
        period: PayPeriod,
        corp: str,
    ) -> tuple[list[CfsRow], bool]:
        """기간 1개 → 행 목록. confirmed 는 동결본, open 은 live preview.

        stores = 기간의 스코프 매장들 (group 기간이면 그룹 전체, 레거시면 1곳)
        — 팁/현금 집계에 쓴다.
        """
        draft = period.status != "confirmed"
        payroll_id = payroll_id_for(period.start_date)

        if draft:
            sources: list = await payroll_calc_service.preview_period(db, period)
            breakdowns = [row.breakdown for row in sources]
        else:
            entries = (
                await db.scalars(
                    select(PayrollEntry)
                    .where(PayrollEntry.pay_period_id == period.id)
                    .order_by(PayrollEntry.member_name.asc())
                )
            ).all()
            sources = list(entries)
            breakdowns = [parse_frozen_breakdown(e.breakdown) for e in entries]

        if not sources:
            return [], draft

        user_ids = {s.user_id for s in sources}
        users = {
            u.id: u
            for u in (await db.scalars(select(User).where(User.id.in_(user_ids)))).all()
        }
        store_ids = [s.id for s in stores]
        eligibility = await self._tip_eligibility(db, store_ids, user_ids)
        cash_map = await self._cash_map(db, store_ids=store_ids, period=period)
        first_period = await self._first_period_users(db, store_ids, period, user_ids)
        rate_changed = await self._rate_changed_users(db, period, user_ids)

        rows: list[CfsRow] = []
        for source, breakdown in zip(sources, breakdowns):
            user = users.get(source.user_id)
            if user is None:
                continue  # 계정 유실 — 표기할 이름이 없다
            name, structured = cfs_name(user)
            warnings: list[str] = []
            if not structured:
                warnings.append(
                    "Name is not split into first/last — export falls back to the "
                    "full name, which may not match the accountant's records"
                )
            if source.empid is None:
                warnings.append(
                    "No EMPID on file — assign one so future exports line up "
                    "automatically"
                )

            notes: list[str] = []
            if source.user_id in first_period:
                notes.append(NOTE_NEW_EMP)
            if source.user_id in rate_changed:
                notes.append(NOTE_RATE_CHANGE)

            days = breakdown.days or []
            rate = self._display_rate(breakdown)
            bonus_rate = breakdown.bonus_rate or _ZERO
            cash = cash_map.get(source.user_id, _ZERO)
            rows.append(
                CfsRow(
                    corp=corp,
                    payroll_id=payroll_id,
                    name=name,
                    emp_id=source.empid,
                    rate=rate,
                    performance_bonus=bonus_rate,
                    # 일자별 반올림 합 (검산 가능한 값). days 가 비면 총분으로 폴백.
                    rgl=(
                        daily_rounded_hours(days, "regular_minutes")
                        if days
                        else minutes_to_hours(source.regular_minutes)
                    ),
                    ovr=(
                        daily_rounded_hours(days, "ot_minutes")
                        if days
                        else minutes_to_hours(source.ot_minutes)
                    ),
                    dbl=(
                        daily_rounded_hours(days, "dt_minutes")
                        if days
                        else minutes_to_hours(source.dt_minutes)
                    ),
                    tip_apply=1 if eligibility.get(source.user_id) else 0,
                    # 회계사 양식의 earnedtip 은 카드+현금 합계다 (실데이터로 확인).
                    earnedtip=(source.card_tips + cash).quantize(_TWO_DP),
                    performanceb=source.bonus_pay,
                    note=", ".join(notes),
                    cash=cash,
                    premium_pay=source.penalty_pay,
                    # 등록 순서 = empid 오름차순. empid 없는 사람은 뒤로.
                    user_id=str(source.user_id),
                    sort_key=(source.empid is None, source.empid or 0, name),
                    warnings=warnings,
                )
            )
        return rows, draft

    @staticmethod
    def _display_rate(breakdown: EntryBreakdown) -> Decimal:
        """rate 칸 한 칸 — 기간 내 단일 rate 가 정책이라 보통 구간이 하나다.

        그래도 구간이 여러 개면(정책 위반 데이터/과거분) 가장 큰 rate 를 쓴다.
        조용히 낮은 값을 보여주는 것보다 눈에 띄는 쪽이 낫다.
        """
        if not breakdown.segments:
            return _ZERO
        return max(s.rate for s in breakdown.segments).quantize(_TWO_DP)

    async def _tip_eligibility(
        self, db: AsyncSession, store_ids: list[UUID], user_ids: set[UUID]
    ) -> dict[UUID, bool]:
        """직원별 tip_apply — 팁 운영 매장 중 한 곳이라도 대상이면 1.

        매장이 팁 미운영(mode=none)이면 그 매장 배정은 계산에서 빠진다
        (설정이 상위 게이트).
        """
        eligible: dict[UUID, bool] = {}
        for store_id in store_ids:
            if await tip_prorate_service.store_mode(db, store_id) == "none":
                continue
            rows = await db.execute(
                select(OrgMember.user_id, OrgMemberStore.tip_eligible)
                .join(OrgMemberStore, OrgMemberStore.org_member_id == OrgMember.id)
                .where(
                    OrgMemberStore.store_id == store_id,
                    OrgMember.user_id.in_(user_ids),
                )
            )
            for r in rows:
                eligible[r.user_id] = eligible.get(r.user_id, False) or bool(
                    r.tip_eligible
                )
        return eligible

    async def _cash_map(
        self, db: AsyncSession, *, store_ids: list[UUID], period: PayPeriod
    ) -> dict[UUID, Decimal]:
        """현금팁 — 시간비례 매장은 배분된 몫, 그 외는 본인 입력값. 매장 합산."""
        from app.models.tip import TipEntry

        cash: dict[UUID, Decimal] = {}
        for store_id in store_ids:
            if await tip_prorate_service.store_mode(db, store_id) == "hours_prorated":
                totals = await tip_prorate_service.totals_for_period(
                    db, store_id=store_id,
                    start=period.start_date, end=period.end_date,
                )
                for user_id, t in totals.items():
                    cash[user_id] = cash.get(user_id, _ZERO) + t["cash"]
                continue

            entries = (
                await db.scalars(
                    select(TipEntry).where(
                        TipEntry.store_id == store_id,
                        TipEntry.date >= period.start_date,
                        TipEntry.date <= period.end_date,
                    )
                )
            ).all()
            for entry in entries:
                cash[entry.employee_id] = (
                    cash.get(entry.employee_id, _ZERO) + entry.cash_tips_kept
                )
        return cash

    async def _first_period_users(
        self,
        db: AsyncSession,
        store_ids: list[UUID],
        period: PayPeriod,
        user_ids: set[UUID],
    ) -> set[UUID]:
        """이 기간이 첫 급여기간인 직원 — note 의 `new emp` 자동 판정.

        직전 기간들(같은 그룹의 group 기간 + 스코프 매장들의 레거시 기간)에
        확정 entry 가 없으면 신규로 본다.
        """
        if not user_ids:
            return set()
        scope_conditions = [PayPeriod.store_id.in_(store_ids)]
        if period.store_group_id is not None:
            scope_conditions.append(
                PayPeriod.store_group_id == period.store_group_id
            )
        from sqlalchemy import or_

        prior = await db.execute(
            select(PayrollEntry.user_id)
            .join(PayPeriod, PayPeriod.id == PayrollEntry.pay_period_id)
            .where(
                or_(*scope_conditions),
                PayPeriod.start_date < period.start_date,
                PayrollEntry.user_id.in_(user_ids),
            )
            .distinct()
        )
        return user_ids - {r.user_id for r in prior}

    async def _rate_changed_users(
        self, db: AsyncSession, period: PayPeriod, user_ids: set[UUID]
    ) -> set[UUID]:
        """이번 기간에 시급이 바뀐 직원 — note 의 `rate change` 자동 판정.

        효력일이 기간 시작일인 이력이 있으면(= 이번 기간부터 새 rate) 표시한다.
        """
        if not user_ids:
            return set()
        rows = await db.execute(
            select(OrgMember.user_id)
            .join(HourlyRateHistory, HourlyRateHistory.org_member_id == OrgMember.id)
            .where(
                OrgMember.user_id.in_(user_ids),
                HourlyRateHistory.effective_date >= period.start_date,
                HourlyRateHistory.effective_date <= period.end_date,
                HourlyRateHistory.old_rate.is_not(None),
            )
            .distinct()
        )
        return {r.user_id for r in rows}


payroll_cfs_export_service = PayrollCfsExportService()
