"""Staff 데이터 xlsx export 서비스.

콘솔 Staff 목록의 사람×매장배정 데이터를 Excel 워크북으로 만든다.
행 = 사람 × 매장배정 (org_member_stores 기준). 배정이 없는 사람은 매장 칸이 빈 1행.

정책:
    - **시급/급여(cost) 컬럼은 아예 넣지 않는다** — cost 마스킹 정책.
      권한 분기로 가리는 게 아니라 export 자체에서 전면 제외한다. 컬럼 추가 시 주의.
    - 매장 배정 소스는 org_member_stores (Model B). 휴면 행(배정 해제 시
      is_work_assignment=false·is_manager=false 로 empid 만 보존하는 정책 A 행)은
      "현재 배정"이 아니므로 제외한다.
    - 미가입(유령, users.is_provisional) 계정은 Account Status 를 "Not signed up" 으로.
"""

from __future__ import annotations

from io import BytesIO
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import and_, or_, select
from sqlalchemy.engine import Row
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.org_member import OrgMember, OrgMemberStore
from app.models.organization import Store
from app.models.user import Role, User

# 컬럼 순서 고정 — 시급/급여 컬럼 금지 (모듈 docstring 참조)
_HEADERS: list[str] = [
    "Username",
    "Name",
    "CREWID",
    "Role",
    "Account Status",
    "Employment Status",
    "Department",
    "Email",
    "Hire Date",
    "Termination Date",
    "Store",
    "Store EMPID",
    "Manager",
]

_COLUMN_WIDTHS: list[int] = [18, 22, 9, 18, 14, 17, 12, 26, 12, 16, 24, 11, 9]


class UserExportService:
    """Staff xlsx export — 조회 + 워크북 생성."""

    async def export_staff_xlsx(
        self,
        db: AsyncSession,
        organization_id: UUID,
        store_ids: list[UUID] | None = None,
    ) -> bytes:
        """조직의 Staff 목록을 xlsx 바이트로 만든다.

        Args:
            db: 비동기 데이터베이스 세션
            organization_id: 조직 ID (호출자 JWT 의 org — 스코프 필수)
            store_ids: 매장 필터. 주어지면 그 매장(들)에 배정된 행만 —
                사람 자체도 배정 기준으로 포함/제외된다 (배정 없는 사람 제외).

        Returns:
            bytes: xlsx 파일 바이트
        """
        rows = await self._fetch_rows(db, organization_id, store_ids)
        return self._build_workbook(rows)

    async def _fetch_rows(
        self,
        db: AsyncSession,
        organization_id: UUID,
        store_ids: list[UUID] | None,
    ) -> list[Row]:
        """사람 × 매장배정 행 조회.

        users 를 기준으로 org_members / org_member_stores 를 LEFT JOIN 한다 —
        전환기(Model B)라 org_member 미생성 레거시 계정도 있을 수 있고, 그 경우
        CREWID/Employment 칸이 비어도 사람 행은 나와야 한다.
        """
        # "현재 배정" 조건 — 휴면 행(둘 다 false, empid 보존용) 제외
        assignment_on = and_(
            OrgMemberStore.org_member_id == OrgMember.id,
            or_(
                OrgMemberStore.is_work_assignment.is_(True),
                OrgMemberStore.is_manager.is_(True),
            ),
        )

        query = (
            select(
                User.username,
                User.full_name,
                User.email,
                User.department,
                User.is_active,
                User.is_provisional,
                Role.name.label("role_name"),
                OrgMember.crewid,
                OrgMember.status.label("member_status"),
                OrgMember.hire_date,
                OrgMember.termination_date,
                Store.name.label("store_name"),
                OrgMemberStore.empid,
                OrgMemberStore.is_manager,
            )
            .join(Role, Role.id == User.role_id)
            .outerjoin(
                OrgMember,
                and_(
                    OrgMember.user_id == User.id,
                    OrgMember.organization_id == organization_id,
                ),
            )
        )

        if store_ids:
            # 매장 필터 — INNER JOIN 으로 그 매장 배정 행만 남긴다.
            # (org_member 없는 레거시 계정은 배정 행이 없으므로 자연히 제외.)
            # store_ids 는 이 org 의 org_member 를 거쳐야만 매칭되므로 타 org
            # 매장 id 를 넣어도 아무것도 안 나온다 (cross-org 누출 없음).
            query = query.join(
                OrgMemberStore,
                and_(assignment_on, OrgMemberStore.store_id.in_(store_ids)),
            ).join(Store, Store.id == OrgMemberStore.store_id)
        else:
            query = query.outerjoin(OrgMemberStore, assignment_on).outerjoin(
                Store, Store.id == OrgMemberStore.store_id
            )

        query = query.where(User.organization_id == organization_id).order_by(
            User.full_name, User.username, Store.name
        )
        result = await db.execute(query)
        return list(result.all())

    def _build_workbook(self, rows: list[Row]) -> bytes:
        """조회 행 → 스타일된 단일 시트 워크북 (dashboard export 와 같은 룩)."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Staff"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="2D3436", end_color="2D3436", fill_type="solid"
        )
        for col_idx, header in enumerate(_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            has_store: bool = row.store_name is not None
            ws.append(
                [
                    row.username,
                    row.full_name,
                    row.crewid,
                    row.role_name,
                    self._account_status(row),
                    row.member_status or "",
                    row.department or "",
                    row.email or "",
                    str(row.hire_date) if row.hire_date else "",
                    str(row.termination_date) if row.termination_date else "",
                    row.store_name or "",
                    row.empid if has_store else "",
                    ("Y" if row.is_manager else "N") if has_store else "",
                ]
            )

        for i, width in enumerate(_COLUMN_WIDTHS, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _account_status(row: Row) -> str:
        """계정 상태 라벨 — 유령(미가입)은 is_active=False 라도 별도 라벨."""
        if row.is_provisional:
            return "Not signed up"
        return "Active" if row.is_active else "Inactive"


user_export_service: UserExportService = UserExportService()
